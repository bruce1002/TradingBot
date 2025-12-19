"""
FastAPI æ‡‰ç”¨ç¨‹å¼å…¥å£é»

æä¾› REST API æ¥æ”¶ TradingView webhookï¼Œä¸¦åœ¨å¹£å®‰æœŸè²¨æ¸¬è©¦ç¶²ä¸‹å–®ã€‚
æ‰€æœ‰è¨‚å–®éƒ½æœƒè¨˜éŒ„åˆ°è³‡æ–™åº«ä¸­ï¼Œæ–¹ä¾¿è¿½è¹¤å’ŒæŸ¥è©¢ã€‚
"""

from fastapi import FastAPI, HTTPException, Depends, Header, Request, Query, Request, status
from fastapi.responses import JSONResponse, RedirectResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import delete, and_, or_
from sqlalchemy.sql import func
from pydantic import BaseModel, Field
from typing import Optional, List
from dataclasses import dataclass
from copy import copy
import os
import asyncio
import logging
from logging.handlers import RotatingFileHandler
import time
import io
import json
import hashlib
from datetime import datetime, timezone, date, timedelta
from fastapi.responses import StreamingResponse
from authlib.integrations.starlette_client import OAuth
from dotenv import load_dotenv

# è¼‰å…¥ .env æª”æ¡ˆä¸­çš„ç’°å¢ƒè®Šæ•¸ï¼ˆå¿…é ˆåœ¨å…¶ä»–æ¨¡çµ„å°å…¥ä¹‹å‰ï¼‰
load_dotenv()

from db import init_db, get_db, SessionLocal
from models import Position, TradingViewSignalLog, BotConfig, TVSignalConfig, PortfolioTrailingConfig
from binance_client import (
    get_client, 
    get_mark_price, 
    open_futures_market_order, 
    close_futures_position,
    get_symbol_info,
    update_all_bots_invest_amount,
    format_quantity
)

# è¨­å®šæ—¥èªŒ
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
LOG_FILE = os.getenv("LOG_FILE", "tvbot.log")

logger = logging.getLogger("tvbot")
logger.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))

# é¿å…é‡è¤‡ handler
if not logger.handlers:
    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )

    # console handler
    ch = logging.StreamHandler()
    ch.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    # rotating file handler
    fh = RotatingFileHandler(
        LOG_FILE,
        maxBytes=5 * 1024 * 1024,  # 5MB
        backupCount=5,             # ä¿ç•™ 5 å€‹èˆŠæª”æ¡ˆ
        encoding="utf-8",
    )
    fh.setLevel(getattr(logging, LOG_LEVEL, logging.INFO))
    fh.setFormatter(formatter)
    logger.addHandler(fh)

# ==================== Dynamic Stop è¨­å®š ====================
# å¾ç’°å¢ƒè®Šæ•¸è®€å– dynamic stop è¨­å®šï¼Œä¸¦æä¾›åˆç†é è¨­å€¼
# TODO: ä¹‹å¾Œæœƒçµ±ä¸€æ”¹ç”¨ TRAILING_CONFIGï¼Œé€™äº›è®Šæ•¸ä¿ç•™ä½œç‚ºå‘å¾Œå…¼å®¹
DYN_TRAILING_ENABLED = os.getenv("DYN_TRAILING_ENABLED", "1") == "1"
DYN_PROFIT_THRESHOLD_PCT = float(os.getenv("DYN_PROFIT_THRESHOLD_PCT", "1.0"))   # PnL% >= 1% æ‰å•Ÿå‹•é–åˆ©
DYN_LOCK_RATIO_DEFAULT = float(os.getenv("DYN_LOCK_RATIO_DEFAULT", "0.5"))       # é è¨­é–ä¸€åŠçš„åˆ©æ½¤ (0.5)
DYN_BASE_SL_PCT = float(os.getenv("DYN_BASE_SL_PCT", "3.0"))                     # é‚„æ²’é”é–€æª»å‰ï¼Œbase åœæ -3%

# ==================== é bot å‰µå»ºçš„ position æ­·å²æœ€é«˜åƒ¹æ ¼è¿½è¹¤ ====================
# ç”¨æ–¼è¿½è¹¤é bot å‰µå»ºçš„ position çš„æ­·å²æœ€é«˜/æœ€ä½åƒ¹æ ¼
# key: f"{symbol}|{position_side}" (ä¾‹å¦‚: "BNBUSDT|LONG")
# value: {"highest_price": float, "entry_price": float, "side": str}
# æ³¨æ„ï¼šé€™å€‹æ˜ å°„åªå­˜åœ¨æ–¼è¨˜æ†¶é«”ä¸­ï¼Œæ‡‰ç”¨é‡å•Ÿå¾Œæœƒé‡ç½®
_non_bot_position_tracking: dict[str, dict] = {}

# ==================== Binance Live Positions åœæé…ç½®è¦†å¯« ====================
# ç”¨æ–¼å­˜å„² Binance Live Positions çš„åœæé…ç½®è¦†å¯«å€¼
# key: f"{symbol}|{position_side}" (ä¾‹å¦‚: "BNBUSDT|LONG")
# value: {"dyn_profit_threshold_pct": float | None, "base_stop_loss_pct": float | None, "trail_callback": float | None}
# æ³¨æ„ï¼šé€™å€‹æ˜ å°„åªå­˜åœ¨æ–¼è¨˜æ†¶é«”ä¸­ï¼Œæ‡‰ç”¨é‡å•Ÿå¾Œæœƒé‡ç½®
_binance_position_stop_overrides: dict[str, dict] = {}

# ==================== Binance Portfolio Trailing Stop ç‹€æ…‹ ====================
# ç”¨æ–¼å­˜å„² Portfolio-level trailing stop çš„é‹è¡Œæ™‚ç‹€æ…‹
# æŒä¹…åŒ–é…ç½®ï¼ˆenabled, target_pnl, lock_ratioï¼‰å­˜å„²åœ¨è³‡æ–™åº«ä¸­ï¼ˆid=1 for LONG, id=2 for SHORTï¼‰
# é‹è¡Œæ™‚ç‹€æ…‹ï¼ˆmax_pnl_reached, last_check_timeï¼‰åªå­˜åœ¨æ–¼è¨˜æ†¶é«”ä¸­ï¼Œæ‡‰ç”¨é‡å•Ÿå¾Œæœƒé‡ç½®
_portfolio_trailing_runtime_state: dict = {
    "long": {
        "max_pnl_reached": None,
        "last_check_time": None
    },
    "short": {
        "max_pnl_reached": None,
        "last_check_time": None
    }
}

# ==================== é¢¨æ§è¨­å®š ====================
# å…è¨±äº¤æ˜“çš„äº¤æ˜“å°åˆ—è¡¨
ALLOWED_SYMBOLS = {"BTCUSDT", "ETHUSDT"}

# æ¯å€‹äº¤æ˜“å°çš„æœ€å¤§æ æ¡¿å€æ•¸
MAX_LEVERAGE_PER_SYMBOL = {
    "BTCUSDT": 20,
    "ETHUSDT": 10
}

# æ¯å€‹äº¤æ˜“å°çš„æœ€å¤§äº¤æ˜“æ•¸é‡
MAX_QTY_PER_SYMBOL = {
    "BTCUSDT": 0.05,
    "ETHUSDT": 1
}

# åˆå§‹åŒ– FastAPI æ‡‰ç”¨ç¨‹å¼
app = FastAPI(
    title="TradingView Binance Bot",
    description="æ¥æ”¶ TradingView webhook ä¸¦åœ¨å¹£å®‰æœŸè²¨æ¸¬è©¦ç¶²ä¸‹å–®",
    version="1.0.0"
)

# è¨­å®š Session Middlewareï¼ˆç”¨æ–¼ Google OAuthï¼‰
# å¾ç’°å¢ƒè®Šæ•¸è®€å– secret keyï¼Œç”¨æ–¼ç°½ç½² session cookie
SESSION_SECRET_KEY = os.getenv("SESSION_SECRET_KEY", "your-secret-key-change-in-production")
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET_KEY)

# åˆå§‹åŒ– Jinja2 Templates
templates = Jinja2Templates(directory="templates")

# æ›è¼‰éœæ…‹æª”æ¡ˆç›®éŒ„ï¼ˆCSSã€JSï¼‰
# æ³¨æ„ï¼šå¿…é ˆåœ¨ static ç›®éŒ„å­˜åœ¨çš„æƒ…æ³ä¸‹æ‰èƒ½æ›è¼‰
static_dir = "static"
if os.path.exists(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")
    logger.info(f"éœæ…‹æª”æ¡ˆç›®éŒ„å·²æ›è¼‰: /static -> {static_dir}")
else:
    logger.warning(f"éœæ…‹æª”æ¡ˆç›®éŒ„ä¸å­˜åœ¨: {static_dir}ã€‚è«‹ç¢ºä¿ static ç›®éŒ„å­˜åœ¨ã€‚")

# è¨­å®š Google OAuth
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")
ADMIN_GOOGLE_EMAIL = os.getenv("ADMIN_GOOGLE_EMAIL", "")

# åˆ¤æ–·æ˜¯å¦å•Ÿç”¨ Google OAuthï¼ˆå¿…é ˆä¸‰å€‹ç’°å¢ƒè®Šæ•¸éƒ½æœ‰å€¼æ‰å•Ÿç”¨ï¼‰
GOOGLE_OAUTH_ENABLED = bool(
    GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET and ADMIN_GOOGLE_EMAIL
)

# OAuth è¨­å®š
oauth = OAuth(app)
if GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET:
    oauth.register(
        name="google",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
        client_kwargs={
            "scope": "openid email profile"
        }
    )

# è¨˜éŒ„ Google OAuth ç‹€æ…‹
if GOOGLE_OAUTH_ENABLED:
    logger.info(f"Google OAuth å·²å•Ÿç”¨ï¼Œç®¡ç†å“¡ email = {ADMIN_GOOGLE_EMAIL}")
else:
    logger.warning("Google OAuth æœªè¨­å®šï¼Œå•Ÿç”¨ã€é–‹ç™¼æ¨¡å¼ç„¡ç™»å…¥ã€ï¼Œæ‰€æœ‰ä½¿ç”¨è€…è¦–ç‚ºç®¡ç†å“¡ï¼ˆå‹¿ç”¨æ–¼æ­£å¼ç’°å¢ƒï¼‰")

# èƒŒæ™¯ä»»å‹™é‹è¡Œæ¨™èªŒ
_trailing_worker_running = False


async def trailing_stop_worker():
    """
    è¿½è¹¤åœæèƒŒæ™¯ä»»å‹™
    
    æ¯ 5 ç§’æª¢æŸ¥ä¸€æ¬¡æ‰€æœ‰ OPEN ç‹€æ…‹çš„å€‰ä½ï¼ˆåŒ…æ‹¬ bot å‰µå»ºå’Œé bot å‰µå»ºçš„å€‰ä½ï¼‰ï¼Œ
    æ ¹æ“š dynamic stop é‚è¼¯æ±ºå®šæ˜¯å¦éœ€è¦é—œå€‰ã€‚
    """
    global _trailing_worker_running
    
    logger.info("è¿½è¹¤åœæèƒŒæ™¯ä»»å‹™å·²å•Ÿå‹•")
    _trailing_worker_running = True
    
    while _trailing_worker_running:
        # æ¯ä¸€è¼ªå»ºç«‹ä¸€å€‹æ–°çš„ DB session
        db = SessionLocal()
        try:
            # å¾è³‡æ–™åº«æ‰¾å‡ºæ‰€æœ‰éœ€è¦æª¢æŸ¥çš„å€‰ä½
            # åªæª¢æŸ¥ bot_stop_loss_enabled=True çš„å€‰ä½ï¼ˆå¦‚æœç‚º Falseï¼Œå‰‡è·³é Bot å…§å»ºçš„åœææ©Ÿåˆ¶ï¼‰
            # åªè¦æ˜¯ status == "OPEN" çš„å€‰ä½ï¼Œå°±è‡³å°‘è¦åƒ base stopï¼ˆå³ä½¿æ²’æœ‰è¨­å®š trail_callbackï¼‰
            # Dynamic Stop æ˜¯å¦å•Ÿç”¨ç”± DYN_TRAILING_ENABLED å’Œ lock_ratio ä¾†æ±ºå®š
            positions = (
                db.query(Position)
                .filter(Position.status == "OPEN")
                .filter(Position.bot_stop_loss_enabled == True)
                .all()
            )
            
            if positions:
                logger.info(f"æª¢æŸ¥ {len(positions)} å€‹é–‹å•Ÿçš„å€‰ä½ï¼ˆbot_stop_loss_enabled=Trueï¼‰")
                # è¨˜éŒ„æ²’æœ‰ lock_ratio çš„å€‰ä½ï¼ˆåªä½¿ç”¨ base stopï¼‰
                positions_without_lock = [p for p in positions if p.trail_callback is None]
                if positions_without_lock:
                    logger.debug(f"å…¶ä¸­ {len(positions_without_lock)} å€‹å€‰ä½æ²’æœ‰è¨­å®š lock_ratioï¼Œå°‡ä½¿ç”¨ base stop")
            
            # å°æ¯å€‹ position é€²è¡Œæª¢æŸ¥
            for position in positions:
                try:
                    await check_trailing_stop(position, db)
                except Exception as e:
                    logger.error(f"æª¢æŸ¥å€‰ä½ {position.id} ({position.symbol}) æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
                    # ç¹¼çºŒè™•ç†ä¸‹ä¸€å€‹å€‰ä½ï¼Œä¸è¦å› ç‚ºå–®ä¸€å€‰ä½éŒ¯èª¤è€Œåœæ­¢æ•´å€‹ä»»å‹™
                    continue
            
            # æª¢æŸ¥ Binance ä¸Šçš„é bot å‰µå»ºå€‰ä½
            try:
                await check_binance_non_bot_positions(db)
            except Exception as e:
                logger.error(f"æª¢æŸ¥ Binance é bot å‰µå»ºå€‰ä½æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
                # ç¹¼çºŒåŸ·è¡Œï¼Œä¸è¦å› ç‚ºé€™å€‹éŒ¯èª¤è€Œåœæ­¢æ•´å€‹ä»»å‹™
            
            # æª¢æŸ¥ Portfolio Trailing Stop
            try:
                await check_portfolio_trailing_stop(db)
            except Exception as e:
                logger.error(f"æª¢æŸ¥ Portfolio Trailing Stop æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
                # ç¹¼çºŒåŸ·è¡Œï¼Œä¸è¦å› ç‚ºé€™å€‹éŒ¯èª¤è€Œåœæ­¢æ•´å€‹ä»»å‹™
        
        except Exception as e:
            logger.error(f"è¿½è¹¤åœæä»»å‹™åŸ·è¡Œæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
            if db:
                db.rollback()
        
        finally:
            # ç¢ºä¿ DB session ä¸€å®šæœƒé—œé–‰
            if db:
                db.close()
        
        # ç­‰å¾… 5 ç§’å¾Œå†æ¬¡åŸ·è¡Œ
        await asyncio.sleep(5)


async def check_binance_non_bot_positions(db: Session):
    """
    æª¢æŸ¥ Binance ä¸Šçš„é bot å‰µå»ºå€‰ä½ï¼Œä¸¦è§¸ç™¼åœæï¼ˆå¦‚æœæ»¿è¶³æ¢ä»¶ï¼‰
    
    é€™å€‹å‡½æ•¸æœƒï¼š
    1. å¾ Binance ç²å–æ‰€æœ‰ open positions
    2. å°æ–¼æ¯å€‹ positionï¼Œæª¢æŸ¥æ˜¯å¦æœ‰å°æ‡‰çš„è³‡æ–™åº«è¨˜éŒ„
    3. å¦‚æœæ²’æœ‰ï¼ˆé bot å‰µå»ºçš„ï¼‰ï¼Œä½¿ç”¨è‡¨æ™‚ Position å°è±¡ä¾†æª¢æŸ¥åœæ
    4. å¦‚æœè§¸ç™¼åœæï¼Œç›´æ¥é—œé–‰ Binance å€‰ä½
    """
    try:
        # å˜—è©¦å–å¾— Binance client
        client = get_client()
        
        # ä½¿ç”¨ USDT-M Futures position info
        raw_positions = client.futures_position_information()
        
        for item in raw_positions:
            try:
                position_amt = float(item.get("positionAmt", "0") or 0)
            except (ValueError, TypeError):
                position_amt = 0.0
            
            if position_amt == 0:
                continue
            
            # è§£æå…¶ä»–æ¬„ä½
            try:
                entry_price = float(item.get("entryPrice", "0") or 0)
                mark_price = float(item.get("markPrice", "0") or 0)
                unrealized_pnl = float(item.get("unRealizedProfit", "0") or 0)
                leverage = int(float(item.get("leverage", "0") or 0))
            except (ValueError, TypeError) as e:
                logger.warning(f"è§£æ Binance position æ¬„ä½å¤±æ•—: {item.get('symbol', 'unknown')}, éŒ¯èª¤: {e}")
                continue
            
            symbol = item.get("symbol", "")
            side_local = "LONG" if position_amt > 0 else "SHORT"
            
            # æŸ¥æ‰¾åŒ¹é…çš„æœ¬åœ° Positionï¼ˆæœ€æ–°çš„ OPEN å€‰ä½ï¼‰
            local_pos = (
                db.query(Position)
                .filter(
                    Position.symbol == symbol.upper(),
                    Position.side == side_local,
                    Position.status == "OPEN",
                )
                .order_by(Position.id.desc())
                .first()
            )
            
            # å¦‚æœæ‰¾åˆ°æœ¬åœ° Positionï¼Œè·³éï¼ˆå·²ç¶“ç”± check_trailing_stop è™•ç†ï¼‰
            if local_pos:
                continue
            
            # é€™æ˜¯é bot å‰µå»ºçš„å€‰ä½ï¼Œéœ€è¦æª¢æŸ¥åœæ
            tracking_key = f"{symbol}|{side_local}"
            override_key = f"{symbol}|{side_local}"
            overrides = _binance_position_stop_overrides.get(override_key, {})
            
            # æª¢æŸ¥æ˜¯å¦å·²æœ‰è¿½è¹¤è¨˜éŒ„
            if tracking_key in _non_bot_position_tracking:
                tracked = _non_bot_position_tracking[tracking_key]
                tracked_entry = tracked.get("entry_price")
                tracked_highest = tracked.get("highest_price")
                
                # å¦‚æœ entry_price æ”¹è®Šï¼Œé‡ç½®è¿½è¹¤
                if tracked_entry is None or (abs(tracked_entry - entry_price) / max(abs(tracked_entry), abs(entry_price), 1.0)) > 0.001:
                    tracked_entry = entry_price
                    tracked_highest = None
            else:
                tracked_entry = entry_price
                tracked_highest = None
            
            # æ›´æ–°æ­·å²æœ€é«˜/æœ€ä½åƒ¹æ ¼
            if side_local == "LONG":
                if tracked_highest is None:
                    tracked_highest = max(mark_price, entry_price) if entry_price > 0 else mark_price
                else:
                    tracked_highest = max(tracked_highest, mark_price)
            else:
                if tracked_highest is None:
                    tracked_highest = min(mark_price, entry_price) if entry_price > 0 else mark_price
                else:
                    tracked_highest = min(tracked_highest, mark_price)
            
            if tracked_highest is None:
                tracked_highest = mark_price
            
            # æ›´æ–°è¿½è¹¤è¨˜éŒ„
            _non_bot_position_tracking[tracking_key] = {
                "entry_price": tracked_entry,
                "highest_price": tracked_highest,
                "side": side_local
            }
            
            # å»ºç«‹è‡¨æ™‚ Position ç‰©ä»¶
            class TempPosition:
                def __init__(self, entry_price, side, highest_price=None):
                    self.entry_price = entry_price
                    self.side = side
                    self.highest_price = highest_price
                    self.trail_callback = overrides.get("trail_callback")
                    self.dyn_profit_threshold_pct = overrides.get("dyn_profit_threshold_pct")
                    self.base_stop_loss_pct = overrides.get("base_stop_loss_pct")
                    self.symbol = symbol
            
            temp_pos = TempPosition(
                entry_price=tracked_entry if tracked_entry else entry_price,
                side=side_local,
                highest_price=tracked_highest
            )
            
            # è¨ˆç®— unrealized_pnl_pct
            calculated_unrealized_pnl_pct = None
            if entry_price > 0 and abs(position_amt) > 0 and leverage > 0:
                notional = abs(position_amt) * entry_price
                if notional > 0:
                    margin = notional / leverage
                    if margin > 0:
                        calculated_unrealized_pnl_pct = (unrealized_pnl / margin) * 100.0
            
            # ä½¿ç”¨ compute_stop_state è¨ˆç®—åœæç‹€æ…‹ï¼ˆå‚³å…¥ leverage å’Œ qty ç”¨æ–¼ margin-based base stop è¨ˆç®—ï¼‰
            stop_state = compute_stop_state(temp_pos, mark_price, calculated_unrealized_pnl_pct, leverage, abs(position_amt))
            
            # æ ¹æ“š stop_mode é¸æ“‡å°æ‡‰çš„åœæåƒ¹æ ¼
            triggered = False
            mode = None
            dyn_stop = None
            
            if stop_state.stop_mode == "dynamic":
                dyn_stop = stop_state.dynamic_stop_price
                if dyn_stop is not None:
                    if side_local == "LONG":
                        triggered = mark_price <= dyn_stop
                    else:  # SHORT
                        triggered = mark_price >= dyn_stop
                    mode = "dynamic_trailing"
            elif stop_state.stop_mode == "base":
                dyn_stop = stop_state.base_stop_price
                if dyn_stop is not None:
                    if side_local == "LONG":
                        triggered = mark_price <= dyn_stop
                    else:  # SHORT
                        triggered = mark_price >= dyn_stop
                    mode = "base_stop"
                else:
                    logger.warning(
                        f"é bot å‰µå»ºå€‰ä½ {symbol} ({side_local}) base mode ä½† base_stop_price ç‚º None"
                    )
            
            # è¨˜éŒ„æª¢æŸ¥çµæœ
            if dyn_stop is not None:
                logger.info(
                    f"é bot å‰µå»ºå€‰ä½ {symbol} ({side_local}) åœææª¢æŸ¥ï¼š"
                    f"mark={mark_price:.6f}, dyn_stop={dyn_stop:.6f}, "
                    f"stop_mode={stop_state.stop_mode}, triggered={triggered}, mode={mode}"
                )
            elif stop_state.stop_mode == "base":
                logger.warning(
                    f"é bot å‰µå»ºå€‰ä½ {symbol} ({side_local}) base mode ä½† dyn_stop ç‚º None, "
                    f"base_stop_price={stop_state.base_stop_price}"
                )
            
            # å¦‚æœè§¸ç™¼åœæï¼Œé—œé–‰å€‰ä½
            if triggered:
                logger.info(
                    f"é bot å‰µå»ºå€‰ä½ {symbol} ({side_local}) è§¸ç™¼ {mode}ï¼Œ"
                    f"ç›®å‰åƒ¹æ ¼: {mark_price}, åœæç·š: {dyn_stop}"
                )
                
                # auto_close_enabled å§‹çµ‚å•Ÿç”¨ï¼ˆå¼·åˆ¶ï¼‰
                # å‘¼å«é—œå€‰å‡½å¼
                try:
                    close_order = close_futures_position(
                        symbol=symbol,
                        position_side=side_local,
                        qty=abs(position_amt),
                        position_id=None  # é bot å‰µå»ºçš„å€‰ä½æ²’æœ‰ position_id
                    )
                    
                    logger.info(
                        f"é bot å‰µå»ºå€‰ä½ {symbol} ({side_local}) å·²é—œå€‰ï¼Œ"
                        f"order_id={close_order.get('orderId', 'unknown')}"
                    )
                    
                    # å–å¾—å¹³å€‰åƒ¹æ ¼
                    exit_price = get_exit_price_from_order(close_order, symbol)
                    
                    # å»ºç«‹ Position è¨˜éŒ„ï¼ˆç”¨æ–¼çµ±è¨ˆè¨ˆç®—ï¼‰
                    position = Position(
                        bot_id=None,  # é bot å‰µå»ºçš„å€‰ä½
                        tv_signal_log_id=None,  # é bot å‰µå»ºçš„å€‰ä½
                        symbol=symbol.upper(),
                        side=side_local,
                        qty=abs(position_amt),
                        entry_price=tracked_entry if tracked_entry else entry_price,
                        exit_price=exit_price,
                        status="CLOSED",
                        closed_at=datetime.now(timezone.utc),
                        exit_reason=mode,  # base_stop æˆ– dynamic_trailing
                        binance_order_id=int(close_order.get("orderId")) if close_order.get("orderId") else None,
                        client_order_id=close_order.get("clientOrderId"),
                        # è¨˜éŒ„åœæç›¸é—œé…ç½®ï¼ˆç”¨æ–¼è¿½è¹¤ï¼‰
                        trail_callback=overrides.get("trail_callback"),
                        dyn_profit_threshold_pct=overrides.get("dyn_profit_threshold_pct"),
                        base_stop_loss_pct=overrides.get("base_stop_loss_pct"),
                        highest_price=tracked_highest if tracked_highest else None,
                    )
                    
                    db.add(position)
                    db.commit()
                    db.refresh(position)
                    
                    logger.info(
                        f"é bot å‰µå»ºå€‰ä½ {symbol} ({side_local}) å·²å»ºç«‹è³‡æ–™åº«è¨˜éŒ„ "
                        f"(position_id={position.id}, exit_reason={mode}, exit_price={exit_price})"
                    )
                    
                    # æ¸…ç†è¿½è¹¤è¨˜éŒ„
                    if tracking_key in _non_bot_position_tracking:
                        del _non_bot_position_tracking[tracking_key]
                        logger.debug(f"æ¸…ç†é bot å€‰ä½è¿½è¹¤è¨˜éŒ„: {tracking_key}")
                    
                except Exception as e:
                    logger.error(f"é—œé–‰é bot å‰µå»ºå€‰ä½ {symbol} ({side_local}) å¤±æ•—: {e}")
                    db.rollback()
                    # ç¹¼çºŒè™•ç†ä¸‹ä¸€å€‹å€‰ä½
                    continue
                    
    except Exception as e:
        logger.error(f"æª¢æŸ¥ Binance é bot å‰µå»ºå€‰ä½æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        # ä¸è¦æ‹‹å‡ºç•°å¸¸ï¼Œè®“ä¸»å¾ªç’°ç¹¼çºŒé‹è¡Œ


async def check_portfolio_trailing_stop(db: Session):
    """
    æª¢æŸ¥ Portfolio-level Trailing Stopï¼ˆåˆ†åˆ¥æª¢æŸ¥ LONG å’Œ SHORTï¼‰
    
    é‚è¼¯ï¼š
    1. åˆ†åˆ¥è¨ˆç®— LONG å’Œ SHORT å€‰ä½çš„ç¸½ PnL
    2. å°æ¯å€‹é¡åˆ¥ï¼ˆLONG/SHORTï¼‰ï¼š
       - å¦‚æœ enabled=True ä¸” target_pnl å·²è¨­å®šï¼š
         - å¦‚æœç¸½ PnL >= target_pnl ä¸” max_pnl_reached ç‚º Noneï¼Œè¨˜éŒ„ max_pnl_reached
         - å¦‚æœç¸½ PnL >= target_pnlï¼Œæ›´æ–° max_pnl_reachedï¼ˆåªå¢ä¸æ¸›ï¼‰
         - å¦‚æœ max_pnl_reached å·²è¨˜éŒ„ï¼Œè¨ˆç®— sell_threshold = max_pnl_reached * lock_ratio
         - å¦‚æœç¸½ PnL <= sell_thresholdï¼Œè§¸ç™¼è‡ªå‹•è³£å‡ºè©²é¡åˆ¥çš„æ‰€æœ‰å€‰ä½
    """
    global _portfolio_trailing_runtime_state
    
    # å¾è³‡æ–™åº«è¼‰å…¥é…ç½®ï¼ˆid=1 for LONG, id=2 for SHORTï¼‰
    long_config = db.query(PortfolioTrailingConfig).filter(PortfolioTrailingConfig.id == 1).first()
    short_config = db.query(PortfolioTrailingConfig).filter(PortfolioTrailingConfig.id == 2).first()
    
    try:
        client = get_client()
        raw_positions = client.futures_position_information()
        
        # åˆ†é›¢ LONG å’Œ SHORT å€‰ä½
        long_positions = []
        short_positions = []
        long_total_pnl = 0.0
        short_total_pnl = 0.0
        
        for item in raw_positions:
            try:
                position_amt = float(item.get("positionAmt", "0") or 0)
            except (ValueError, TypeError):
                continue
            
            if position_amt == 0:
                continue
            
            symbol = item.get("symbol", "")
            if not symbol:
                continue
            
            # LONG: positionAmt > 0, SHORT: positionAmt < 0
            position_side = "LONG" if position_amt > 0 else "SHORT"
            
            try:
                # Binance API çš„ unRealizedProfit å·²ç¶“æ­£ç¢ºè¨ˆç®—ï¼š
                # - LONG: åƒ¹æ ¼ä¸Šæ¼²æ™‚ç‚ºæ­£ï¼Œåƒ¹æ ¼ä¸‹è·Œæ™‚ç‚ºè² 
                # - SHORT: åƒ¹æ ¼ä¸‹è·Œæ™‚ç‚ºæ­£ï¼Œåƒ¹æ ¼ä¸Šæ¼²æ™‚ç‚ºè² 
                unrealized_pnl = float(item.get("unRealizedProfit", "0") or 0)
                if position_side == "LONG":
                    long_positions.append(item)
                    long_total_pnl += unrealized_pnl
                else:
                    short_positions.append(item)
                    short_total_pnl += unrealized_pnl
            except (ValueError, TypeError):
                continue
        
        # è¨˜éŒ„ç¸½ PnLï¼ˆç”¨æ–¼èª¿è©¦ï¼‰
        logger.debug(
            f"[Portfolio Trailing] å€‰ä½çµ±è¨ˆ: LONG={len(long_positions)} å€‹ï¼Œç¸½ PnL={long_total_pnl:.4f} USDTï¼›"
            f"SHORT={len(short_positions)} å€‹ï¼Œç¸½ PnL={short_total_pnl:.4f} USDT"
        )
        
        # åˆ†åˆ¥æª¢æŸ¥ LONG å’Œ SHORT
        for side_name, config, positions_list, total_pnl in [
            ("LONG", long_config, long_positions, long_total_pnl),
            ("SHORT", short_config, short_positions, short_total_pnl)
        ]:
            logger.debug(
                f"[Portfolio Trailing {side_name}] ğŸ” æª¢æŸ¥é–‹å§‹: "
                f"total_pnl={total_pnl:.4f} USDT, å€‰ä½æ•¸={len(positions_list)}, "
                f"configå­˜åœ¨={config is not None}, enabled={config.enabled if config else False}, "
                f"target_pnl={config.target_pnl if config else None} USDT"
            )
            
            # å¦‚æœé…ç½®ä¸å­˜åœ¨æˆ–æœªå•Ÿç”¨ï¼Œè·³é
            if not config or not config.enabled:
                logger.debug(f"[Portfolio Trailing {side_name}] â­ï¸ è·³éï¼ˆé…ç½®ä¸å­˜åœ¨æˆ–æœªå•Ÿç”¨ï¼‰")
                continue
            
            target_pnl = config.target_pnl
            if target_pnl is None:
                logger.debug(f"[Portfolio Trailing {side_name}] â­ï¸ è·³éï¼ˆtarget_pnl æœªè¨­å®šï¼‰")
                continue
            
            # ç¢ºä¿è©²é¡åˆ¥çš„é‹è¡Œæ™‚ç‹€æ…‹å­˜åœ¨
            if side_name.lower() not in _portfolio_trailing_runtime_state:
                _portfolio_trailing_runtime_state[side_name.lower()] = {"max_pnl_reached": None, "last_check_time": None}
            
            # å–å¾—è©²é¡åˆ¥çš„é‹è¡Œæ™‚ç‹€æ…‹
            side_state = _portfolio_trailing_runtime_state.get(side_name.lower(), {})
            max_pnl_reached = side_state.get("max_pnl_reached")
            
            # å¦‚æœé”åˆ°ç›®æ¨™ï¼Œè¨˜éŒ„æˆ–æ›´æ–° max_pnl_reached
            # è¦å‰‡ï¼šä¸€æ—¦è¨­å®šï¼Œåªå¢ä¸æ¸›ï¼Œå³ä½¿ PnL å†æ¬¡é”åˆ°ç›®æ¨™ï¼ˆä½†ä½æ–¼ maxï¼‰ä¹Ÿä¸æœƒé‡ç½®
            old_max_pnl_reached = max_pnl_reached  # ä¿å­˜èˆŠå€¼ç”¨æ–¼æ—¥èªŒ
            if total_pnl >= target_pnl:
                if max_pnl_reached is None:
                    # é¦–æ¬¡é”åˆ°ç›®æ¨™ï¼Œè¨˜éŒ„ max_pnl_reached
                    _portfolio_trailing_runtime_state[side_name.lower()]["max_pnl_reached"] = total_pnl
                    max_pnl_reached = total_pnl
                    logger.info(
                        f"[Portfolio Trailing {side_name}] âœ… ç¸½ PnL ({total_pnl:.4f}) é¦–æ¬¡é”åˆ°ç›®æ¨™ {target_pnl:.4f}ï¼Œ"
                        f"è¨˜éŒ„æœ€å¤§ PnL: {max_pnl_reached:.4f}"
                    )
                elif total_pnl > max_pnl_reached:
                    # å¦‚æœç•¶å‰ PnL é«˜æ–¼å·²è¨˜éŒ„çš„æœ€å¤§å€¼ï¼Œæ›´æ–°æœ€å¤§å€¼ï¼ˆåªå¢ä¸æ¸›ï¼‰
                    old_max_pnl_reached = max_pnl_reached  # ä¿å­˜æ›´æ–°å‰çš„å€¼
                    _portfolio_trailing_runtime_state[side_name.lower()]["max_pnl_reached"] = total_pnl
                    max_pnl_reached = total_pnl
                    logger.info(
                        f"[Portfolio Trailing {side_name}] âœ… ç¸½ PnL ({total_pnl:.4f}) è¶…è¶Šä¹‹å‰è¨˜éŒ„ ({old_max_pnl_reached:.4f})ï¼Œ"
                        f"æ›´æ–°æœ€å¤§ PnL: {max_pnl_reached:.4f}"
                    )
                else:
                    # ç•¶å‰ PnL åœ¨ç›®æ¨™ä»¥ä¸Šä½†ä½æ–¼å·²è¨˜éŒ„çš„æœ€å¤§å€¼ï¼ˆä¸æ›´æ–°ï¼Œä¿æŒæœ€å¤§è¨˜éŒ„ï¼‰
                    logger.debug(
                        f"[Portfolio Trailing {side_name}] ç¸½ PnL ({total_pnl:.4f}) >= ç›®æ¨™ ({target_pnl:.4f}) "
                        f"ä½†ä½æ–¼å·²è¨˜éŒ„æœ€å¤§å€¼ ({max_pnl_reached:.4f})ï¼Œä¸æ›´æ–°"
                    )
            
            # å¦‚æœå·²è¨˜éŒ„ max_pnl_reachedï¼Œæª¢æŸ¥æ˜¯å¦éœ€è¦è³£å‡º
            if max_pnl_reached is not None:
                # å–å¾—æœ‰æ•ˆçš„ lock_ratioï¼ˆportfolio è¨­å®šå„ªå…ˆï¼Œå¦å‰‡ä½¿ç”¨å…¨å±€ï¼‰
                lock_ratio = config.lock_ratio
                if lock_ratio is None:
                    # ä½¿ç”¨å°æ‡‰æ–¹å‘çš„å…¨å±€ lock_ratio
                    side_config = TRAILING_CONFIG.get_config_for_side(side_name)
                    lock_ratio = side_config.lock_ratio if side_config.lock_ratio is not None else DYN_LOCK_RATIO_DEFAULT
                
                # è¨ˆç®—è³£å‡ºé–€æª»
                sell_threshold = max_pnl_reached * lock_ratio
                
                # è¨˜éŒ„æª¢æŸ¥ç‹€æ…‹ï¼ˆç”¨æ–¼èª¿è©¦ï¼‰ - ä½¿ç”¨ INFO ç´šåˆ¥ä»¥ä¾¿åœ¨ç”Ÿç”¢ç’°å¢ƒä¸­ä¹Ÿèƒ½çœ‹åˆ°
                should_trigger_sell = total_pnl <= sell_threshold
                
                # ç‚º SHORT æ·»åŠ ç‰¹æ®Šèªªæ˜ï¼ˆå› ç‚º SHORT çš„ PnL è¡Œç‚ºèˆ‡ LONG ç›¸åï¼‰
                side_note = ""
                if side_name == "SHORT":
                    side_note = " (SHORT: åƒ¹æ ¼ä¸‹è·Œæ™‚ PnL å¢åŠ ï¼Œåƒ¹æ ¼ä¸Šæ¼²æ™‚ PnL æ¸›å°‘)"
                
                logger.info(
                    f"[Portfolio Trailing {side_name}] ğŸ“Š è³£å‡ºæ¢ä»¶æª¢æŸ¥{side_note}: "
                    f"total_pnl={total_pnl:.4f} USDT, "
                    f"max_pnl_reached={max_pnl_reached:.4f} USDT, "
                    f"lock_ratio={lock_ratio:.4f}, "
                    f"sell_threshold={sell_threshold:.4f} USDT (= {max_pnl_reached:.4f} Ã— {lock_ratio:.4f}), "
                    f"æ¢ä»¶æª¢æŸ¥: {total_pnl:.4f} <= {sell_threshold:.4f} = {should_trigger_sell}, "
                    f"å€‰ä½æ•¸é‡={len(positions_list)}"
                )
                
                # æª¢æŸ¥æ˜¯å¦æ‡‰è©²è§¸ç™¼è‡ªå‹•è³£å‡º
                if should_trigger_sell:
                    logger.warning(
                        f"[Portfolio Trailing {side_name}] âš ï¸ ç¸½ PnL ({total_pnl:.4f}) å·²é™è‡³è³£å‡ºé–€æª» ({sell_threshold:.4f} = "
                        f"{max_pnl_reached:.4f} Ã— {lock_ratio:.4f})ï¼Œé–‹å§‹é—œé–‰æ‰€æœ‰ {side_name} å€‰ä½ï¼ˆå…± {len(positions_list)} å€‹ï¼‰"
                    )
                    
                    # é©—è­‰å€‰ä½åˆ—è¡¨ä¸ç‚ºç©º
                    if not positions_list:
                        logger.warning(
                            f"[Portfolio Trailing {side_name}] è­¦å‘Šï¼šè§¸ç™¼è‡ªå‹•è³£å‡ºä½†å€‰ä½åˆ—è¡¨ç‚ºç©ºï¼Œè·³éé—œå€‰æ“ä½œ"
                        )
                        # ä»ç„¶é‡ç½® max_pnl_reachedï¼ˆé¿å…é‡è¤‡è§¸ç™¼ï¼‰
                        _portfolio_trailing_runtime_state[side_name.lower()]["max_pnl_reached"] = None
                        continue
                    
                    # é—œé–‰è©²é¡åˆ¥çš„æ‰€æœ‰å€‰ä½
                    closed_count = 0
                    errors = []
                    
                    for item in positions_list:
                        try:
                            position_amt = float(item.get("positionAmt", "0") or 0)
                        except (ValueError, TypeError):
                            continue
                        
                        if position_amt == 0:
                            continue
                        
                        symbol = item.get("symbol", "")
                        if not symbol:
                            continue
                        
                        try:
                            # é—œé–‰å€‰ä½çš„è¨‚å–®æ–¹å‘ï¼š
                            # - LONG å€‰ä½ï¼šä½¿ç”¨ SELL è¨‚å–®ï¼ˆè³£å‡ºå¹³å¤šï¼‰
                            # - SHORT å€‰ä½ï¼šä½¿ç”¨ BUY è¨‚å–®ï¼ˆè²·å…¥å¹³ç©ºï¼‰
                            side = "SELL" if side_name == "LONG" else "BUY"
                            qty = abs(position_amt)  # SHORT çš„ position_amt ç‚ºè² æ•¸ï¼Œä½¿ç”¨ abs å–å¾—æ­£æ•¸é‡
                            
                            timestamp = int(time.time() * 1000)
                            client_order_id = f"TVBOT_PORTFOLIO_TRAILING_{side_name}_{timestamp}_{closed_count}"
                            
                            logger.info(f"[Portfolio Trailing {side_name}] é—œé–‰ {symbol}ï¼Œæ•¸é‡: {qty}")
                            
                            order = client.futures_create_order(
                                symbol=symbol,
                                side=side,
                                type="MARKET",
                                quantity=qty,
                                reduceOnly=True,
                                newClientOrderId=client_order_id
                            )
                            
                            closed_count += 1
                            logger.info(
                                f"[Portfolio Trailing {side_name}] æˆåŠŸé—œé–‰ {symbol}ï¼Œ"
                                f"è¨‚å–®ID: {order.get('orderId')}"
                            )
                            
                            # æ›´æ–°æœ¬åœ°è³‡æ–™åº«ä¸­çš„ Position è¨˜éŒ„ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
                            try:
                                local_positions = db.query(Position).filter(
                                    Position.symbol == symbol.upper(),
                                    Position.side == side_name,
                                    Position.status == "OPEN"
                                ).all()
                                
                                for local_pos in local_positions:
                                    exit_price = get_exit_price_from_order(order, symbol)
                                    local_pos.status = "CLOSED"
                                    local_pos.exit_price = exit_price
                                    local_pos.closed_at = datetime.now(timezone.utc)
                                    local_pos.exit_reason = f"portfolio_trailing_auto_sell_{side_name.lower()}"
                                    logger.debug(
                                        f"[Portfolio Trailing {side_name}] æ›´æ–°æœ¬åœ°å€‰ä½è¨˜éŒ„ {local_pos.id} "
                                        f"({symbol}) ç‚º CLOSED"
                                    )
                                
                                if local_positions:
                                    db.commit()
                            except Exception as db_error:
                                logger.warning(
                                    f"[Portfolio Trailing {side_name}] æ›´æ–°æœ¬åœ°å€‰ä½è¨˜éŒ„æ™‚ç™¼ç”ŸéŒ¯èª¤: {db_error}"
                                )
                                db.rollback()
                            
                        except Exception as e:
                            error_msg = f"{symbol}: {str(e)}"
                            errors.append(error_msg)
                            logger.error(f"[Portfolio Trailing {side_name}] é—œé–‰ {symbol} å¤±æ•—: {e}")
                    
                    # é‡ç½®è©²é¡åˆ¥çš„ max_pnl_reachedï¼ˆæ‰€æœ‰å€‰ä½å·²é—œé–‰æˆ–è‡ªå‹•è³£å‡ºå·²è§¸ç™¼ï¼‰
                    _portfolio_trailing_runtime_state[side_name.lower()]["max_pnl_reached"] = None
                    logger.info(
                        f"[Portfolio Trailing {side_name}] âœ… è‡ªå‹•è³£å‡ºè§¸ç™¼ï¼Œå·²é‡ç½® Max PnL Reachedã€‚"
                        f"é—œé–‰ {closed_count} å€‹å€‰ä½ï¼ŒéŒ¯èª¤æ•¸: {len(errors)}"
                    )
                    if errors:
                        logger.error(f"[Portfolio Trailing {side_name}] âŒ é—œé–‰å€‰ä½æ™‚çš„éŒ¯èª¤è©³æƒ…: {errors}")
            
            # æ›´æ–°æœ€å¾Œæª¢æŸ¥æ™‚é–“
            _portfolio_trailing_runtime_state[side_name.lower()]["last_check_time"] = time.time()
        
    except Exception as e:
        logger.error(f"[Portfolio Trailing] æª¢æŸ¥æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}", exc_info=True)


def get_exit_price_from_order(close_order: dict, symbol: str) -> float:
    """
    å¾é—œå€‰è¨‚å–®å›å‚³ä¸­å–å¾—å¹³å€‰åƒ¹æ ¼
    
    å„ªå…ˆé †åºï¼š
    1. close_order.get("avgPrice") - å¹³å‡æˆäº¤åƒ¹æ ¼ï¼ˆå­˜åœ¨ä¸”éç©ºå­—ä¸²ä¸” > 0ï¼‰
    2. æŸ¥è©¢è¨‚å–®è©³æƒ…å–å¾— avgPriceï¼ˆå¦‚æœè¨‚å–® ID å­˜åœ¨ï¼‰
    3. close_order.get("price") - è¨‚å–®åƒ¹æ ¼ï¼ˆå­˜åœ¨ä¸” > 0ï¼‰
    4. get_mark_price(symbol) - æ¨™è¨˜åƒ¹æ ¼ï¼ˆfallbackï¼‰
    
    æ­¤å‡½å¼ä¿è­‰ä¸€å®šæœƒå›å‚³ä¸€å€‹ float å€¼ï¼Œä¸æœƒæ‹‹å‡ºä¾‹å¤–ã€‚
    
    Args:
        close_order: å¹£å®‰ API å›å‚³çš„é—œå€‰è¨‚å–®è³‡è¨Š
        symbol: äº¤æ˜“å°
    
    Returns:
        float: å¹³å€‰åƒ¹æ ¼
    """
    try:
        # å„ªå…ˆä½¿ç”¨ avgPriceï¼ˆå¹³å‡æˆäº¤åƒ¹æ ¼ï¼‰- é€™æ˜¯ Binance çš„å¯¦éš›æˆäº¤å¹³å‡åƒ¹æ ¼
        # æª¢æŸ¥å­˜åœ¨ä¸”ä¸æ˜¯ç©ºå­—ä¸²ã€ä¸æ˜¯ "0"ã€ä¸” > 0
        avg_price = close_order.get("avgPrice")
        if avg_price is not None and avg_price != "" and str(avg_price).strip() != "0":
            try:
                avg_price_float = float(avg_price)
                if avg_price_float > 0:
                    logger.info(f"å¾è¨‚å–®å›å‚³å–å¾— {symbol} å¹³å€‰åƒ¹æ ¼ (avgPrice): {avg_price_float}")
                    return avg_price_float
            except (ValueError, TypeError):
                pass
        
        # å¦‚æœè¨‚å–®å›å‚³ä¸­æ²’æœ‰æœ‰æ•ˆçš„ avgPriceï¼Œå˜—è©¦æŸ¥è©¢è¨‚å–®è©³æƒ…
        order_id = close_order.get("orderId")
        if order_id:
            try:
                client = get_client()
                # ç­‰å¾…ä¸€å°æ®µæ™‚é–“ç¢ºä¿è¨‚å–®å·²æˆäº¤
                time.sleep(0.3)
                # æŸ¥è©¢è¨‚å–®è©³æƒ…
                order_detail = client.futures_get_order(symbol=symbol, orderId=order_id)
                avg_price_detail = order_detail.get("avgPrice")
                if avg_price_detail:
                    try:
                        avg_price_float = float(avg_price_detail)
                        if avg_price_float > 0:
                            logger.info(f"å¾è¨‚å–®è©³æƒ…å–å¾— {symbol} å¹³å€‰åƒ¹æ ¼ (avgPrice): {avg_price_float}")
                            return avg_price_float
                    except (ValueError, TypeError):
                        pass
            except Exception as e:
                logger.debug(f"æŸ¥è©¢è¨‚å–® {order_id} è©³æƒ…æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        
        # å…¶æ¬¡ä½¿ç”¨ priceï¼ˆè¨‚å–®åƒ¹æ ¼ï¼‰
        price = close_order.get("price")
        if price is not None:
            try:
                price_float = float(price)
                if price_float > 0:
                    return price_float
            except (ValueError, TypeError):
                pass
        
        # å¦‚æœéƒ½æ²’æœ‰ï¼Œä½¿ç”¨æ¨™è¨˜åƒ¹æ ¼ä½œç‚º fallback
        logger.warning(f"ç„¡æ³•å¾è¨‚å–®ä¸­å–å¾—å¹³å€‰åƒ¹æ ¼ï¼Œä½¿ç”¨ {symbol} æ¨™è¨˜åƒ¹æ ¼ä½œç‚º fallback")
        return get_mark_price(symbol)
    
    except (ValueError, TypeError) as e:
        # å¦‚æœè½‰æ› float å¤±æ•—ï¼Œä½¿ç”¨æ¨™è¨˜åƒ¹æ ¼ä½œç‚º fallback
        logger.warning(
            f"å¾è¨‚å–®ä¸­è§£æå¹³å€‰åƒ¹æ ¼æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}ï¼Œ"
            f"ä½¿ç”¨ {symbol} æ¨™è¨˜åƒ¹æ ¼ä½œç‚º fallback"
        )
        return get_mark_price(symbol)
    
    except Exception as e:
        # å…¶ä»–æœªé æœŸçš„éŒ¯èª¤ï¼ˆä¾‹å¦‚ get_mark_price å¤±æ•—ï¼‰
        # è¨˜éŒ„éŒ¯èª¤ä¸¦å›å‚³ä¸€å€‹é è¨­å€¼ï¼Œç¢ºä¿å‡½å¼ä¸æœƒå´©æ½°
        logger.error(
            f"å–å¾—å¹³å€‰åƒ¹æ ¼æ™‚ç™¼ç”Ÿæœªé æœŸéŒ¯èª¤: {e}ï¼Œ"
            f"ä½¿ç”¨ {symbol} æ¨™è¨˜åƒ¹æ ¼ä½œç‚º fallback"
        )
        try:
            return get_mark_price(symbol)
        except Exception:
            # å¦‚æœé€£æ¨™è¨˜åƒ¹æ ¼éƒ½å–å¾—å¤±æ•—ï¼Œå›å‚³ 0.0 ä½œç‚ºæœ€å¾Œçš„ fallback
            logger.critical(
                f"ç„¡æ³•å–å¾— {symbol} æ¨™è¨˜åƒ¹æ ¼ï¼Œå›å‚³ 0.0 ä½œç‚ºé è¨­å€¼ã€‚"
                f"é€™è¡¨ç¤ºå¹£å®‰ API é€£ç·šå¯èƒ½æœ‰å•é¡Œã€‚"
            )
            return 0.0


@dataclass
class StopState:
    """åœæç‹€æ…‹è³‡è¨Š"""
    stop_mode: str  # "dynamic", "base", "none", "dynamic_locked"
    base_stop_price: Optional[float]
    dynamic_stop_price: Optional[float]


def compute_stop_state(position: Position, mark_price: float, unrealized_pnl_pct: Optional[float] = None, leverage: Optional[int] = None, qty: Optional[float] = None) -> StopState:
    """
    è¨ˆç®—å€‰ä½çš„åœæç‹€æ…‹ï¼ˆç´”è¨ˆç®—å‡½æ•¸ï¼Œä¸ä¿®æ”¹ DB æˆ–ä¸‹å–®ï¼‰
    
    çµ¦å®šä¸€å€‹ Position å’Œç•¶å‰æ¨™è¨˜åƒ¹æ ¼ï¼Œè¨ˆç®—ï¼š
    - ç•¶å‰ç”Ÿæ•ˆçš„åœææ¨¡å¼
    - base_stop_price å’Œ dynamic_stop_price
    
    æ­¤å‡½æ•¸ä¸æœƒä¿®æ”¹è³‡æ–™åº«æˆ–ä¸‹å–®ï¼Œåªé€²è¡Œè¨ˆç®—ã€‚
    
    Args:
        position: Position æ¨¡å‹å¯¦ä¾‹
        mark_price: ç•¶å‰æ¨™è¨˜åƒ¹æ ¼
        unrealized_pnl_pct: æœªå¯¦ç¾ç›ˆè™§ç™¾åˆ†æ¯”ï¼ˆPnL%ï¼ŒåŸºæ–¼ margin è¨ˆç®—ï¼‰ï¼Œå¦‚æœç‚º None å‰‡ä½¿ç”¨åƒ¹æ ¼ç™¾åˆ†æ¯”è¨ˆç®—
        leverage: æ æ¡¿å€æ•¸ï¼Œç”¨æ–¼è¨ˆç®—åŸºæ–¼ margin çš„ Base SL%ã€‚å¦‚æœç‚º Noneï¼Œå‰‡ä½¿ç”¨é è¨­å€¼ 20
        qty: å€‰ä½æ•¸é‡ï¼Œç”¨æ–¼è¨ˆç®— marginã€‚å¦‚æœç‚º Noneï¼Œå‰‡å˜—è©¦å¾ position.qty ç²å–
    
    Returns:
        StopState: åœæç‹€æ…‹è³‡è¨Š
    """
    try:
        entry = position.entry_price
        best = position.highest_price  # LONG: æœ€é«˜åƒ¹; SHORT: æœ€ä½åƒ¹
        mark = mark_price
        
        # è‹¥ entry <= 0ï¼Œè¿”å› none
        if entry <= 0:
            return StopState(
                stop_mode="none",
                base_stop_price=None,
                dynamic_stop_price=None
            )
        
        # ç²å– qty å’Œ leverageï¼ˆç”¨æ–¼è¨ˆç®— margin-based base stop priceï¼‰
        position_qty = qty if qty is not None else getattr(position, 'qty', 0)
        position_leverage = leverage if leverage is not None else 20  # é»˜èªæ æ¡¿
        
        # æ ¹æ“šå€‰ä½æ–¹å‘å–å¾—å°æ‡‰çš„å…¨å±€è¨­å®š
        side_config = TRAILING_CONFIG.get_config_for_side(position.side)
        trailing_enabled = TRAILING_CONFIG.trailing_enabled if TRAILING_CONFIG.trailing_enabled is not None else DYN_TRAILING_ENABLED
        
        # ä½¿ç”¨å°æ‡‰æ–¹å‘çš„å…¨å±€è¨­å®šä½œç‚ºé»˜èªå€¼ï¼Œå¦‚æœæ²’æœ‰å‰‡ä½¿ç”¨ç’°å¢ƒè®Šæ•¸
        base_sl_pct_default = side_config.base_sl_pct if side_config.base_sl_pct is not None else DYN_BASE_SL_PCT
        profit_threshold_pct_default = side_config.profit_threshold_pct if side_config.profit_threshold_pct is not None else DYN_PROFIT_THRESHOLD_PCT
        
        # å„ªå…ˆä½¿ç”¨å€‰ä½è¦†å¯«å€¼ï¼Œå¦‚æœæ²’æœ‰å‰‡ä½¿ç”¨å…¨å±€é…ç½®
        if position.base_stop_loss_pct is not None:
            base_sl_pct = position.base_stop_loss_pct
        else:
            base_sl_pct = base_sl_pct_default
        
        if position.dyn_profit_threshold_pct is not None:
            profit_threshold_pct = position.dyn_profit_threshold_pct
        else:
            profit_threshold_pct = profit_threshold_pct_default
        
        # å…ˆæ±ºå®šé€™ç­†å–®ä½¿ç”¨çš„ lock_ratio
        # trail_callback: null â†’ ä½¿ç”¨å…¨å±€é…ç½®, 0 â†’ base stop only, >0 â†’ ä½¿ç”¨è©²å€¼ä½œç‚º lock_ratio
        # ä½¿ç”¨ getattr å®‰å…¨åœ°è¨ªå•å±¬æ€§ï¼ˆæ”¯æŒ TempPosition å’Œ Position å°è±¡ï¼‰
        trail_callback_override = getattr(position, 'trail_callback', None)
        if trail_callback_override is None:
            # ä½¿ç”¨å°æ‡‰æ–¹å‘çš„ TRAILING_CONFIG lock_ratioï¼ˆå¦‚æœæœ‰çš„è©±ï¼‰ï¼Œå¦å‰‡ä½¿ç”¨é è¨­å€¼
            lock_ratio = side_config.lock_ratio if side_config.lock_ratio is not None else DYN_LOCK_RATIO_DEFAULT
        elif trail_callback_override == 0:
            lock_ratio = None
        else:
            lock_ratio = trail_callback_override
        
        # ç¯„åœé˜²å‘†
        if lock_ratio is not None:
            if lock_ratio <= 0:
                lock_ratio = None
            elif lock_ratio > 1:
                lock_ratio = 1.0
        
        # è™•ç† LONG å€‰ä½
        if position.side == "LONG":
            # å¦‚æœ best ç‚º Noneï¼Œä½¿ç”¨ç•¶å‰åƒ¹æ ¼
            if best is None:
                best = mark
            
            # æ›´æ–° bestï¼ˆåƒ…ç”¨æ–¼è¨ˆç®—ï¼Œä¸ä¿®æ”¹ DBï¼‰
            # é‡è¦ï¼šbest åªèƒ½ä¸Šå‡ï¼Œä¸èƒ½ä¸‹é™ï¼Œé€™æ¨£ dynamic stop æ‰èƒ½ä¿æŒç©©å®š
            # é€™æ˜¯ trailing stop çš„æ ¸å¿ƒé‚è¼¯ï¼šåŸºæ–¼æ­·å²æœ€é«˜åƒ¹æ ¼è¨ˆç®—åœæ
            if mark > best:
                best = mark
            
            # profit_pct åŸºæ–¼æ­·å²æœ€é«˜åƒ¹æ ¼ï¼ˆbestï¼‰è¨ˆç®—ï¼Œè€Œä¸æ˜¯ç•¶å‰åƒ¹æ ¼ï¼ˆmarkï¼‰
            # é€™æ¨£å³ä½¿ç•¶å‰åƒ¹æ ¼ä¸‹è·Œï¼Œåªè¦æ­·å²æœ€é«˜ profit_pct >= thresholdï¼Œå°±æœƒä¿æŒåœ¨ dynamic mode
            profit_pct = (best - entry) / entry * 100.0 if entry > 0 else 0.0
            
            # è¨ˆç®— base stop priceï¼ˆåŸºæ–¼ marginï¼‰
            # æ ¹æ“šéœ€æ±‚ï¼šStop price = best - (Margin * Base SL% / 100) / qty
            # Margin = (Entry Price * Qty) / Leverage
            # Base SL Amount (USDT) = Margin * (Base SL% / 100.0)
            # å°æ–¼ LONG: stop_price = best - (margin * base_sl_pct / 100) / qty
            base_stop_price = None
            if base_sl_pct > 0 and entry > 0 and position_qty > 0 and position_leverage > 0:
                # è¨ˆç®— marginï¼ˆä½¿ç”¨ entry price è¨ˆç®— marginï¼‰
                notional = entry * position_qty
                margin = notional / position_leverage
                # ä½¿ç”¨ best åƒ¹æ ¼ï¼ˆå¦‚æœ best ç‚º Noneï¼Œä½¿ç”¨ entryï¼‰
                best_price = best if best is not None else entry
                # Base stop price = best - (margin * base_sl_pct / 100) / qty
                base_stop_price = best_price - (margin * base_sl_pct / 100.0) / position_qty
            
            # è¨ˆç®— dynamic stop price
            dynamic_stop_price = None
            stop_mode = "none"
            
            # æª¢æŸ¥æ˜¯å¦æœ‰è¦†å¯«å€¼ï¼ˆç”¨æ–¼æ±ºå®šæ˜¯å¦å•Ÿç”¨åœæï¼‰
            dyn_profit_threshold_pct_override = getattr(position, 'dyn_profit_threshold_pct', None)
            base_stop_loss_pct_override = getattr(position, 'base_stop_loss_pct', None)
            has_override = (
                trail_callback_override is not None or 
                dyn_profit_threshold_pct_override is not None or 
                base_stop_loss_pct_override is not None
            )
            # å¦‚æœæœ‰è¦†å¯«å€¼ï¼Œå³ä½¿å…¨å±€ trailing_enabled ç‚º Falseï¼Œä¹Ÿæ‡‰è©²å•Ÿç”¨åœæ
            effective_trailing_enabled = trailing_enabled or (has_override and lock_ratio is not None)
            
            # é—œéµé‚è¼¯ï¼šä¸€æ—¦é€²å…¥ dynamic modeï¼Œå°±æ‡‰è©²ä¿æŒåœ¨ dynamic mode
            # åˆ¤æ–·æ˜¯å¦æ‡‰è©²é€²å…¥æˆ–ä¿æŒåœ¨ dynamic modeï¼š
            # 1. å¦‚æœ bestï¼ˆæ­·å²æœ€é«˜åƒ¹æ ¼ï¼‰æ›¾ç¶“é”åˆ°é thresholdï¼Œå°±æ‡‰è©²ä¿æŒåœ¨ dynamic mode
            # 2. ä½¿ç”¨åŸºæ–¼ best çš„ PnL% ä¾†åˆ¤æ–·ï¼Œè€Œä¸æ˜¯ç•¶å‰åƒ¹æ ¼
            
            # è¨ˆç®—åŸºæ–¼ bestï¼ˆæ­·å²æœ€é«˜åƒ¹æ ¼ï¼‰çš„ profit_pctï¼ˆåƒ¹æ ¼ç™¾åˆ†æ¯”ï¼‰
            profit_pct_based_on_best = profit_pct  # å·²ç¶“åŸºæ–¼ best è¨ˆç®—
            
            # å¦‚æœæä¾›äº† unrealized_pnl_pctï¼Œæˆ‘å€‘éœ€è¦è¨ˆç®—åŸºæ–¼ best çš„ PnL%
            # åŸºæ–¼ best çš„ PnL% = (best - entry) / entry * 100 * (entry * qty / leverage) / (entry * qty / leverage)
            # ç°¡åŒ–å¾Œï¼šåŸºæ–¼ best çš„ PnL% = (best - entry) / entry * 100 * leverage / leverage
            # å¯¦éš›ä¸Šï¼ŒPnL% å’Œåƒ¹æ ¼ç™¾åˆ†æ¯”çš„æ¯”ä¾‹æ˜¯å›ºå®šçš„ï¼ˆéƒ½åŸºæ–¼ç›¸åŒçš„ marginï¼‰
            # æ‰€ä»¥ï¼šåŸºæ–¼ best çš„ PnL% = unrealized_pnl_pct * (best - entry) / (mark - entry)
            if unrealized_pnl_pct is not None:
                # ä½¿ç”¨ç•¶å‰ PnL% ä¾†åˆ¤æ–·æ˜¯å¦é€²å…¥ dynamic mode
                profit_pct_for_threshold = unrealized_pnl_pct
                # è¨ˆç®—åŸºæ–¼ best çš„ PnL%ï¼Œç”¨æ–¼åˆ¤æ–·æ˜¯å¦æ‡‰è©²ä¿æŒåœ¨ dynamic mode
                # å¦‚æœ mark != entryï¼Œä½¿ç”¨æ¯”ä¾‹è¨ˆç®—ï¼›å¦å‰‡ä½¿ç”¨ç•¶å‰ PnL%
                if mark != entry and entry > 0:
                    # è¨ˆç®—åŸºæ–¼ best çš„ unrealized PnLï¼ˆç›¸å°æ–¼ entryï¼‰
                    if position.side == "LONG":
                        best_unrealized_pnl_ratio = (best - entry) / (mark - entry) if (mark - entry) != 0 else 1.0
                    else:  # SHORT
                        best_unrealized_pnl_ratio = (entry - best) / (entry - mark) if (entry - mark) != 0 else 1.0
                    # åŸºæ–¼ best çš„ PnL% = ç•¶å‰ PnL% * (best ç›¸å°æ–¼ entry çš„ç²åˆ©æ¯”ä¾‹) / (mark ç›¸å°æ–¼ entry çš„ç²åˆ©æ¯”ä¾‹)
                    profit_pct_based_on_best_for_threshold = unrealized_pnl_pct * best_unrealized_pnl_ratio
                else:
                    # å¦‚æœ mark == entryï¼Œä½¿ç”¨ç•¶å‰ PnL%ï¼ˆæ­¤æ™‚ best æ‡‰è©²ä¹Ÿç­‰æ–¼ entryï¼‰
                    profit_pct_based_on_best_for_threshold = unrealized_pnl_pct
            else:
                # Fallbackï¼šä½¿ç”¨åƒ¹æ ¼ç™¾åˆ†æ¯”ï¼ˆåŸºæ–¼æ­·å²æœ€é«˜åƒ¹æ ¼ï¼‰
                profit_pct_for_threshold = profit_pct
                profit_pct_based_on_best_for_threshold = profit_pct_based_on_best
            
            # èª¿è©¦æ—¥èªŒï¼šè¨˜éŒ„åœææ¨¡å¼åˆ¤æ–·çš„é—œéµåƒæ•¸
            logger.info(
                f"compute_stop_state LONG: symbol={getattr(position, 'symbol', 'unknown')}, "
                f"entry={entry:.4f}, best={best:.4f}, mark={mark:.4f}, "
                f"profit_pct(price)={profit_pct:.2f}%, unrealized_pnl_pct={unrealized_pnl_pct}, "
                f"profit_pct_for_threshold={profit_pct_for_threshold:.2f}%, profit_pct_based_on_best={profit_pct_based_on_best:.2f}%, "
                f"profit_threshold_pct={profit_threshold_pct:.2f}%, "
                f"trailing_enabled={trailing_enabled}, effective_trailing_enabled={effective_trailing_enabled}, "
                f"lock_ratio={lock_ratio}, has_override={has_override}, "
                f"trail_callback_override={trail_callback_override}"
            )
            
            # Case 1ï¼šDynamic Trailingï¼ˆé–åˆ©ï¼‰
            # é—œéµï¼šä¸€æ—¦ best æ›¾ç¶“é”åˆ°é thresholdï¼Œå°±æ‡‰è©²ä¿æŒåœ¨ dynamic mode
            # ä½¿ç”¨åŸºæ–¼ best çš„ profit_pct ä¾†åˆ¤æ–·ï¼Œè€Œä¸æ˜¯ç•¶å‰åƒ¹æ ¼
            # é€™æ¨£å³ä½¿ç•¶å‰åƒ¹æ ¼ä¸‹è·Œï¼Œåªè¦ best æ›¾ç¶“é”åˆ°é thresholdï¼Œå°±æœƒä¿æŒåœ¨ dynamic mode
            if (
                effective_trailing_enabled
                and lock_ratio is not None
                and profit_pct_based_on_best_for_threshold >= profit_threshold_pct
            ):
                # dynamic_stop_price åŸºæ–¼ bestï¼ˆæ­·å²æœ€é«˜åƒ¹æ ¼ï¼‰è¨ˆç®—ï¼Œæ°¸é ä¸æœƒä¸‹é™
                dynamic_stop_price = entry + (best - entry) * lock_ratio
                stop_mode = "dynamic"
                logger.info(f"âœ“ é€²å…¥ Dynamic æ¨¡å¼: dynamic_stop_price={dynamic_stop_price:.4f}, best={best:.4f}, entry={entry:.4f}, lock_ratio={lock_ratio}")
            else:
                # è¨˜éŒ„ç‚ºä»€éº¼æ²’æœ‰é€²å…¥ dynamic mode
                reasons = []
                if not effective_trailing_enabled:
                    reasons.append(f"effective_trailing_enabled=False (trailing_enabled={trailing_enabled}, has_override={has_override})")
                if lock_ratio is None:
                    reasons.append(f"lock_ratio=None (trail_callback_override={trail_callback_override})")
                if profit_pct_for_threshold < profit_threshold_pct:
                    reasons.append(f"profit_pct_for_threshold({profit_pct_for_threshold:.2f}%) < threshold({profit_threshold_pct:.2f}%)")
                logger.info(f"âœ— æœªé€²å…¥ Dynamic æ¨¡å¼: {', '.join(reasons)}")
            
            # Case 2ï¼šBase Stop-Lossï¼ˆåªæœ‰ç•¶å¾æœªé€²å…¥é dynamic mode æ™‚æ‰é¡¯ç¤º base stopï¼‰
            # ä¸€æ—¦é€²å…¥ dynamic modeï¼Œå°±ä¸æœƒè¿”å› base mode
            # å¦‚æœæœ‰è¦†å¯«å€¼ï¼Œå³ä½¿å¾æœªé”åˆ° thresholdï¼Œä¹Ÿæ‡‰è©²é¡¯ç¤º base stop
            if stop_mode != "dynamic" and base_sl_pct > 0 and (profit_pct_based_on_best_for_threshold < profit_threshold_pct or has_override):
                stop_mode = "base"
                logger.debug(f"ä¿æŒåœ¨ Base æ¨¡å¼: profit_pct_based_on_best={profit_pct_based_on_best:.2f}% < threshold={profit_threshold_pct:.2f}% æˆ– has_override={has_override}")
            
            return StopState(
                stop_mode=stop_mode,
                base_stop_price=base_stop_price,
                dynamic_stop_price=dynamic_stop_price
            )
        
        # è™•ç† SHORT å€‰ä½
        elif position.side == "SHORT":
            # å¦‚æœ best ç‚º Noneï¼Œä½¿ç”¨ç•¶å‰åƒ¹æ ¼
            if best is None:
                best = mark
            
            # æ›´æ–° bestï¼ˆåƒ…ç”¨æ–¼è¨ˆç®—ï¼Œä¸ä¿®æ”¹ DBï¼‰
            # é‡è¦ï¼šå°æ–¼ SHORTï¼Œbest æ˜¯æœ€ä½åƒ¹æ ¼ï¼Œåªèƒ½ä¸‹é™ï¼Œä¸èƒ½ä¸Šå‡
            # é€™æ¨£ dynamic stop æ‰èƒ½ä¿æŒç©©å®šï¼šåŸºæ–¼æ­·å²æœ€ä½åƒ¹æ ¼è¨ˆç®—åœæ
            if mark < best:
                best = mark
            
            # profit_pct åŸºæ–¼æ­·å²æœ€ä½åƒ¹æ ¼ï¼ˆbestï¼‰è¨ˆç®—ï¼Œè€Œä¸æ˜¯ç•¶å‰åƒ¹æ ¼ï¼ˆmarkï¼‰
            # é€™æ¨£å³ä½¿ç•¶å‰åƒ¹æ ¼ä¸Šæ¼²ï¼Œåªè¦æ­·å²æœ€é«˜ profit_pct >= thresholdï¼Œå°±æœƒä¿æŒåœ¨ dynamic mode
            profit_pct = (entry - best) / entry * 100.0 if entry > 0 else 0.0
            
            # è¨ˆç®— base stop priceï¼ˆåŸºæ–¼ marginï¼‰
            # æ ¹æ“šéœ€æ±‚ï¼šStop price = best + (Margin * Base SL% / 100) / qty
            # Margin = (Entry Price * Qty) / Leverage
            # Base SL Amount (USDT) = Margin * (Base SL% / 100.0)
            # å°æ–¼ SHORT: stop_price = best + (margin * base_sl_pct / 100) / qty
            base_stop_price = None
            if base_sl_pct > 0 and entry > 0 and position_qty > 0 and position_leverage > 0:
                # è¨ˆç®— marginï¼ˆä½¿ç”¨ entry price è¨ˆç®— marginï¼‰
                notional = entry * position_qty
                margin = notional / position_leverage
                # ä½¿ç”¨ best åƒ¹æ ¼ï¼ˆå¦‚æœ best ç‚º Noneï¼Œä½¿ç”¨ entryï¼‰
                best_price = best if best is not None else entry
                # Base stop price = best + (margin * base_sl_pct / 100) / qty
                base_stop_price = best_price + (margin * base_sl_pct / 100.0) / position_qty
            
            # è¨ˆç®— dynamic stop price
            dynamic_stop_price = None
            stop_mode = "none"
            
            # æª¢æŸ¥æ˜¯å¦æœ‰è¦†å¯«å€¼ï¼ˆç”¨æ–¼æ±ºå®šæ˜¯å¦å•Ÿç”¨åœæï¼‰
            dyn_profit_threshold_pct_override = getattr(position, 'dyn_profit_threshold_pct', None)
            base_stop_loss_pct_override = getattr(position, 'base_stop_loss_pct', None)
            has_override = (
                trail_callback_override is not None or 
                dyn_profit_threshold_pct_override is not None or 
                base_stop_loss_pct_override is not None
            )
            # å¦‚æœæœ‰è¦†å¯«å€¼ï¼Œå³ä½¿å…¨å±€ trailing_enabled ç‚º Falseï¼Œä¹Ÿæ‡‰è©²å•Ÿç”¨åœæ
            effective_trailing_enabled = trailing_enabled or (has_override and lock_ratio is not None)
            
            # é—œéµé‚è¼¯ï¼šä¸€æ—¦é€²å…¥ dynamic modeï¼Œå°±æ‡‰è©²ä¿æŒåœ¨ dynamic mode
            # åˆ¤æ–·æ˜¯å¦æ‡‰è©²é€²å…¥æˆ–ä¿æŒåœ¨ dynamic modeï¼š
            # 1. å¦‚æœ bestï¼ˆæ­·å²æœ€ä½åƒ¹æ ¼ï¼‰æ›¾ç¶“é”åˆ°é thresholdï¼Œå°±æ‡‰è©²ä¿æŒåœ¨ dynamic mode
            # 2. ä½¿ç”¨åŸºæ–¼ best çš„ PnL% ä¾†åˆ¤æ–·ï¼Œè€Œä¸æ˜¯ç•¶å‰åƒ¹æ ¼
            
            # è¨ˆç®—åŸºæ–¼ bestï¼ˆæ­·å²æœ€ä½åƒ¹æ ¼ï¼‰çš„ profit_pctï¼ˆåƒ¹æ ¼ç™¾åˆ†æ¯”ï¼‰
            profit_pct_based_on_best = profit_pct  # å·²ç¶“åŸºæ–¼ best è¨ˆç®—
            
            # å¦‚æœæä¾›äº† unrealized_pnl_pctï¼Œæˆ‘å€‘éœ€è¦è¨ˆç®—åŸºæ–¼ best çš„ PnL%
            # åŸºæ–¼ best çš„ PnL% = (entry - best) / entry * 100 * (entry * qty / leverage) / (entry * qty / leverage)
            # ç°¡åŒ–å¾Œï¼šåŸºæ–¼ best çš„ PnL% = (entry - best) / entry * 100 * leverage / leverage
            # å¯¦éš›ä¸Šï¼ŒPnL% å’Œåƒ¹æ ¼ç™¾åˆ†æ¯”çš„æ¯”ä¾‹æ˜¯å›ºå®šçš„ï¼ˆéƒ½åŸºæ–¼ç›¸åŒçš„ marginï¼‰
            # æ‰€ä»¥ï¼šåŸºæ–¼ best çš„ PnL% = unrealized_pnl_pct * (entry - best) / (entry - mark)
            if unrealized_pnl_pct is not None:
                # ä½¿ç”¨ç•¶å‰ PnL% ä¾†åˆ¤æ–·æ˜¯å¦é€²å…¥ dynamic mode
                profit_pct_for_threshold = unrealized_pnl_pct
                # è¨ˆç®—åŸºæ–¼ best çš„ PnL%ï¼Œç”¨æ–¼åˆ¤æ–·æ˜¯å¦æ‡‰è©²ä¿æŒåœ¨ dynamic mode
                # å¦‚æœ mark != entryï¼Œä½¿ç”¨æ¯”ä¾‹è¨ˆç®—ï¼›å¦å‰‡ä½¿ç”¨ç•¶å‰ PnL%
                if mark != entry and entry > 0:
                    # è¨ˆç®—åŸºæ–¼ best çš„ unrealized PnLï¼ˆç›¸å°æ–¼ entryï¼‰
                    # å°æ–¼ SHORTï¼šbest æ˜¯æœ€ä½åƒ¹æ ¼ï¼Œç²åˆ© = entry - best
                    best_unrealized_pnl_ratio = (entry - best) / (entry - mark) if (entry - mark) != 0 else 1.0
                    # åŸºæ–¼ best çš„ PnL% = ç•¶å‰ PnL% * (best ç›¸å°æ–¼ entry çš„ç²åˆ©æ¯”ä¾‹) / (mark ç›¸å°æ–¼ entry çš„ç²åˆ©æ¯”ä¾‹)
                    profit_pct_based_on_best_for_threshold = unrealized_pnl_pct * best_unrealized_pnl_ratio
                else:
                    # å¦‚æœ mark == entryï¼Œä½¿ç”¨ç•¶å‰ PnL%ï¼ˆæ­¤æ™‚ best æ‡‰è©²ä¹Ÿç­‰æ–¼ entryï¼‰
                    profit_pct_based_on_best_for_threshold = unrealized_pnl_pct
            else:
                # Fallbackï¼šä½¿ç”¨åƒ¹æ ¼ç™¾åˆ†æ¯”ï¼ˆåŸºæ–¼æ­·å²æœ€ä½åƒ¹æ ¼ï¼‰
                profit_pct_for_threshold = profit_pct
                profit_pct_based_on_best_for_threshold = profit_pct_based_on_best
            
            # èª¿è©¦æ—¥èªŒï¼šè¨˜éŒ„åœææ¨¡å¼åˆ¤æ–·çš„é—œéµåƒæ•¸
            logger.info(
                f"compute_stop_state SHORT: symbol={getattr(position, 'symbol', 'unknown')}, "
                f"entry={entry:.4f}, best={best:.4f}, mark={mark:.4f}, "
                f"profit_pct(price)={profit_pct:.2f}%, unrealized_pnl_pct={unrealized_pnl_pct}, "
                f"profit_pct_for_threshold={profit_pct_for_threshold:.2f}%, profit_pct_based_on_best={profit_pct_based_on_best:.2f}%, "
                f"profit_threshold_pct={profit_threshold_pct:.2f}%, "
                f"trailing_enabled={trailing_enabled}, effective_trailing_enabled={effective_trailing_enabled}, "
                f"lock_ratio={lock_ratio}, has_override={has_override}, "
                f"trail_callback_override={trail_callback_override}"
            )
            
            # Case 1ï¼šDynamic Trailingï¼ˆé–åˆ©ï¼‰
            # é—œéµï¼šä¸€æ—¦ best æ›¾ç¶“é”åˆ°é thresholdï¼Œå°±æ‡‰è©²ä¿æŒåœ¨ dynamic mode
            # ä½¿ç”¨åŸºæ–¼ best çš„ profit_pct ä¾†åˆ¤æ–·ï¼Œè€Œä¸æ˜¯ç•¶å‰åƒ¹æ ¼
            # é€™æ¨£å³ä½¿ç•¶å‰åƒ¹æ ¼ä¸Šæ¼²ï¼Œåªè¦ best æ›¾ç¶“é”åˆ°é thresholdï¼Œå°±æœƒä¿æŒåœ¨ dynamic mode
            if (
                effective_trailing_enabled
                and lock_ratio is not None
                and profit_pct_based_on_best_for_threshold >= profit_threshold_pct
            ):
                # dynamic_stop_price åŸºæ–¼ bestï¼ˆæ­·å²æœ€ä½åƒ¹æ ¼ï¼‰è¨ˆç®—ï¼Œæ°¸é ä¸æœƒä¸Šå‡
                dynamic_stop_price = entry - (entry - best) * lock_ratio
                stop_mode = "dynamic"
                logger.info(f"âœ“ é€²å…¥ Dynamic æ¨¡å¼: dynamic_stop_price={dynamic_stop_price:.4f}, best={best:.4f}, entry={entry:.4f}, lock_ratio={lock_ratio}")
            else:
                # è¨˜éŒ„ç‚ºä»€éº¼æ²’æœ‰é€²å…¥ dynamic mode
                reasons = []
                if not effective_trailing_enabled:
                    reasons.append(f"effective_trailing_enabled=False (trailing_enabled={trailing_enabled}, has_override={has_override})")
                if lock_ratio is None:
                    reasons.append(f"lock_ratio=None (trail_callback_override={trail_callback_override})")
                if profit_pct_for_threshold < profit_threshold_pct:
                    reasons.append(f"profit_pct_for_threshold({profit_pct_for_threshold:.2f}%) < threshold({profit_threshold_pct:.2f}%)")
                logger.info(f"âœ— æœªé€²å…¥ Dynamic æ¨¡å¼: {', '.join(reasons)}")
            
            # Case 2ï¼šBase Stop-Lossï¼ˆåªæœ‰ç•¶å¾æœªé€²å…¥é dynamic mode æ™‚æ‰é¡¯ç¤º base stopï¼‰
            # ä¸€æ—¦é€²å…¥ dynamic modeï¼Œå°±ä¸æœƒè¿”å› base mode
            # å¦‚æœæœ‰è¦†å¯«å€¼ï¼Œå³ä½¿å¾æœªé”åˆ° thresholdï¼Œä¹Ÿæ‡‰è©²é¡¯ç¤º base stop
            if stop_mode != "dynamic" and base_sl_pct > 0 and (profit_pct_based_on_best_for_threshold < profit_threshold_pct or has_override):
                stop_mode = "base"
            
            return StopState(
                stop_mode=stop_mode,
                base_stop_price=base_stop_price,
                dynamic_stop_price=dynamic_stop_price
            )
        
        # æœªçŸ¥æ–¹å‘
        else:
            return StopState(
                stop_mode="none",
                base_stop_price=None,
                dynamic_stop_price=None
            )
    
    except Exception as e:
        # å®‰å…¨åœ°ç²å– position.idï¼ˆTempPosition å¯èƒ½æ²’æœ‰ id å±¬æ€§ï¼‰
        pos_id = getattr(position, 'id', None)
        pos_symbol = getattr(position, 'symbol', 'unknown')
        if pos_id:
            logger.warning(f"è¨ˆç®—å€‰ä½ {pos_id} ({pos_symbol}) åœæç‹€æ…‹æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        else:
            logger.warning(f"è¨ˆç®—å€‰ä½ ({pos_symbol}) åœæç‹€æ…‹æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        return StopState(
            stop_mode="none",
            base_stop_price=None,
            dynamic_stop_price=None
        )


async def check_trailing_stop(position: Position, db: Session):
    """
    æª¢æŸ¥å–®ä¸€å€‰ä½çš„ Dynamic Stopï¼ˆå‹•æ…‹åœæï¼‰
    
    ä½¿ç”¨ dynamic stop é‚è¼¯ï¼š
    1. ç•¶ PnL% é”åˆ°é–€æª»æ™‚ï¼Œé–ä½ä¸€éƒ¨åˆ†åˆ©æ½¤ä½œç‚ºåœæç·š
    2. å¦å‰‡ä½¿ç”¨ base stop-lossï¼ˆå›ºå®šç™¾åˆ†æ¯”åœæï¼‰
    
    æ³¨æ„ï¼šæ­¤å‡½æ•¸åªæœƒåœ¨ bot_stop_loss_enabled=True æ™‚è¢«èª¿ç”¨ã€‚
    
    Args:
        position: Position æ¨¡å‹å¯¦ä¾‹
        db: è³‡æ–™åº« Session
    """
    # é›™é‡æª¢æŸ¥ï¼šå¦‚æœ bot_stop_loss_enabled ç‚º Falseï¼Œç›´æ¥è¿”å›
    if not position.bot_stop_loss_enabled:
        logger.debug(f"å€‰ä½ {position.id} ({position.symbol}) bot_stop_loss_enabled=Falseï¼Œè·³é Bot åœææª¢æŸ¥")
        return
    
    try:
        # å–å¾—ç›®å‰æ¨™è¨˜åƒ¹æ ¼
        current_price = get_mark_price(position.symbol)
        
        # è¨ˆç®— dynamic stop æ‰€éœ€çš„å…±ç”¨è®Šæ•¸
        entry = position.entry_price
        best = position.highest_price  # LONG: æœ€é«˜åƒ¹; SHORT: æœ€ä½åƒ¹
        mark = current_price
        
        # è‹¥ entry <= 0ï¼Œå˜—è©¦å¾ Binance æŸ¥è©¢å¯¦éš›çš„ entry price
        if entry <= 0:
            logger.warning(
                f"å€‰ä½ {position.id} ({position.symbol}) entry_price={entry} ç„¡æ•ˆï¼Œ"
                f"å˜—è©¦å¾ Binance æŸ¥è©¢å¯¦éš› entry price"
            )
            try:
                client = get_client()
                positions_info = client.futures_position_information(symbol=position.symbol)
                for pos_info in positions_info:
                    position_amt = float(pos_info.get("positionAmt", "0") or 0)
                    if abs(position_amt) < 1e-8:
                        continue
                    
                    # åˆ¤æ–·æ–¹å‘æ˜¯å¦åŒ¹é…
                    side_local = "LONG" if position_amt > 0 else "SHORT"
                    if side_local != position.side:
                        continue
                    
                    # å–å¾— Binance çš„ entry price
                    binance_entry = float(pos_info.get("entryPrice", "0") or 0)
                    if binance_entry > 0:
                        logger.info(
                            f"å¾ Binance å–å¾—å€‰ä½ {position.id} ({position.symbol}) çš„ entry_price: {binance_entry}"
                        )
                        position.entry_price = binance_entry
                        entry = binance_entry
                        # å¦‚æœ highest_price ä¹Ÿç„¡æ•ˆï¼Œä½¿ç”¨ç•¶å‰ mark_price åˆå§‹åŒ–
                        if best is None or best <= 0:
                            position.highest_price = current_price
                            best = current_price
                        db.commit()
                        break
            except Exception as e:
                logger.error(
                    f"å¾ Binance æŸ¥è©¢å€‰ä½ {position.id} ({position.symbol}) entry_price å¤±æ•—: {e}"
                )
                # å¦‚æœæŸ¥è©¢å¤±æ•—ï¼Œä»ç„¶è·³éåœææª¢æŸ¥
                return
        
        # å¦‚æœæ›´æ–°å¾Œ entry ä»ç„¶ <= 0ï¼Œè·³éåœææª¢æŸ¥
        if entry <= 0:
            logger.warning(f"å€‰ä½ {position.id} ({position.symbol}) entry_price={entry} ä»ç„¶ç„¡æ•ˆï¼Œè·³éåœææª¢æŸ¥")
            return
        
        # è¨ˆç®— unrealized_pnl_pctï¼ˆPnL%ï¼‰ç”¨æ–¼åˆ¤æ–·æ˜¯å¦é€²å…¥ dynamic mode
        # PnL% = (Unrealized PnL / ä¿è­‰é‡‘) * 100
        # ä¿è­‰é‡‘ = åç¾©åƒ¹å€¼ / æ æ¡¿ = (Entry Price * Qty) / Leverage
        calculated_unrealized_pnl_pct = None
        if entry > 0 and position.qty > 0:
            # è¨ˆç®— unrealized PnL
            if position.side == "LONG":
                unrealized_pnl_amount = (mark - entry) * position.qty
            else:  # SHORT
                unrealized_pnl_amount = (entry - mark) * position.qty
            # ç²å– leverageï¼ˆå¾é—œè¯çš„ Bot æˆ–ä½¿ç”¨é»˜èªå€¼ï¼‰
            leverage = 20  # é»˜èªæ æ¡¿
            if position.bot_id:
                try:
                    from models import BotConfig
                    bot = db.query(BotConfig).filter(BotConfig.id == position.bot_id).first()
                    if bot and bot.leverage:
                        leverage = bot.leverage
                except Exception as e:
                    logger.debug(f"ç„¡æ³•å–å¾—å€‰ä½ {position.id} çš„ Bot leverage: {e}")
            # è¨ˆç®—ä¿è­‰é‡‘
            notional = entry * position.qty
            if notional > 0 and leverage > 0:
                margin = notional / leverage
                if margin > 0:
                    calculated_unrealized_pnl_pct = (unrealized_pnl_amount / margin) * 100.0
        
        # ç²å– leverage å’Œ qtyï¼ˆç”¨æ–¼ margin-based base stop è¨ˆç®—ï¼‰
        leverage_for_stop = 20  # é»˜èªæ æ¡¿
        qty_for_stop = getattr(position, 'qty', 0)
        if position.bot_id:
            try:
                from models import BotConfig
                bot = db.query(BotConfig).filter(BotConfig.id == position.bot_id).first()
                if bot and bot.leverage:
                    leverage_for_stop = bot.leverage
            except Exception as e:
                logger.debug(f"ç„¡æ³•å–å¾—å€‰ä½ {position.id} çš„ Bot leverage: {e}")
        
        # ä½¿ç”¨ compute_stop_state è¨ˆç®—åœæç‹€æ…‹ï¼ˆå‚³å…¥ leverage å’Œ qtyï¼‰
        stop_state = compute_stop_state(position, current_price, calculated_unrealized_pnl_pct, leverage_for_stop, qty_for_stop)
        
        # æ ¹æ“šå€‰ä½æ–¹å‘å–å¾—å°æ‡‰çš„å…¨å±€è¨­å®š
        side_config = TRAILING_CONFIG.get_config_for_side(position.side)
        trailing_enabled = TRAILING_CONFIG.trailing_enabled if TRAILING_CONFIG.trailing_enabled is not None else DYN_TRAILING_ENABLED
        
        # ä½¿ç”¨å°æ‡‰æ–¹å‘çš„å…¨å±€è¨­å®šä½œç‚ºé»˜èªå€¼ï¼Œå¦‚æœæ²’æœ‰å‰‡ä½¿ç”¨ç’°å¢ƒè®Šæ•¸
        base_sl_pct_default = side_config.base_sl_pct if side_config.base_sl_pct is not None else DYN_BASE_SL_PCT
        profit_threshold_pct_default = side_config.profit_threshold_pct if side_config.profit_threshold_pct is not None else DYN_PROFIT_THRESHOLD_PCT
        
        # å„ªå…ˆä½¿ç”¨å€‰ä½è¦†å¯«å€¼ï¼Œå¦‚æœæ²’æœ‰å‰‡ä½¿ç”¨å…¨å±€é…ç½®
        if position.base_stop_loss_pct is not None:
            base_sl_pct = position.base_stop_loss_pct
        else:
            base_sl_pct = base_sl_pct_default
        
        if position.dyn_profit_threshold_pct is not None:
            profit_threshold_pct = position.dyn_profit_threshold_pct
        else:
            profit_threshold_pct = profit_threshold_pct_default
        
        # å…ˆæ±ºå®šé€™ç­†å–®ä½¿ç”¨çš„ lock_ratio
        # trail_callback: null â†’ ä½¿ç”¨å…¨å±€é…ç½®, 0 â†’ base stop only, >0 â†’ ä½¿ç”¨è©²å€¼ä½œç‚º lock_ratio
        if position.trail_callback is None:
            # ä½¿ç”¨å°æ‡‰æ–¹å‘çš„ TRAILING_CONFIG lock_ratioï¼ˆå¦‚æœæœ‰çš„è©±ï¼‰ï¼Œå¦å‰‡ä½¿ç”¨é è¨­å€¼
            lock_ratio = side_config.lock_ratio if side_config.lock_ratio is not None else DYN_LOCK_RATIO_DEFAULT
        elif position.trail_callback == 0:
            logger.info(
                f"å€‰ä½ {position.id} ({position.symbol}) trail_callback=0ï¼Œåƒ…ä½¿ç”¨ base stop-loss"
            )
            lock_ratio = None
        else:
            # æ­£å¸¸æƒ…æ³ï¼Œä½¿ç”¨æ¯ç­†å€‰ä½è‡ªå·±çš„ lock_ratio
            lock_ratio = position.trail_callback
        
        # è¨˜éŒ„ä½¿ç”¨çš„åœæé…ç½®å€¼ï¼ˆç”¨æ–¼èª¿è©¦å’Œé©—è­‰ï¼‰
        logger.info(
            f"[DynamicStop] pos_id={position.id} symbol={position.symbol} "
            f"profit_threshold={profit_threshold_pct}% "
            f"(override={position.dyn_profit_threshold_pct if position.dyn_profit_threshold_pct is not None else 'global'}) "
            f"lock_ratio={lock_ratio if lock_ratio is not None else 'base-only'} "
            f"(override={position.trail_callback if position.trail_callback is not None else 'global'}) "
            f"base_sl={base_sl_pct}% "
            f"(override={position.base_stop_loss_pct if position.base_stop_loss_pct is not None else 'global'})"
        )
        
        # ä¹‹å¾Œå†åšç¯„åœé˜²å‘†ï¼ˆåªå° >0 çš„ lock_ratio åš clampï¼‰
        if lock_ratio is not None:
            if lock_ratio <= 0:
                logger.warning(
                    f"å€‰ä½ {position.id} ({position.symbol}) lock_ratio <= 0ï¼ˆå€¼={lock_ratio}ï¼‰ï¼Œå¿½ç•¥ dynamicï¼Œæ”¹ç”¨ base stop"
                )
                lock_ratio = None
            elif lock_ratio > 1:
                logger.warning(
                    f"å€‰ä½ {position.id} ({position.symbol}) lock_ratio > 1ï¼ˆå€¼={lock_ratio}ï¼‰ï¼Œå·²å¼·åˆ¶èª¿æ•´ç‚º 1.0"
                )
                lock_ratio = 1.0
        
        # è™•ç† LONG å€‰ä½
        if position.side == "LONG":
            # æ›´æ–° highest_priceï¼ˆè¨˜éŒ„æœ€é«˜åƒ¹æ ¼ï¼‰
            if best is None:
                # å¦‚æœé‚„æ²’æœ‰è¨­å®šæœ€é«˜åƒ¹ï¼Œè¨­å®šç‚ºç›®å‰åƒ¹æ ¼
                position.highest_price = current_price
                logger.info(f"å€‰ä½ {position.id} ({position.symbol}) LONG åˆå§‹åŒ–æœ€é«˜åƒ¹: {position.highest_price}")
                db.commit()
                return
            elif current_price > best:
                position.highest_price = current_price
                best = current_price  # æ›´æ–° best è®Šæ•¸
                logger.info(f"å€‰ä½ {position.id} ({position.symbol}) LONG æ›´æ–°æœ€é«˜åƒ¹: {best}")
            
            # è¨ˆç®— unrealized_pnl_pctï¼ˆPnL%ï¼‰ç”¨æ–¼åˆ¤æ–·æ˜¯å¦é€²å…¥ dynamic mode
            # PnL% = (Unrealized PnL / ä¿è­‰é‡‘) * 100
            # ä¿è­‰é‡‘ = åç¾©åƒ¹å€¼ / æ æ¡¿ = (Entry Price * Qty) / Leverage
            calculated_unrealized_pnl_pct = None
            if entry > 0 and position.qty > 0:
                # è¨ˆç®— unrealized PnL
                unrealized_pnl_amount = (mark - entry) * position.qty
                # ç²å– leverageï¼ˆå¾é—œè¯çš„ Bot æˆ–ä½¿ç”¨é»˜èªå€¼ï¼‰
                leverage = 20  # é»˜èªæ æ¡¿
                if position.bot_id:
                    try:
                        from models import BotConfig
                        bot = db.query(BotConfig).filter(BotConfig.id == position.bot_id).first()
                        if bot and bot.leverage:
                            leverage = bot.leverage
                    except Exception as e:
                        logger.debug(f"ç„¡æ³•å–å¾—å€‰ä½ {position.id} çš„ Bot leverage: {e}")
                # è¨ˆç®—ä¿è­‰é‡‘
                notional = entry * position.qty
                if notional > 0 and leverage > 0:
                    margin = notional / leverage
                    if margin > 0:
                        calculated_unrealized_pnl_pct = (unrealized_pnl_amount / margin) * 100.0
            
            # å…ˆæ›´æ–°å®Œ best (= position.highest_price) ä¹‹å¾Œï¼Œä½¿ç”¨ compute_stop_state è¨ˆç®—åœæç‹€æ…‹ï¼ˆå‚³å…¥ leverage å’Œ qtyï¼‰
            stop_state = compute_stop_state(position, mark, calculated_unrealized_pnl_pct, leverage, position.qty)
            
            # å¾ stop_state å–å¾—åœæåƒ¹æ ¼å’Œæ¨¡å¼
            # æ ¹æ“š stop_mode é¸æ“‡å°æ‡‰çš„åœæåƒ¹æ ¼
            triggered = False
            mode = None
            dyn_stop = None
            
            # åˆ¤æ–·æ˜¯å¦è§¸ç™¼åœæ
            if stop_state.stop_mode == "dynamic":
                # Dynamic mode: åªä½¿ç”¨ dynamic_stop_price
                dyn_stop = stop_state.dynamic_stop_price
                if dyn_stop is not None:
                    # LONG: ç•¶åƒ¹æ ¼ä¸‹è·Œåˆ° dynamic_stop_price ä»¥ä¸‹æ™‚è§¸ç™¼
                    triggered = mark <= dyn_stop
                    mode = "dynamic_trailing"
            elif stop_state.stop_mode == "base":
                # Base mode: åªä½¿ç”¨ base_stop_price
                dyn_stop = stop_state.base_stop_price
                if dyn_stop is not None:
                    # LONG: ç•¶åƒ¹æ ¼ä¸‹è·Œåˆ° base_stop_price ä»¥ä¸‹æ™‚è§¸ç™¼
                    triggered = mark <= dyn_stop
                    mode = "base_stop"
                else:
                    logger.warning(
                        f"å€‰ä½ {position.id} ({position.symbol}) LONG base mode ä½† base_stop_price ç‚º None"
                    )
            
            profit_pct = (best - entry) / entry * 100.0 if entry > 0 else 0.0
            
            # å¦‚æœæœ‰è¨ˆç®—å‡º dyn_stopï¼Œå°±å¯« info log
            if dyn_stop is not None:
                logger.info(
                    f"å€‰ä½ {position.id} ({position.symbol}) LONG æ¨¡å¼: {mode}, "
                    f"ç›®å‰åƒ¹æ ¼: {mark:.6f}, best: {best:.6f}, dyn_stop: {dyn_stop:.6f}, "
                    f"ç²åˆ©%: {profit_pct:.2f}, lock_ratio: {lock_ratio}, base_sl_pct: {base_sl_pct}, "
                    f"è§¸ç™¼æ¢ä»¶: mark <= dyn_stop ({mark:.6f} <= {dyn_stop:.6f} = {mark <= dyn_stop}), "
                    f"triggered={triggered}, stop_mode={stop_state.stop_mode}, "
                    f"dynamic_stop_price={stop_state.dynamic_stop_price}, base_stop_price={stop_state.base_stop_price}"
                )
            else:
                logger.warning(
                    f"å€‰ä½ {position.id} ({position.symbol}) LONG æ²’æœ‰åœæåƒ¹æ ¼ï¼"
                    f"stop_mode={stop_state.stop_mode}, "
                    f"dynamic_stop_price={stop_state.dynamic_stop_price}, "
                    f"base_stop_price={stop_state.base_stop_price}, "
                    f"triggered={triggered}, mode={mode}"
                )
            
            if triggered:
                logger.info(
                    f"å€‰ä½ {position.id} ({position.symbol}) LONG è§¸ç™¼ {mode}ï¼Œ"
                    f"ç›®å‰åƒ¹æ ¼: {mark}, åœæç·š: {dyn_stop}, ç²åˆ©%: {profit_pct:.2f}"
                )
                
                # auto_close_enabled å§‹çµ‚å•Ÿç”¨ï¼ˆå¼·åˆ¶ï¼‰
                # å‘¼å«é—œå€‰å‡½å¼
                try:
                    close_order = close_futures_position(
                        symbol=position.symbol,
                        position_side=position.side,  # "LONG"
                        qty=position.qty,
                        position_id=position.id
                    )
                    
                    # å–å¾—å¹³å€‰åƒ¹æ ¼
                    exit_price = get_exit_price_from_order(close_order, position.symbol)
                    
                    # æ›´æ–°å€‰ä½ç‹€æ…‹èˆ‡å¹³å€‰è³‡è¨Š
                    position.status = "CLOSED"
                    position.closed_at = datetime.now(timezone.utc)
                    position.exit_price = exit_price
                    # exit_reasonï¼šdynamic_trailing or base_stop
                    position.exit_reason = "dynamic_stop" if mode == "dynamic_trailing" else "base_stop"
                    db.commit()
                    
                    logger.info(
                        f"å€‰ä½ {position.id} ({position.symbol}) LONG å·²æˆåŠŸé—œå€‰ï¼ˆ{mode}ï¼‰ï¼Œ"
                        f"è¨‚å–® ID: {close_order.get('orderId')}, å¹³å€‰åƒ¹æ ¼: {exit_price}"
                    )
                
                except Exception as e:
                    logger.error(f"é—œé–‰å€‰ä½ {position.id} å¤±æ•—: {e}")
                    position.status = "ERROR"
                    db.commit()
                    raise
        
        # è™•ç† SHORT å€‰ä½
        elif position.side == "SHORT":
            # å°æ–¼ SHORTï¼Œhighest_price æ¬„ä½ç”¨ä¾†è¨˜éŒ„æœ€ä½åƒ¹æ ¼ï¼ˆlowest_priceï¼‰
            if best is None:
                # å¦‚æœé‚„æ²’æœ‰è¨­å®šæœ€ä½åƒ¹ï¼Œè¨­å®šç‚ºç›®å‰åƒ¹æ ¼ï¼ˆç¬¬ä¸€æ¬¡è¨˜éŒ„ï¼‰
                position.highest_price = current_price
                best = current_price  # æ›´æ–° best è®Šæ•¸
                logger.info(f"å€‰ä½ {position.id} ({position.symbol}) SHORT åˆå§‹åŒ–æœ€ä½åƒ¹: {best}")
                db.commit()
                return
            elif current_price < best:
                # å¦‚æœç›®å‰åƒ¹æ ¼æ›´ä½ï¼Œæ›´æ–°ç‚ºæ–°çš„æœ€ä½åƒ¹ï¼ˆå‰µæ–°ä½ï¼‰
                position.highest_price = current_price
                best = current_price  # æ›´æ–° best è®Šæ•¸
                logger.info(f"å€‰ä½ {position.id} ({position.symbol}) SHORT æ›´æ–°æœ€ä½åƒ¹: {best}")
            
            profit_pct = (entry - best) / entry * 100.0 if entry > 0 else 0.0
            
            # è¨ˆç®— unrealized_pnl_pctï¼ˆPnL%ï¼‰ç”¨æ–¼åˆ¤æ–·æ˜¯å¦é€²å…¥ dynamic mode
            # PnL% = (Unrealized PnL / ä¿è­‰é‡‘) * 100
            # ä¿è­‰é‡‘ = åç¾©åƒ¹å€¼ / æ æ¡¿ = (Entry Price * Qty) / Leverage
            calculated_unrealized_pnl_pct = None
            if entry > 0 and position.qty > 0:
                # è¨ˆç®— unrealized PnL
                unrealized_pnl_amount = (entry - mark) * position.qty  # SHORT: entry - mark
                # ç²å– leverageï¼ˆå¾é—œè¯çš„ Bot æˆ–ä½¿ç”¨é»˜èªå€¼ï¼‰
                leverage = 20  # é»˜èªæ æ¡¿
                if position.bot_id:
                    try:
                        from models import BotConfig
                        bot = db.query(BotConfig).filter(BotConfig.id == position.bot_id).first()
                        if bot and bot.leverage:
                            leverage = bot.leverage
                    except Exception as e:
                        logger.debug(f"ç„¡æ³•å–å¾—å€‰ä½ {position.id} çš„ Bot leverage: {e}")
                # è¨ˆç®—ä¿è­‰é‡‘
                notional = entry * position.qty
                if notional > 0 and leverage > 0:
                    margin = notional / leverage
                    if margin > 0:
                        calculated_unrealized_pnl_pct = (unrealized_pnl_amount / margin) * 100.0
            
            # ä½¿ç”¨ compute_stop_state è¨ˆç®—åœæç‹€æ…‹ï¼ˆå‚³å…¥ leverage å’Œ qtyï¼‰
            stop_state = compute_stop_state(position, mark, calculated_unrealized_pnl_pct, leverage, position.qty)
            
            # å¾ stop_state å–å¾—åœæåƒ¹æ ¼å’Œæ¨¡å¼
            # æ ¹æ“š stop_mode é¸æ“‡å°æ‡‰çš„åœæåƒ¹æ ¼
            triggered = False
            mode = None
            dyn_stop = None
            
            # åˆ¤æ–·æ˜¯å¦è§¸ç™¼åœæ
            if stop_state.stop_mode == "dynamic":
                # Dynamic mode: åªä½¿ç”¨ dynamic_stop_price
                dyn_stop = stop_state.dynamic_stop_price
                if dyn_stop is not None:
                    # SHORT: ç•¶åƒ¹æ ¼ä¸Šæ¼²åˆ° dynamic_stop_price ä»¥ä¸Šæ™‚è§¸ç™¼
                    triggered = mark >= dyn_stop
                    mode = "dynamic_trailing"
            elif stop_state.stop_mode == "base":
                # Base mode: åªä½¿ç”¨ base_stop_price
                dyn_stop = stop_state.base_stop_price
                if dyn_stop is not None:
                    # SHORT: ç•¶åƒ¹æ ¼ä¸Šæ¼²åˆ° base_stop_price ä»¥ä¸Šæ™‚è§¸ç™¼
                    triggered = mark >= dyn_stop
                    mode = "base_stop"
            
            profit_pct = (entry - best) / entry * 100.0 if entry > 0 else 0.0
            if dyn_stop is not None:
                logger.info(
                    f"å€‰ä½ {position.id} ({position.symbol}) SHORT æ¨¡å¼: {mode}, "
                    f"ç›®å‰åƒ¹æ ¼: {mark:.6f}, æœ€ä½åƒ¹(best): {best:.6f}, dyn_stop: {dyn_stop:.6f}, "
                    f"ç²åˆ©%: {profit_pct:.2f}, lock_ratio: {lock_ratio}, base_sl_pct: {base_sl_pct}, "
                    f"è§¸ç™¼æ¢ä»¶: mark >= dyn_stop ({mark:.6f} >= {dyn_stop:.6f} = {mark >= dyn_stop}), "
                    f"triggered={triggered}, stop_mode={stop_state.stop_mode}, "
                    f"dynamic_stop_price={stop_state.dynamic_stop_price}, base_stop_price={stop_state.base_stop_price}"
                )
            else:
                logger.warning(
                    f"å€‰ä½ {position.id} ({position.symbol}) SHORT æ²’æœ‰åœæåƒ¹æ ¼ï¼"
                    f"stop_mode={stop_state.stop_mode}, "
                    f"dynamic_stop_price={stop_state.dynamic_stop_price}, "
                    f"base_stop_price={stop_state.base_stop_price}"
                )
            
            if triggered:
                logger.info(
                    f"å€‰ä½ {position.id} ({position.symbol}) SHORT è§¸ç™¼ {mode}ï¼Œ"
                    f"ç›®å‰åƒ¹æ ¼: {mark}, åœæç·š: {dyn_stop}, ç²åˆ©%: {profit_pct:.2f}"
                )
                
                # auto_close_enabled å§‹çµ‚å•Ÿç”¨ï¼ˆå¼·åˆ¶ï¼‰
                # å‘¼å«é—œå€‰å‡½å¼
                try:
                    close_order = close_futures_position(
                        symbol=position.symbol,
                        position_side=position.side,  # "SHORT"
                        qty=position.qty,
                        position_id=position.id
                    )
                    
                    # å–å¾—å¹³å€‰åƒ¹æ ¼
                    exit_price = get_exit_price_from_order(close_order, position.symbol)
                    
                    # æ›´æ–°å€‰ä½ç‹€æ…‹èˆ‡å¹³å€‰è³‡è¨Š
                    position.status = "CLOSED"
                    position.closed_at = datetime.now(timezone.utc)
                    position.exit_price = exit_price
                    position.exit_reason = "dynamic_stop" if mode == "dynamic_trailing" else "base_stop"
                    db.commit()
                    
                    logger.info(
                        f"å€‰ä½ {position.id} ({position.symbol}) SHORT å·²æˆåŠŸé—œå€‰ï¼ˆ{mode}ï¼‰ï¼Œ"
                        f"è¨‚å–® ID: {close_order.get('orderId')}, å¹³å€‰åƒ¹æ ¼: {exit_price}"
                    )
                
                except Exception as e:
                    logger.error(f"é—œé–‰å€‰ä½ {position.id} å¤±æ•—: {e}")
                    position.status = "ERROR"
                    db.commit()
                    raise
        
        # å¦‚æœå€‰ä½æ–¹å‘ä¸æ˜¯ LONG æˆ– SHORTï¼Œè¨˜éŒ„è­¦å‘Š
        else:
            logger.warning(f"å€‰ä½ {position.id} ({position.symbol}) æœ‰æœªçŸ¥çš„æ–¹å‘: {position.side}")
            return
        
        db.commit()
    
    except Exception as e:
        logger.error(f"æª¢æŸ¥å€‰ä½ {position.id} è¿½è¹¤åœææ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        raise


@app.on_event("shutdown")
async def shutdown_event():
    """æ‡‰ç”¨ç¨‹å¼é—œé–‰æ™‚åŸ·è¡Œ"""
    global _trailing_worker_running
    _trailing_worker_running = False
    logger.info("è¿½è¹¤åœæèƒŒæ™¯ä»»å‹™å·²åœæ­¢")


# æ‡‰ç”¨ç¨‹å¼å•Ÿå‹•æ™‚åˆå§‹åŒ–è³‡æ–™åº«
@app.on_event("startup")
async def startup_event():
    """æ‡‰ç”¨ç¨‹å¼å•Ÿå‹•æ™‚åŸ·è¡Œ"""
    init_db()
    
    # åˆå§‹åŒ– Portfolio Trailing Configï¼ˆå¦‚æœä¸å­˜åœ¨å‰‡å‰µå»ºé è¨­å€¼ï¼‰
    # id=1 for LONG, id=2 for SHORT
    db = SessionLocal()
    try:
        for config_id, side_name in [(1, "LONG"), (2, "SHORT")]:
            config = db.query(PortfolioTrailingConfig).filter(PortfolioTrailingConfig.id == config_id).first()
            if not config:
                try:
                    # å˜—è©¦å‰µå»ºæ–°è¨˜éŒ„ï¼ˆæ˜ç¢ºæŒ‡å®š idï¼‰
                    config = PortfolioTrailingConfig(
                        id=config_id,
                        enabled=False,
                        target_pnl=None,
                        lock_ratio=None
                    )
                    db.add(config)
                    db.commit()
                    db.refresh(config)
                    logger.info(f"Portfolio Trailing Config ({side_name}) å·²åˆå§‹åŒ–ï¼ˆé è¨­å€¼ï¼‰")
                except Exception as create_error:
                    # å¦‚æœå‰µå»ºå¤±æ•—ï¼ˆå¯èƒ½æ˜¯å› ç‚º id å·²å­˜åœ¨æˆ–å…¶ä»–åŸå› ï¼‰ï¼Œå˜—è©¦æŸ¥è©¢
                    db.rollback()
                    config = db.query(PortfolioTrailingConfig).filter(PortfolioTrailingConfig.id == config_id).first()
                    if config:
                        logger.info(f"Portfolio Trailing Config ({side_name}) å·²å­˜åœ¨: enabled={config.enabled}, target_pnl={config.target_pnl}, lock_ratio={config.lock_ratio}")
                    else:
                        logger.error(f"å‰µå»º Portfolio Trailing Config ({side_name}) å¤±æ•—ï¼Œä¸”æŸ¥è©¢ä¹Ÿå¤±æ•—: {create_error}")
            else:
                logger.info(f"Portfolio Trailing Config ({side_name}) å·²è¼‰å…¥: enabled={config.enabled}, target_pnl={config.target_pnl}, lock_ratio={config.lock_ratio}")
    except Exception as e:
        logger.error(f"åˆå§‹åŒ– Portfolio Trailing Config æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}", exc_info=True)
        try:
            db.rollback()
        except:
            pass
    finally:
        try:
            db.close()
        except:
            pass
    
    # å˜—è©¦åˆå§‹åŒ–å¹£å®‰å®¢æˆ¶ç«¯ï¼ˆæª¢æŸ¥ç’°å¢ƒè®Šæ•¸æ˜¯å¦è¨­å®šï¼‰
    try:
        get_client()
        logger.info("å¹£å®‰å®¢æˆ¶ç«¯åˆå§‹åŒ–æˆåŠŸ")
    except Exception as e:
        logger.warning(f"å¹£å®‰å®¢æˆ¶ç«¯åˆå§‹åŒ–å¤±æ•—: {e}")
        logger.warning("è«‹ç¢ºä¿å·²è¨­å®š BINANCE_API_KEY å’Œ BINANCE_API_SECRET ç’°å¢ƒè®Šæ•¸")
    
    # è¨˜éŒ„ Dynamic Stop è¨­å®šå€¼
    logger.info("Dynamic Stop è¨­å®š:")
    logger.info(f"  DYN_TRAILING_ENABLED: {DYN_TRAILING_ENABLED}")
    logger.info(f"  DYN_PROFIT_THRESHOLD_PCT: {DYN_PROFIT_THRESHOLD_PCT}%")
    logger.info(f"  DYN_LOCK_RATIO_DEFAULT: {DYN_LOCK_RATIO_DEFAULT}")
    logger.info(f"  DYN_BASE_SL_PCT: {DYN_BASE_SL_PCT}%")
    
    # å•Ÿå‹•è¿½è¹¤åœæèƒŒæ™¯ä»»å‹™
    asyncio.create_task(trailing_stop_worker())
    logger.info("è¿½è¹¤åœæèƒŒæ™¯ä»»å‹™å·²åœ¨å•Ÿå‹•äº‹ä»¶ä¸­å»ºç«‹")


# ==================== Pydantic æ¨¡å‹å®šç¾© ====================

class TradingViewSignal(BaseModel):
    """TradingView Webhook è¨Šè™Ÿæ ¼å¼ï¼ˆèˆŠç‰ˆï¼Œä¿ç•™å‘å¾Œå…¼å®¹ï¼‰"""
    secret: str = Field(..., description="Webhook å¯†é‘°ï¼Œç”¨æ–¼é©—è­‰è«‹æ±‚ä¾†æº")
    symbol: str = Field(..., description="äº¤æ˜“å°ï¼Œä¾‹å¦‚ï¼šBTCUSDT")
    side: str = Field(..., description="äº¤æ˜“æ–¹å‘ï¼šBUY æˆ– SELL")
    qty: float = Field(..., description="äº¤æ˜“æ•¸é‡")
    leverage: Optional[int] = Field(None, description="æ æ¡¿å€æ•¸ï¼Œä¾‹å¦‚ï¼š10")
    trailing_callback_percent: Optional[float] = Field(None, description="è¿½è¹¤åœæå›èª¿ç™¾åˆ†æ¯”ï¼Œä¾‹å¦‚ï¼š2.0 ä»£è¡¨ 2%")
    tag: Optional[str] = Field(None, description="å¯é¸çš„æ¨™ç±¤ï¼Œç”¨æ–¼æ¨™è¨˜è¨‚å–®ä¾†æº")


class TradingViewSignalIn(BaseModel):
    """
    TradingView Webhook è¼¸å…¥æ ¼å¼
    
    æ”¯æ´å…©ç¨®æ¨¡å¼ï¼š
    1. æ–°æ ¼å¼ï¼ˆæ¨è–¦ï¼‰ï¼šä½¿ç”¨ signal_keyï¼Œç³»çµ±æœƒæŸ¥æ‰¾æ‰€æœ‰ enabled=True ä¸” signal_id åŒ¹é…çš„ bots
    2. èˆŠæ ¼å¼ï¼ˆå…¼å®¹ï¼‰ï¼šä½¿ç”¨ bot_keyï¼Œç›´æ¥æŸ¥æ‰¾å°æ‡‰çš„ bot
    
    è‹¥åŒæ™‚æä¾› signal_key å’Œ bot_keyï¼Œå„ªå…ˆä½¿ç”¨ signal_keyã€‚
    
    ä½ç½®å°å‘æ¨¡å¼ï¼ˆposition-basedï¼‰ï¼š
    - å¦‚æœæä¾› position_sizeï¼Œç³»çµ±æœƒæ ¹æ“šç›®æ¨™å€‰ä½å¤§å°èª¿æ•´ç¾æœ‰å€‰ä½
    - position_size > 0: ç›®æ¨™ç‚ºå¤šå€‰
    - position_size < 0: ç›®æ¨™ç‚ºç©ºå€‰
    - position_size == 0: ç›®æ¨™ç‚ºå¹³å€‰
    - å¦‚æœæœªæä¾› position_sizeï¼Œå‰‡ä½¿ç”¨èˆŠçš„è¨‚å–®å°å‘æ¨¡å¼ï¼ˆBUY=é–‹å¤šï¼ŒSELL=é–‹ç©ºï¼‰
    """
    secret: str = Field(..., description="Webhook å¯†é‘°ï¼Œç”¨æ–¼é©—è­‰è«‹æ±‚ä¾†æº")
    signal_key: Optional[str] = Field(None, description="æ–°æ ¼å¼ï¼šå°æ‡‰ TVSignalConfig.signal_key")
    bot_key: Optional[str] = Field(None, description="èˆŠæ ¼å¼ï¼šå°æ‡‰ BotConfig.bot_keyï¼ˆå…¼å®¹ï¼‰")
    symbol: str = Field(..., description="äº¤æ˜“å°ï¼Œä¾‹å¦‚ï¼šBTCUSDT")
    side: str = Field(..., description="äº¤æ˜“æ–¹å‘ï¼šBUY æˆ– SELL")
    qty: float = Field(..., description="äº¤æ˜“æ•¸é‡")
    position_size: Optional[float] = Field(None, description="ç›®æ¨™å€‰ä½å¤§å°ï¼ˆä½ç½®å°å‘æ¨¡å¼ï¼‰ã€‚>0=å¤šå€‰ï¼Œ<0=ç©ºå€‰ï¼Œ0=å¹³å€‰ã€‚å¦‚æœæœªæä¾›å‰‡ä½¿ç”¨èˆŠçš„è¨‚å–®å°å‘æ¨¡å¼")
    time: Optional[str] = Field(None, description="TradingView å‚³ä¾†çš„æ™‚é–“å­—ä¸²ï¼ˆå¯é¸ï¼‰")
    extra: Optional[dict] = Field(None, description="å½ˆæ€§æ¬„ä½ï¼ˆå¯é¸ï¼‰")


class PositionOut(BaseModel):
    """Position å›æ‡‰æ ¼å¼
    
    æ™‚é–“æ¬„ä½ä½¿ç”¨ datetime å‹åˆ¥ï¼ŒFastAPI æœƒè‡ªå‹•åºåˆ—åŒ–ç‚º ISO8601 æ ¼å¼ã€‚
    Pydantic æœƒè‡ªå‹•è§£æ ISO8601 å­—ä¸²ç‚º datetime ç‰©ä»¶ã€‚
    """
    id: int
    symbol: str
    side: str
    qty: float
    entry_price: float
    status: str
    binance_order_id: Optional[int] = None
    client_order_id: Optional[str] = None
    highest_price: Optional[float] = None
    trail_callback: Optional[float] = None
    dyn_profit_threshold_pct: Optional[float] = None
    base_stop_loss_pct: Optional[float] = None
    exit_price: Optional[float] = None
    exit_reason: Optional[str] = None
    created_at: datetime
    closed_at: Optional[datetime] = None
    # æ·»åŠ å¯¦éš›ä½¿ç”¨çš„å€¼å’Œä¾†æºæ¨™è¨˜ï¼ˆç”¨æ–¼å‰ç«¯é¡¯ç¤ºå’Œé¡è‰²æ¨™è¨˜ï¼‰
    profit_threshold_value: Optional[float] = None
    profit_threshold_source: Optional[str] = None  # "override", "global", "default"
    lock_ratio_value: Optional[float] = None
    lock_ratio_source: Optional[str] = None  # "override", "global", "default"
    base_sl_value: Optional[float] = None
    base_sl_source: Optional[str] = None  # "override", "global", "default"
    # æ·»åŠ åœæç‹€æ…‹ï¼ˆåƒ…å° OPEN ç‹€æ…‹çš„å€‰ä½æœ‰æ•ˆï¼‰
    stop_mode: Optional[str] = None  # "dynamic", "base", "none"
    base_stop_price: Optional[float] = None
    dynamic_stop_price: Optional[float] = None
    # åœæ/æ­¢ç›ˆæ©Ÿåˆ¶æ§åˆ¶
    bot_stop_loss_enabled: bool = True
    tv_signal_close_enabled: bool = True


class TrailingUpdate(BaseModel):
    """æ›´æ–°è¿½è¹¤åœæè«‹æ±‚æ ¼å¼"""
    trailing_callback_percent: float = Field(..., description="è¿½è¹¤åœæå›èª¿ç™¾åˆ†æ¯”ï¼Œä¾‹å¦‚ï¼š2.0 ä»£è¡¨ 2%")
    activation_profit_percent: Optional[float] = Field(None, description="å•Ÿå‹•è¿½è¹¤åœæçš„ç²åˆ©ç™¾åˆ†æ¯”ï¼Œä¾‹å¦‚ï¼š1.0 ä»£è¡¨å…ˆè³º 1% å†å•Ÿå‹•è¿½è¹¤")


class TrailingSideConfig(BaseModel):
    """å–®å´ (LONG æˆ– SHORT) çš„ Trailing Stop è¨­å®š"""
    profit_threshold_pct: float = 1.0      # PnL% >= æ­¤å€¼æ‰å•Ÿå‹•é–åˆ©
    lock_ratio: float = 2.0 / 3.0          # é–åˆ©æ¯”ä¾‹ (ç´„ 0.67 = 2/3)
    base_sl_pct: float = 0.5               # åŸºç¤åœæè·é›¢ (%)


class TrailingConfig(BaseModel):
    """Trailing Stop å…¨åŸŸè¨­å®šæ¨¡å‹ï¼ˆåˆ† LONG å’Œ SHORTï¼‰"""
    trailing_enabled: bool = True
    long_config: TrailingSideConfig = Field(default_factory=lambda: TrailingSideConfig())
    short_config: TrailingSideConfig = Field(default_factory=lambda: TrailingSideConfig())
    auto_close_enabled: bool = True        # Dynamic Stop è§¸ç™¼æ™‚æ˜¯å¦è‡ªå‹•é—œå€‰
    
    # å‘å¾Œå…¼å®¹ï¼šæä¾›èˆŠçš„å±¬æ€§è¨ªå•æ–¹å¼ï¼ˆè¿”å› LONG çš„è¨­å®šï¼‰
    @property
    def profit_threshold_pct(self) -> float:
        """å‘å¾Œå…¼å®¹ï¼šè¿”å› LONG çš„ profit_threshold_pct"""
        return self.long_config.profit_threshold_pct
    
    @property
    def lock_ratio(self) -> float:
        """å‘å¾Œå…¼å®¹ï¼šè¿”å› LONG çš„ lock_ratio"""
        return self.long_config.lock_ratio
    
    @property
    def base_sl_pct(self) -> float:
        """å‘å¾Œå…¼å®¹ï¼šè¿”å› LONG çš„ base_sl_pct"""
        return self.long_config.base_sl_pct
    
    def get_config_for_side(self, side: str) -> TrailingSideConfig:
        """æ ¹æ“šå€‰ä½æ–¹å‘å–å¾—å°æ‡‰çš„è¨­å®š"""
        side_upper = side.upper()
        if side_upper == "LONG":
            return self.long_config
        elif side_upper == "SHORT":
            return self.short_config
        else:
            # é è¨­è¿”å› LONG è¨­å®š
            return self.long_config


class TrailingSideConfigUpdate(BaseModel):
    """æ›´æ–°å–®å´ Trailing è¨­å®šçš„è«‹æ±‚æ ¼å¼"""
    profit_threshold_pct: Optional[float] = None
    lock_ratio: Optional[float] = None
    base_sl_pct: Optional[float] = None


class TrailingConfigUpdate(BaseModel):
    """æ›´æ–° Trailing è¨­å®šçš„è«‹æ±‚æ ¼å¼"""
    trailing_enabled: Optional[bool] = None
    long_config: Optional[TrailingSideConfigUpdate] = None
    short_config: Optional[TrailingSideConfigUpdate] = None
    auto_close_enabled: Optional[bool] = None
    
    # å‘å¾Œå…¼å®¹ï¼šæ”¯æ´èˆŠæ ¼å¼ï¼ˆæœƒåŒæ™‚æ›´æ–° LONG å’Œ SHORTï¼‰
    profit_threshold_pct: Optional[float] = None
    lock_ratio: Optional[float] = None
    base_sl_pct: Optional[float] = None


class BinanceCloseRequest(BaseModel):
    """Binance Live Position é—œå€‰è«‹æ±‚æ ¼å¼"""
    symbol: str = Field(..., description="äº¤æ˜“å°ï¼Œä¾‹å¦‚ BTCUSDT")
    position_side: str = Field(..., description="å€‰ä½æ–¹å‘ï¼ŒLONG æˆ– SHORT")


class TVSignalConfigBase(BaseModel):
    """TradingView Signal Config åŸºç¤æ¨¡å‹"""
    name: str
    signal_key: str
    description: Optional[str] = None
    symbol_hint: Optional[str] = None
    timeframe_hint: Optional[str] = None
    enabled: bool = True


class TVSignalConfigCreate(TVSignalConfigBase):
    """å»ºç«‹ Signal Config çš„è«‹æ±‚æ ¼å¼"""
    pass


class TVSignalConfigUpdate(BaseModel):
    """æ›´æ–° Signal Config çš„è«‹æ±‚æ ¼å¼"""
    name: Optional[str] = None
    signal_key: Optional[str] = None
    description: Optional[str] = None
    symbol_hint: Optional[str] = None
    timeframe_hint: Optional[str] = None
    enabled: Optional[bool] = None


class TVSignalConfigOut(TVSignalConfigBase):
    """Signal Config å›æ‡‰æ ¼å¼"""
    id: int
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    
    class Config:
        from_attributes = True
        orm_mode = True


class BotConfigBase(BaseModel):
    """Bot è¨­å®šåŸºç¤æ¨¡å‹"""
    name: str
    bot_key: str
    enabled: bool = True
    symbol: str = "BTCUSDT"
    use_signal_side: bool = True
    fixed_side: Optional[str] = None
    qty: float = 0.01
    max_invest_usdt: Optional[float] = None
    leverage: int = 20
    use_dynamic_stop: bool = True
    trailing_callback_percent: Optional[float] = None
    base_stop_loss_pct: float = 3.0
    signal_id: Optional[int] = None


class BotConfigCreate(BotConfigBase):
    """å»ºç«‹ Bot è¨­å®šçš„è«‹æ±‚æ ¼å¼"""
    pass


class BotConfigUpdate(BaseModel):
    """æ›´æ–° Bot è¨­å®šçš„è«‹æ±‚æ ¼å¼"""
    name: Optional[str] = None
    enabled: Optional[bool] = None
    symbol: Optional[str] = None
    use_signal_side: Optional[bool] = None
    fixed_side: Optional[str] = None
    qty: Optional[float] = None
    max_invest_usdt: Optional[float] = None
    leverage: Optional[int] = None
    use_dynamic_stop: Optional[bool] = None
    trailing_callback_percent: Optional[float] = None
    base_stop_loss_pct: Optional[float] = None
    signal_id: Optional[int] = None
    max_invest_password: Optional[str] = Field(None, description="æ›´æ–° max_invest_usdt æ™‚éœ€è¦çš„å¯†ç¢¼")


class BotConfigOut(BotConfigBase):
    """Bot è¨­å®šå›æ‡‰æ ¼å¼"""
    id: int
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None
    signal: Optional[TVSignalConfigOut] = None  # é—œè¯çš„ Signal Configï¼ˆå¯é¸ï¼‰
    
    class Config:
        from_attributes = True
        orm_mode = True


class WebhookResponse(BaseModel):
    """Webhook å›æ‡‰æ ¼å¼"""
    success: bool
    message: str
    position_id: int
    binance_order: dict


# ==================== Trailing å…¨åŸŸè¨­å®šï¼ˆin-memoryï¼‰====================
# ç›®å‰åŸ·è¡Œç·’å…§ä½¿ç”¨çš„å…¨åŸŸè¨­å®šï¼ˆin-memoryï¼›ä¹‹å¾Œå¿…è¦æ™‚å†å¯« DBï¼‰
# TODO: ä¹‹å¾Œ dynamic stop ç›¸é—œé‚è¼¯æœƒçµ±ä¸€æ”¹ç”¨æ­¤é…ç½®ï¼Œè€Œä¸æ˜¯ç›´æ¥è®€å– DYN_* å¸¸æ•¸
# æ³¨æ„ï¼šå¿…é ˆåœ¨ TrailingConfig é¡å®šç¾©ä¹‹å¾Œæ‰èƒ½åˆå§‹åŒ–
# å¾ç’°å¢ƒè®Šæ•¸è®€å–é è¨­å€¼ï¼Œå¦‚æœæ²’æœ‰è¨­å®šå‰‡ä½¿ç”¨é è¨­å€¼
# é€™äº›å€¼å¯ä»¥é€šé .env æª”æ¡ˆè¨­å®šï¼š
#   DYN_PROFIT_THRESHOLD_PCT=1.0    (PnL% threshold)
#   DYN_LOCK_RATIO_DEFAULT=0.666    (Lock ratio, ä¾‹å¦‚ 0.666 ä»£è¡¨ 2/3)
#   DYN_BASE_SL_PCT=0.5              (Base SL %)
# åˆå§‹åŒ– TRAILING_CONFIGï¼Œå¾ç’°å¢ƒè®Šæ•¸è®€å–è¨­å®š
_default_profit_threshold = float(os.getenv("DYN_PROFIT_THRESHOLD_PCT", "1.0"))
_default_lock_ratio = float(os.getenv("DYN_LOCK_RATIO_DEFAULT", str(2.0 / 3.0)))
_default_base_sl = float(os.getenv("DYN_BASE_SL_PCT", "0.5"))

TRAILING_CONFIG = TrailingConfig(
    trailing_enabled=True,
    long_config=TrailingSideConfig(
        profit_threshold_pct=_default_profit_threshold,
        lock_ratio=_default_lock_ratio,
        base_sl_pct=_default_base_sl
    ),
    short_config=TrailingSideConfig(
        profit_threshold_pct=_default_profit_threshold,
        lock_ratio=_default_lock_ratio,
        base_sl_pct=_default_base_sl
    ),
    auto_close_enabled=True
)


# ==================== èªè­‰ä¾è³´ ====================

def verify_admin_api_key(x_api_key: str = Header(...)):
    """
    é©—è­‰ç®¡ç†å“¡ API Keyï¼ˆèˆŠç‰ˆï¼Œä¿ç•™å‘å¾Œå…¼å®¹ï¼‰
    
    Args:
        x_api_key: å¾ Header å–å¾—çš„ X-API-KEY
    
    Raises:
        HTTPException: ç•¶ API Key ä¸æ­£ç¢ºæ™‚
    """
    admin_api_key = os.getenv("ADMIN_API_KEY", "")
    
    if not admin_api_key:
        raise HTTPException(
            status_code=500,
            detail="ä¼ºæœå™¨æœªè¨­å®š ADMIN_API_KEYï¼Œè«‹è¯ç¹«ç®¡ç†å“¡"
        )
    
    if x_api_key != admin_api_key:
        raise HTTPException(
            status_code=401,
            detail="ç„¡æ•ˆçš„ API Key"
        )
    
    return x_api_key


async def require_admin_user(request: Request):
    """
    é©—è­‰ä½¿ç”¨è€…æ˜¯å¦ç‚ºç®¡ç†å“¡ï¼ˆä½¿ç”¨ Google OAuth Sessionï¼‰
    
    å¦‚æœæœªè¨­å®š Google OAuthï¼Œå‰‡å•Ÿç”¨ã€é–‹ç™¼æ¨¡å¼ã€ï¼Œæ‰€æœ‰è«‹æ±‚ç›´æ¥è¦–ç‚ºç®¡ç†å“¡ã€‚
    å¦‚æœå·²å•Ÿç”¨ Google OAuthï¼Œå‰‡æª¢æŸ¥ session ä¸­æ˜¯å¦æœ‰ is_admin=True å’Œ emailã€‚
    å¦‚æœä¸æ˜¯ç®¡ç†å“¡ï¼Œçµ±ä¸€ä½¿ç”¨ HTTPException ä¸­æ–·æµç¨‹ï¼š
    - å°æ–¼ HTML è«‹æ±‚ï¼šä½¿ç”¨ 307 + Location header å°å‘ç™»å…¥é é¢
    - å°æ–¼ API è«‹æ±‚ï¼šå›å‚³ 401/403
    
    Args:
        request: FastAPI Request ç‰©ä»¶
    
    Returns:
        dict: åŒ…å« email çš„ä½¿ç”¨è€…è³‡è¨Š
    
    Raises:
        HTTPException: ç•¶ä½¿ç”¨è€…æœªç™»å…¥æˆ–ä¸æ˜¯ç®¡ç†å“¡æ™‚ï¼ˆåƒ…åœ¨ OAuth å•Ÿç”¨æ™‚ï¼‰
            - æœªç™»å…¥ä¸”ç‚º HTML è«‹æ±‚ï¼š307 é‡å®šå‘åˆ° /auth/login
            - æœªç™»å…¥ä¸”ç‚º API è«‹æ±‚ï¼š401 Unauthorized
            - å·²ç™»å…¥ä½†ä¸æ˜¯ç®¡ç†å“¡ï¼š403 Forbidden
    """
    # é–‹ç™¼æ¨¡å¼ï¼šGoogle OAuth æœªè¨­å®š â†’ ç›´æ¥å›å‚³ä¸€å€‹å‡çš„ admin ä½¿ç”¨è€…
    if not GOOGLE_OAUTH_ENABLED:
        dev_email = ADMIN_GOOGLE_EMAIL or "dev-admin@example.com"
        return {"email": dev_email}
    
    session = request.session
    accept = request.headers.get("accept", "")
    
    # æª¢æŸ¥ session ä¸­æ˜¯å¦æœ‰ç®¡ç†å“¡æ¨™è¨˜
    if not session.get("is_admin") or not session.get("email"):
        # å¦‚æœæ˜¯ HTML è«‹æ±‚ï¼Œä½¿ç”¨ 307 é‡å®šå‘åˆ°ç™»å…¥é é¢
        if "text/html" in accept:
            raise HTTPException(
                status_code=status.HTTP_307_TEMPORARY_REDIRECT,
                detail="è«‹å…ˆç™»å…¥",
                headers={"Location": "/auth/login"},
            )
        # å¦å‰‡å›å‚³ 401
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="è«‹å…ˆç™»å…¥",
        )
    
    # æª¢æŸ¥ email æ˜¯å¦ç‚ºç®¡ç†å“¡ email
    email = session.get("email")
    if email != ADMIN_GOOGLE_EMAIL:
        # çµ±ä¸€ä½¿ç”¨ 403ï¼Œä¸å†å›å‚³ HTMLResponse
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="ä½ æ²’æœ‰æ¬Šé™ä½¿ç”¨æ­¤ç³»çµ±",
        )
    
    # é©—è­‰é€šéï¼Œå›å‚³ä½¿ç”¨è€…è³‡è¨Š
    return {"email": email}


def verify_tradingview_secret(secret: str):
    """
    é©—è­‰ TradingView Webhook Secret
    
    Args:
        secret: å¾è«‹æ±‚å–å¾—çš„ secret
    
    Raises:
        HTTPException: ç•¶ secret ä¸æ­£ç¢ºæ™‚
    """
    tradingview_secret = os.getenv("TRADINGVIEW_SECRET", "")
    
    if not tradingview_secret:
        raise HTTPException(
            status_code=500,
            detail="ä¼ºæœå™¨æœªè¨­å®š TRADINGVIEW_SECRETï¼Œè«‹è¯ç¹«ç®¡ç†å“¡"
        )
    
    if secret != tradingview_secret:
        raise HTTPException(
            status_code=401,
            detail="ç„¡æ•ˆçš„ Webhook Secret"
        )


# ==================== API ç«¯é» ====================

@app.get("/")
async def root():
    """æ ¹è·¯å¾‘ï¼Œç”¨æ–¼å¥åº·æª¢æŸ¥"""
    return {
        "status": "ok",
        "message": "TradingView Binance Bot is running",
        "version": "1.0.0"
    }


# ==================== Google OAuth è·¯ç”± ====================

@app.get("/auth/login")
async def login(request: Request):
    """
    å°å‘ Google OAuth æˆæ¬Šé é¢ã€‚
    å¦‚æœæœªè¨­å®š Google OAuthï¼Œé¡¯ç¤ºç°¡å–®æç¤ºé ã€‚
    """
    if not GOOGLE_OAUTH_ENABLED:
        # é–‹ç™¼æ¨¡å¼ï¼šé¡¯ç¤ºæç¤ºé å³å¯ï¼Œé¿å… 500
        return HTMLResponse(
            content="""
            <html>
                <head>
                    <meta charset="UTF-8">
                    <title>Google OAuth æœªè¨­å®š</title>
                    <style>
                        body {
                            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                            max-width: 600px;
                            margin: 50px auto;
                            padding: 20px;
                            line-height: 1.6;
                        }
                        h1 { color: #333; }
                        a {
                            display: inline-block;
                            margin-top: 20px;
                            padding: 10px 20px;
                            background-color: #007bff;
                            color: white;
                            text-decoration: none;
                            border-radius: 4px;
                        }
                        a:hover { background-color: #0056b3; }
                    </style>
                </head>
                <body>
                    <h1>Google OAuth æœªè¨­å®š</h1>
                    <p>ç›®å‰ç³»çµ±åœ¨é–‹ç™¼æ¨¡å¼ä¸‹åŸ·è¡Œï¼Œæœªå•Ÿç”¨ Google ç™»å…¥ã€‚</p>
                    <p>è«‹è¨­å®š GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET / ADMIN_GOOGLE_EMAIL å¾Œé‡æ–°å•Ÿå‹•ä¼ºæœå™¨ã€‚</p>
                    <p><a href="/dashboard">å›åˆ° Dashboard</a></p>
                </body>
            </html>
            """,
            status_code=200,
        )
    
    # æ­£å¸¸æƒ…æ³ï¼šå°å‘ Google OAuth
    redirect_uri = str(request.url_for("auth_callback"))
    return await oauth.google.authorize_redirect(request, redirect_uri)


@app.get("/auth/callback")
async def auth_callback(request: Request):
    """
    Google OAuth å›èª¿è™•ç†
    
    æ¥æ”¶ Google å›å‚³çš„ codeï¼Œäº¤æ› tokenï¼Œå–å¾—ä½¿ç”¨è€… emailï¼Œ
    ä¸¦æª¢æŸ¥æ˜¯å¦ç‚ºç®¡ç†å“¡ emailã€‚
    """
    try:
        token = await oauth.google.authorize_access_token(request)
        
        # å–å¾—ä½¿ç”¨è€…è³‡è¨Š
        resp = await oauth.google.get('userinfo', token=token)
        user_info = resp.json()
        
        email = user_info.get("email")
        if not email:
            raise HTTPException(status_code=400, detail="ç„¡æ³•å–å¾— email")
        
        # æª¢æŸ¥æ˜¯å¦ç‚ºç®¡ç†å“¡ email
        if email == ADMIN_GOOGLE_EMAIL:
            # è¨­å®š session
            request.session["is_admin"] = True
            request.session["email"] = email
            request.session["name"] = user_info.get("name", email)
            
            logger.info(f"ç®¡ç†å“¡ç™»å…¥æˆåŠŸ: {email}")
            return RedirectResponse(url="/dashboard")
        else:
            logger.warning(f"éç®¡ç†å“¡å˜—è©¦ç™»å…¥: {email}")
            return HTMLResponse(
                content=f"""
                <html>
                    <body>
                        <h1>æœªæˆæ¬Š</h1>
                        <p>ä½ æ²’æœ‰æ¬Šé™ä½¿ç”¨æ­¤ç³»çµ±ã€‚</p>
                        <p>Email: {email}</p>
                        <p><a href="/auth/logout">è¿”å›</a></p>
                    </body>
                </html>
                """,
                status_code=403
            )
    
    except Exception as e:
        logger.error(f"OAuth å›èª¿è™•ç†å¤±æ•—: {e}")
        raise HTTPException(status_code=500, detail=f"ç™»å…¥å¤±æ•—: {str(e)}")


@app.get("/auth/logout")
async def logout(request: Request):
    """
    ç™»å‡ºï¼Œæ¸…é™¤ session
    """
    request.session.clear()
    return RedirectResponse(url="/")


# ==================== Dashboard è·¯ç”± ====================

@app.get("/me")
async def me(user: dict = Depends(require_admin_user)):
    """
    å–å¾—ç•¶å‰ä½¿ç”¨è€…è³‡è¨Šå’Œ Binance æ¨¡å¼ã€‚
    åœ¨ Demo æ¨¡å¼ä¸‹ï¼ˆGOOGLE_OAUTH_ENABLED=Falseï¼‰ï¼Œæœƒè‡ªå‹•è¿”å›å‡çš„ç®¡ç†å“¡è³‡è¨Šã€‚
    """
    # å¦‚æœå·²ç¶“åœ¨ require_admin_user ä¸­é€šéé©—è­‰ï¼Œç›´æ¥è¿”å›
    user_email = user.get("email", "demo@example.com")
    
    # å–å¾— Binance æ¨¡å¼
    binance_mode = "demo"
    try:
        client, mode = get_client()
        binance_mode = mode
    except Exception:
        pass
    
    # å–å¾— TRADINGVIEW_SECRETï¼ˆåƒ…ä¾›ç®¡ç†å“¡ä½¿ç”¨ï¼Œç”¨æ–¼ç”Ÿæˆ TradingView Alert Templateï¼‰
    tradingview_secret = os.getenv("TRADINGVIEW_SECRET", "")
    
    return {
        "user_email": user_email,
        "binance_mode": binance_mode,
        "tradingview_secret": tradingview_secret,  # ç”¨æ–¼å‰ç«¯ç”Ÿæˆæ¨¡æ¿
    }


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(
    request: Request,
    user: dict = Depends(require_admin_user)
):
    """
    é¡¯ç¤ºå€‰ä½ Dashboard
    
    åƒ…é™å·²ç™»å…¥ä¸”é€šéç®¡ç†å“¡é©—è­‰çš„ä½¿ç”¨è€…ä½¿ç”¨ï¼ˆGoogle OAuth + ADMIN_GOOGLE_EMAILï¼‰ã€‚
    å‰ç«¯ä½¿ç”¨ JavaScript å‹•æ…‹å¾ /positions API è¼‰å…¥è³‡æ–™ã€‚
    """
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "user": user if isinstance(user, dict) else {},
            "user_email": user.get("email") if isinstance(user, dict) else "",
        }
    )


# ==================== å·¥å…· API ====================

@app.get("/api/mark-price/{symbol}")
async def get_mark_price_api(
    symbol: str,
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾—äº¤æ˜“å°çš„æ¨™è¨˜åƒ¹æ ¼ï¼ˆç”¨æ–¼å‰ç«¯è¨ˆç®— qtyï¼‰
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        symbol: äº¤æ˜“å°ï¼Œä¾‹å¦‚ BTCUSDT
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: åŒ…å« mark_price çš„å­—å…¸
    """
    try:
        mark_price = get_mark_price(symbol.upper())
        return {
            "symbol": symbol.upper(),
            "mark_price": mark_price
        }
    except Exception as e:
        logger.error(f"å–å¾— {symbol} æ¨™è¨˜åƒ¹æ ¼å¤±æ•—: {e}")
        raise HTTPException(status_code=500, detail=f"ç„¡æ³•å–å¾— {symbol} çš„æ¨™è¨˜åƒ¹æ ¼: {str(e)}")


@app.get("/api/symbol-info/{symbol}")
async def get_symbol_info_api(
    symbol: str,
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾—äº¤æ˜“å°çš„ç²¾åº¦è³‡è¨Šï¼ˆæ•¸é‡ç²¾åº¦ã€åƒ¹æ ¼ç²¾åº¦ã€æ­¥é•·ç­‰ï¼‰
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        symbol: äº¤æ˜“å°ï¼Œä¾‹å¦‚ BTCUSDT
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: åŒ…å«ç²¾åº¦è³‡è¨Šçš„å­—å…¸
    """
    try:
        symbol_info = get_symbol_info(symbol.upper())
        # åªå›å‚³å¿…è¦çš„è³‡è¨Šï¼Œä¸åŒ…å« raw åŸå§‹è³‡æ–™
        return {
            "symbol": symbol.upper(),
            "quantityPrecision": symbol_info["quantityPrecision"],
            "stepSize": symbol_info["stepSize"],
            "pricePrecision": symbol_info["pricePrecision"],
            "tickSize": symbol_info["tickSize"],
            "minQty": symbol_info["minQty"],
            "maxQty": symbol_info["maxQty"] if symbol_info["maxQty"] != float("inf") else None
        }
    except Exception as e:
        logger.error(f"å–å¾— {symbol} ç²¾åº¦è³‡è¨Šå¤±æ•—: {e}")
        raise HTTPException(status_code=500, detail=f"ç„¡æ³•å–å¾— {symbol} çš„ç²¾åº¦è³‡è¨Š: {str(e)}")


# èˆŠç‰ˆ webhook endpoint å·²ç§»é™¤ï¼Œæ”¹ç”¨ä¸‹é¢çš„æ–°ç‰ˆ endpoint æ”¯æ´ bot_key


def calculate_qty_from_max_invest(bot: BotConfig, symbol: str, target_qty: Optional[float] = None) -> float:
    """
    æ ¹æ“š max_invest_usdt è¨ˆç®— qtyï¼ˆåœ¨ä¸‹å–®ç•¶ä¸‹æ ¹æ“šå³æ™‚åƒ¹æ ¼è¨ˆç®—ï¼‰
    
    æ­¤å‡½æ•¸åœ¨ webhook è¢«è§¸ç™¼ä¸‹å–®æ™‚æ‰èª¿ç”¨ï¼Œæœƒå–å¾—ç•¶ä¸‹çš„å³æ™‚æ¨™è¨˜åƒ¹æ ¼ä¾†è¨ˆç®—æ•¸é‡ã€‚
    
    Args:
        bot: Bot è¨­å®š
        symbol: äº¤æ˜“å°
        target_qty: ç›®æ¨™æ•¸é‡ï¼ˆä½ç½®å°å‘æ¨¡å¼ä½¿ç”¨ï¼‰ï¼Œå¦‚æœç‚º None å‰‡ä½¿ç”¨ bot.qty
    
    Returns:
        è¨ˆç®—å¾Œçš„ qty
    """
    if bot.max_invest_usdt is not None and bot.max_invest_usdt > 0:
        try:
            # å–å¾—ç•¶ä¸‹çš„å³æ™‚æ¨™è¨˜åƒ¹æ ¼
            current_price = get_mark_price(symbol)
            if current_price and current_price > 0:
                max_qty_from_invest = bot.max_invest_usdt / current_price
                if target_qty is not None:
                    # ä½ç½®å°å‘æ¨¡å¼ï¼šä½¿ç”¨è¼ƒå°å€¼ï¼ˆä¸è¶…é max_invest_usdtï¼‰
                    qty = min(target_qty, max_qty_from_invest)
                    logger.info(
                        f"Bot {bot.id} ä½ç½®å°å‘æ¨¡å¼ï¼šmax_invest_usdt={bot.max_invest_usdt} USDT, "
                        f"å³æ™‚åƒ¹æ ¼={current_price}, target_qty={target_qty}, "
                        f"max_qty_from_invest={max_qty_from_invest}, æœ€çµ‚ qty={qty}"
                    )
                else:
                    # è¨‚å–®å°å‘æ¨¡å¼ï¼šç›´æ¥ä½¿ç”¨ max_invest_usdt è¨ˆç®—
                    qty = max_qty_from_invest
                    logger.info(
                        f"Bot {bot.id} è¨‚å–®å°å‘æ¨¡å¼ï¼šmax_invest_usdt={bot.max_invest_usdt} USDT, "
                        f"å³æ™‚åƒ¹æ ¼={current_price}, è¨ˆç®— qty={qty}"
                    )
                return qty
            else:
                logger.warning(f"Bot {bot.id} ç„¡æ³•å–å¾— {symbol} çš„å³æ™‚åƒ¹æ ¼ï¼Œä½¿ç”¨é è¨­ qty")
                return target_qty if target_qty is not None else bot.qty
        except Exception as e:
            logger.warning(f"Bot {bot.id} è¨ˆç®— qty æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}ï¼Œä½¿ç”¨é è¨­ qty")
            return target_qty if target_qty is not None else bot.qty
    else:
        # æ²’æœ‰è¨­å®š max_invest_usdtï¼Œä½¿ç”¨åŸå§‹å€¼
        return target_qty if target_qty is not None else bot.qty


def normalize_symbol_from_tv(tv_symbol: str) -> str:
    """
    Normalize TradingView symbol into a Binance-compatible futures symbol.
    
    Examples:
    - "BTCUSDT.P"           -> "BTCUSDT"
    - "BINANCE:BTCUSDT.P"   -> "BTCUSDT"
    - "btcusdt.p"           -> "BTCUSDT"
    - "BTCUSDT"             -> "BTCUSDT"
    - "BINANCE:ETHUSDT"     -> "ETHUSDT"
    
    Args:
        tv_symbol: TradingView symbol string
    
    Returns:
        str: Normalized Binance-compatible symbol (uppercase, no exchange prefix, no .P suffix)
    
    Raises:
        ValueError: If the normalized symbol is empty
    """
    if not tv_symbol:
        return tv_symbol
    
    s = tv_symbol.strip().upper()
    
    # Remove exchange prefix like "BINANCE:"
    if ":" in s:
        _, s = s.split(":", 1)
        s = s.strip()
    
    # Remove common TradingView perp suffix ".P"
    if s.endswith(".P"):
        s = s[:-2]
    
    # Final validation
    if not s:
        raise ValueError(f"Normalized symbol is empty. Original: '{tv_symbol}'")
    
    return s


def get_current_position_signed_qty(db: Session, bot_id: int, symbol: str) -> tuple[Optional[Position], float]:
    """
    å–å¾—ç•¶å‰ bot çš„å€‰ä½ï¼ˆä»¥å¸¶ç¬¦è™Ÿçš„æ•¸é‡è¡¨ç¤ºï¼‰
    
    Args:
        db: è³‡æ–™åº« Session
        bot_id: Bot ID
        symbol: äº¤æ˜“å°
    
    Returns:
        tuple: (Position ç‰©ä»¶æˆ– None, å¸¶ç¬¦è™Ÿçš„æ•¸é‡)
        - å¦‚æœæ²’æœ‰å€‰ä½ï¼Œè¿”å› (None, 0.0)
        - å¦‚æœæ˜¯ LONGï¼Œè¿”å› (position, +qty)
        - å¦‚æœæ˜¯ SHORTï¼Œè¿”å› (position, -qty)
    """
    position = db.query(Position).filter(
        Position.bot_id == bot_id,
        Position.symbol == symbol.upper(),
        Position.status == "OPEN"
    ).first()
    
    if not position:
        return None, 0.0
    
    # è½‰æ›ç‚ºå¸¶ç¬¦è™Ÿçš„æ•¸é‡
    if position.side == "LONG":
        return position, position.qty
    elif position.side == "SHORT":
        return position, -position.qty
    else:
        return position, 0.0


@app.post("/webhook/tradingview", response_model=dict)
async def webhook_tradingview(
    request: Request,
    db: Session = Depends(get_db),
):
    """
    æ¥æ”¶ TradingView Webhookï¼ˆæ”¯æ´ signal_key å’Œ bot_keyï¼‰
    
    è™•ç†æµç¨‹ï¼š
    1. è®€å–åŸå§‹ body ä¸¦å„²å­˜ç‚º raw_payload
    2. è§£æ JSON ä¸¦é©—è­‰ secret
    3. å…ˆå»ºç«‹ TradingViewSignalLog è¨˜éŒ„ï¼ˆç¸½æ˜¯åŸ·è¡Œï¼‰
    4. æ ¹æ“š signal_key æˆ– bot_key æŸ¥æ‰¾å°æ‡‰çš„ enabled bots
    5. ç‚ºæ¯å€‹ bot ä¸‹å–®ä¸¦å»ºç«‹ Position
    6. æ›´æ–° log çš„ processed å’Œ process_resultï¼ˆç¸½æ˜¯åŸ·è¡Œï¼‰
    
    æ”¯æ´å…©ç¨®æ¨¡å¼ï¼š
    - æ–°æ ¼å¼ï¼ˆæ¨è–¦ï¼‰ï¼šä½¿ç”¨ signal_keyï¼ŒæŸ¥æ‰¾æ‰€æœ‰ enabled=True ä¸” signal_id åŒ¹é…çš„ bots
    - èˆŠæ ¼å¼ï¼ˆå…¼å®¹ï¼‰ï¼šä½¿ç”¨ bot_keyï¼Œç›´æ¥æŸ¥æ‰¾å°æ‡‰çš„ bot
    
    Args:
        request: FastAPI Request ç‰©ä»¶
        db: è³‡æ–™åº« Session
    
    Returns:
        dict: è™•ç†çµæœï¼ŒåŒ…å« successã€messageã€signal_log_idã€results ç­‰
    """
    TRADINGVIEW_SECRET = os.getenv("TRADINGVIEW_SECRET", "")
    log = None
    signal_config = None
    
    try:
        # è®€å–åŸå§‹ body
        raw_bytes = await request.body()
        raw_text = raw_bytes.decode("utf-8") if raw_bytes else ""
        
        # è§£æ JSON
        try:
            data = json.loads(raw_text) if raw_text else {}
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"ç„¡æ•ˆçš„ JSON æ ¼å¼: {str(e)}")
        
        # å»ºç«‹ Pydantic æ¨¡å‹
        signal = TradingViewSignalIn(**data)
        
        # é©—è­‰ secret
        if TRADINGVIEW_SECRET and signal.secret != TRADINGVIEW_SECRET:
            raise HTTPException(status_code=401, detail="ç„¡æ•ˆçš„ Webhook Secret")
        
        # Normalize symbol from TradingView
        raw_symbol = signal.symbol
        try:
            normalized_symbol = normalize_symbol_from_tv(raw_symbol)
            logger.info(
                f"[TV Webhook] Symbol normalization: raw='{raw_symbol}' -> normalized='{normalized_symbol}'"
            )
        except (ValueError, AttributeError) as e:
            logger.error(f"[TV Webhook] Symbol normalization failed: {e}, raw='{raw_symbol}'")
            raise HTTPException(
                status_code=400,
                detail=f"Invalid symbol from TradingView: '{raw_symbol}'. Normalization failed: {str(e)}"
            )
        
        # Replace signal.symbol with normalized version for all business logic
        # Note: We preserve the raw_symbol in the raw_payload for debugging
        signal.symbol = normalized_symbol
        
        # æ±ºå®šä½¿ç”¨å“ªç¨®æ¨¡å¼ï¼ˆå„ªå…ˆä½¿ç”¨ signal_keyï¼‰
        use_signal_key = signal.signal_key is not None and signal.signal_key != ""
        use_bot_key = signal.bot_key is not None and signal.bot_key != ""
        
        if not use_signal_key and not use_bot_key:
            raise HTTPException(status_code=400, detail="å¿…é ˆæä¾› signal_key æˆ– bot_key")
        
        # æª¢æŸ¥é‡è¤‡è¨Šè™Ÿï¼ˆé˜²æ­¢ TradingView é‡è¤‡ç™¼é€ï¼‰
        # å»ºç«‹è¨Šè™Ÿçš„å”¯ä¸€è­˜åˆ¥ç¢¼ï¼ˆæ’é™¤ time æ¬„ä½ï¼Œå› ç‚ºå¯èƒ½æ¯æ¬¡ä¸åŒï¼‰
        signal_fingerprint = {
            "signal_key": signal.signal_key,
            "bot_key": signal.bot_key,
            "symbol": normalized_symbol,
            "side": signal.side.upper(),
            "qty": signal.qty,
            "position_size": signal.position_size,
        }
        # å»ºç«‹ hash
        fingerprint_str = json.dumps(signal_fingerprint, sort_keys=True)
        signal_hash = hashlib.sha256(fingerprint_str.encode("utf-8")).hexdigest()[:16]
        
        # æª¢æŸ¥æœ€è¿‘ 60 ç§’å…§æ˜¯å¦æœ‰ç›¸åŒçš„è¨Šè™Ÿ
        time_threshold = datetime.now(timezone.utc) - timedelta(seconds=60)
        duplicate_log = (
            db.query(TradingViewSignalLog)
            .filter(
                TradingViewSignalLog.symbol == normalized_symbol.upper(),
                TradingViewSignalLog.side == signal.side.upper(),
                TradingViewSignalLog.qty == signal.qty,
                TradingViewSignalLog.position_size == signal.position_size,
                TradingViewSignalLog.received_at >= time_threshold
            )
        )
        
        # å¦‚æœä½¿ç”¨ signal_keyï¼Œä¹Ÿæª¢æŸ¥ signal_id
        if use_signal_key:
            signal_config = db.query(TVSignalConfig).filter(TVSignalConfig.signal_key == signal.signal_key).first()
            if not signal_config:
                raise HTTPException(status_code=400, detail=f"æ‰¾ä¸åˆ° signal_key='{signal.signal_key}' çš„ Signal Config")
            signal_id = signal_config.id
            duplicate_log = duplicate_log.filter(TradingViewSignalLog.signal_id == signal_id)
        else:
            signal_id = None
            duplicate_log = duplicate_log.filter(TradingViewSignalLog.bot_key == signal.bot_key)
        
        existing_duplicate = duplicate_log.first()
        if existing_duplicate:
            logger.warning(
                f"æª¢æ¸¬åˆ°é‡è¤‡è¨Šè™Ÿï¼Œè·³éè™•ç†ã€‚ç•¶å‰è¨Šè™Ÿ hash={signal_hash}, "
                f"é‡è¤‡çš„ log_id={existing_duplicate.id}, "
                f"é‡è¤‡æ™‚é–“={existing_duplicate.received_at}"
            )
            return {
                "success": False,
                "message": "é‡è¤‡è¨Šè™Ÿï¼Œå·²æ–¼æœ€è¿‘è™•ç†é",
                "duplicate_log_id": existing_duplicate.id,
                "signal_hash": signal_hash
            }
        
        # 1) å…ˆå»ºç«‹ signal logï¼ˆç¸½æ˜¯åŸ·è¡Œï¼‰
        signal_dict = signal.model_dump() if hasattr(signal, 'model_dump') else signal.dict()
        
        # ç¸½æ˜¯å„²å­˜ bot_keyï¼ˆå¦‚æœè«‹æ±‚ä¸­æœ‰æä¾›ï¼‰ï¼Œç„¡è«–ä½¿ç”¨å“ªç¨®æ¨¡å¼
        bot_key_to_store = signal.bot_key if signal.bot_key else None
        
        log = TradingViewSignalLog(
            bot_key=bot_key_to_store,
            signal_id=signal_id,
            symbol=signal.symbol.upper(),
            side=signal.side.upper(),
            qty=signal.qty,
            position_size=signal.position_size,
            raw_body=signal_dict,
            raw_payload=raw_text
        )
        db.add(log)
        db.commit()
        db.refresh(log)
        
        logger.info(
            f"[TV RAW] signal_log_id={log.id} payload={raw_text}"
        )
        logger.info(
            f"æ”¶åˆ° TradingView signal, log_id={log.id}, "
            f"signal_key={signal.signal_key if use_signal_key else None}, "
            f"bot_key={signal.bot_key if use_bot_key else None}, "
            f"raw_symbol='{raw_symbol}', normalized_symbol='{normalized_symbol}', "
            f"side={log.side}, qty={log.qty}, "
            f"position_size={log.position_size} (mode={'position-based' if log.position_size is not None else 'order-based'})"
        )
        
        # 2) æŸ¥æ‰¾ active bots
        if use_signal_key:
            # æ–°æ ¼å¼ï¼šæ ¹æ“š signal_id æŸ¥æ‰¾æ‰€æœ‰ enabled bots
            bots = (
                db.query(BotConfig)
                .filter(BotConfig.signal_id == signal_id, BotConfig.enabled == True)
                .all()
            )
        else:
            # èˆŠæ ¼å¼ï¼šæ ¹æ“š bot_key æŸ¥æ‰¾
            bots = (
                db.query(BotConfig)
                .filter(BotConfig.bot_key == signal.bot_key, BotConfig.enabled == True)
                .all()
            )
        
        if not bots:
            log.processed = True
            log.process_result = "no_active_bot"
            db.commit()
            return {
                "success": False,
                "message": "æ²’æœ‰å•Ÿç”¨ä¸­çš„ bot",
                "signal_log_id": log.id
            }
        
        client = get_client()
        results = []
        EPS = 1e-8  # æµ®é»æ•¸æ¯”è¼ƒçš„èª¤å·®ç¯„åœ
        
        # 3) ç‚ºæ¯å€‹ bot è™•ç†ï¼ˆä½ç½®å°å‘æˆ–è¨‚å–®å°å‘ï¼‰
        for bot in bots:
            try:
                # æ±ºå®š symbolï¼ˆå„ªå…ˆä½¿ç”¨ bot.symbolï¼Œå¦å‰‡ä½¿ç”¨å·²ç¶“è¦ç¯„åŒ–å¾Œçš„ signal.symbolï¼‰
                # å¦‚æœ bot.symbol å­˜åœ¨ï¼Œä¹Ÿéœ€è¦è¦ç¯„åŒ–ï¼›å¦å‰‡ä½¿ç”¨å·²è¦ç¯„åŒ–çš„ signal.symbol
                if bot.symbol:
                    try:
                        symbol = normalize_symbol_from_tv(bot.symbol)
                        logger.debug(f"Bot {bot.id} ä½¿ç”¨ bot.symbol='{bot.symbol}' -> normalized='{symbol}'")
                    except (ValueError, AttributeError) as e:
                        logger.warning(f"Bot {bot.id} çš„ symbol='{bot.symbol}' è¦ç¯„åŒ–å¤±æ•—ï¼Œä½¿ç”¨ signal.symbol: {e}")
                        symbol = normalized_symbol
                else:
                    symbol = normalized_symbol
                
                # æª¢æŸ¥æ˜¯å¦ä½¿ç”¨ä½ç½®å°å‘æ¨¡å¼
                target_position_size = signal.position_size
                
                # å¦‚æœæœªæä¾› position_sizeï¼Œä½¿ç”¨èˆŠçš„è¨‚å–®å°å‘æ¨¡å¼
                if target_position_size is None:
                    # ========== èˆŠçš„è¨‚å–®å°å‘æ¨¡å¼ï¼ˆå‘å¾Œå…¼å®¹ï¼‰ ==========
                    side = signal.side.upper()
                    if not bot.use_signal_side and bot.fixed_side:
                        side = bot.fixed_side.upper()
                    
                    if side not in ["BUY", "SELL"]:
                        raise ValueError(f"ç„¡æ•ˆçš„ä¸‹å–®æ–¹å‘: {side}")
                    
                    # è¨ˆç®— qtyï¼šä½¿ç”¨ bot è¨­å®šï¼ˆmax_invest_usdt æˆ– qtyï¼‰
                    # å¦‚æœè¨­å®šäº† max_invest_usdtï¼Œå‰‡æ ¹æ“šç•¶å‰åƒ¹æ ¼è¨ˆç®—ï¼›å¦å‰‡ä½¿ç”¨ bot.qty
                    qty = calculate_qty_from_max_invest(bot, symbol, target_qty=None)
                    
                    # è¨­å®šæ æ¡¿
                    try:
                        client.futures_change_leverage(symbol=symbol, leverage=bot.leverage)
                        logger.info(f"æˆåŠŸè¨­å®š {symbol} æ æ¡¿ç‚º {bot.leverage}x")
                    except Exception as e:
                        logger.warning(f"è¨­å®šæ æ¡¿æ™‚ç™¼ç”Ÿè­¦å‘Š: {e}")
                    
                    # ä¸‹å¸‚åƒ¹å–®
                    timestamp = int(time.time() * 1000)
                    client_order_id = f"bot_{bot.id}_{timestamp}"
                    
                    logger.info(f"Bot {bot.id} ({bot.name}) ä¸‹å–®ï¼ˆè¨‚å–®å°å‘ï¼‰: {symbol} {side} {qty}")
                    
                    # æ ¼å¼åŒ–æ•¸é‡ä»¥ç¬¦åˆ Binance ç²¾åº¦è¦æ±‚
                    formatted_qty = format_quantity(symbol, qty)
                    formatted_qty_float = float(formatted_qty)
                    
                    order = client.futures_create_order(
                        symbol=symbol,
                        side="BUY" if side == "BUY" else "SELL",
                        type="MARKET",
                        quantity=formatted_qty,
                        newClientOrderId=client_order_id
                    )
                    
                    logger.info(f"Bot {bot.id} æˆåŠŸä¸‹å–®: {order.get('orderId')}")
                    
                    # å–å¾— entry_price
                    entry_price = None
                    if order.get("avgPrice"):
                        try:
                            avg_price = float(order["avgPrice"])
                            if avg_price > 0:
                                entry_price = avg_price
                        except (ValueError, TypeError):
                            pass
                    
                    if entry_price is None or entry_price <= 0:
                        if order.get("price"):
                            try:
                                price = float(order["price"])
                                if price > 0:
                                    entry_price = price
                            except (ValueError, TypeError):
                                pass
                    
                    if entry_price is None or entry_price <= 0:
                        try:
                            entry_price = get_mark_price(symbol)
                            if entry_price <= 0:
                                logger.warning(f"Bot {bot.id} ä¸‹å–® {symbol} å¾Œï¼Œç„¡æ³•å–å¾—æœ‰æ•ˆçš„ entry_price")
                        except Exception as e:
                            logger.error(f"Bot {bot.id} ä¸‹å–® {symbol} å¾Œï¼Œå–å¾— mark_price å¤±æ•—: {e}")
                            entry_price = 0.0
                    
                    position_side = "LONG" if side == "BUY" else "SHORT"
                    
                    # è¨­å®š trail_callbackï¼ˆåªä½¿ç”¨ bot çš„è¦†å¯«å€¼ï¼Œå¦å‰‡è¨­ç‚º None ä»¥ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼‰
                    trail_callback = None
                    if bot.use_dynamic_stop and bot.trailing_callback_percent is not None:
                        # Bot æœ‰æ˜ç¢ºçš„è¦†å¯«å€¼ï¼Œä½¿ç”¨å®ƒ
                        trail_callback = bot.trailing_callback_percent / 100.0
                    # å¦å‰‡è¨­ç‚º Noneï¼Œè®“ä½ç½®ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼ˆä¸å„²å­˜å…¨å±€å€¼ä½œç‚ºè¦†å¯«ï¼‰
                    
                    # å»ºç«‹ Position è¨˜éŒ„
                    position = Position(
                        bot_id=bot.id,
                        tv_signal_log_id=log.id,
                        symbol=symbol,
                        side=position_side,
                        qty=formatted_qty_float,
                        status="OPEN",
                        binance_order_id=int(order.get("orderId")) if order.get("orderId") else None,
                        client_order_id=order.get("clientOrderId"),
                        entry_price=entry_price,
                        trail_callback=trail_callback,
                        highest_price=entry_price if trail_callback else None,
                    )
                    
                    db.add(position)
                    db.commit()
                    db.refresh(position)
                    
                    results.append(f"bot={bot.id}, position_id={position.id}, mode=order_based")
                    logger.info(f"Bot {bot.id} æˆåŠŸå»ºç«‹ Position {position.id} (è¨‚å–®å°å‘)")
                    
                else:
                    # ========== æ–°çš„ä½ç½®å°å‘æ¨¡å¼ ==========
                    # æ­£è¦åŒ– targetï¼ˆå°‡éå¸¸å°çš„å€¼è¦–ç‚º 0ï¼‰
                    if abs(target_position_size) < EPS:
                        target_position_size = 0.0
                    
                    # å–å¾—ç•¶å‰å€‰ä½
                    current_position, current_qty_signed = get_current_position_signed_qty(db, bot.id, symbol)
                    
                    logger.info(
                        f"Bot {bot.id} ({bot.name}) ä½ç½®å°å‘æ¨¡å¼: "
                        f"symbol={symbol}, target_direction={'LONG' if target_position_size > 0 else 'SHORT' if target_position_size < 0 else 'CLOSE'}, "
                        f"current={current_qty_signed}, qtyå°‡ç”±botè¨­å®šè¨ˆç®—(max_invest_usdt={bot.max_invest_usdt}, qty={bot.qty})"
                    )
                    
                    # è¨­å®šæ æ¡¿ï¼ˆåœ¨æ±ºå®šæ“ä½œå‰å…ˆè¨­å®šï¼‰
                    try:
                        client.futures_change_leverage(symbol=symbol, leverage=bot.leverage)
                        logger.info(f"æˆåŠŸè¨­å®š {symbol} æ æ¡¿ç‚º {bot.leverage}x")
                    except Exception as e:
                        logger.warning(f"è¨­å®šæ æ¡¿æ™‚ç™¼ç”Ÿè­¦å‘Š: {e}")
                    
                    # Case A: ç›®æ¨™ç‚ºå¹³å€‰ (target == 0)
                    if abs(target_position_size) < EPS:
                        if abs(current_qty_signed) < EPS:
                            # å·²ç¶“å¹³å€‰ï¼Œç„¡éœ€æ“ä½œ
                            results.append(f"bot={bot.id}, result=flat_no_position")
                            logger.info(f"Bot {bot.id} ç›®æ¨™ç‚ºå¹³å€‰ï¼Œç•¶å‰å·²ç„¡å€‰ä½ï¼Œç„¡éœ€æ“ä½œ")
                        else:
                            # é—œé–‰ç¾æœ‰å€‰ä½
                            if not current_position:
                                results.append(f"bot={bot.id}, error=current_position_not_found")
                                logger.warning(f"Bot {bot.id} ç›®æ¨™ç‚ºå¹³å€‰ï¼Œä½†æ‰¾ä¸åˆ°ç•¶å‰å€‰ä½è¨˜éŒ„")
                            else:
                                # æª¢æŸ¥ tv_signal_close_enabled æ¨™èªŒ
                                if not current_position.tv_signal_close_enabled:
                                    results.append(f"bot={bot.id}, result=tv_signal_close_disabled, position_id={current_position.id}")
                                    logger.info(
                                        f"Bot {bot.id} æ”¶åˆ° TradingView å¹³å€‰è¨Šè™Ÿï¼Œä½†å€‰ä½ {current_position.id} "
                                        f"tv_signal_close_enabled=Falseï¼Œè·³éé—œå€‰"
                                    )
                                else:
                                    try:
                                        close_order = close_futures_position(
                                            symbol=symbol,
                                            position_side=current_position.side,
                                            qty=current_position.qty,
                                            position_id=current_position.id
                                        )
                                        
                                        # ä½¿ç”¨çµ±ä¸€çš„å‡½æ•¸å–å¾— exit_priceï¼ˆå„ªå…ˆä½¿ç”¨ avgPriceï¼‰
                                        exit_price = get_exit_price_from_order(close_order, symbol)
                                        
                                        current_position.status = "CLOSED"
                                        current_position.closed_at = datetime.now(timezone.utc)
                                        current_position.exit_price = exit_price
                                        current_position.exit_reason = "tv_exit"
                                        db.commit()
                                        
                                        results.append(f"bot={bot.id}, closed_position_id={current_position.id}")
                                        logger.info(f"Bot {bot.id} æˆåŠŸå¹³å€‰ Position {current_position.id}")
                                    except Exception as e:
                                        error_msg = f"bot={bot.id}, error=close_failed: {str(e)}"
                                        logger.exception(f"Bot {bot.id} å¹³å€‰å¤±æ•—: {e}")
                                        results.append(error_msg)
                    
                    # Case B: ç›®æ¨™ç‚ºå¤šå€‰ (target > 0)
                    elif target_position_size > 0:
                        # å¿½ç•¥ position_size çš„æ•¸é‡å€¼ï¼Œæ”¹ç‚ºä½¿ç”¨ bot è¨­å®šè¨ˆç®—æ•¸é‡
                        # ä½¿ç”¨ bot.max_invest_usdt è¨ˆç®—ï¼Œå¦‚æœæœªè¨­å®šå‰‡ä½¿ç”¨ bot.qty
                        target_qty = calculate_qty_from_max_invest(bot, symbol, target_qty=None)
                        
                        if current_qty_signed > 0:
                            # å·²ç¶“æ˜¯å¤šå€‰
                            diff = target_qty - current_qty_signed
                            if abs(diff) < EPS:
                                # æ•¸é‡å·²åŒ¹é…ï¼Œç„¡éœ€æ“ä½œ
                                results.append(f"bot={bot.id}, result=long_qty_already_match, qty={current_qty_signed}")
                                logger.info(f"Bot {bot.id} ç›®æ¨™å¤šå€‰æ•¸é‡å·²åŒ¹é…ï¼Œç„¡éœ€æ“ä½œ")
                            else:
                                # éœ€è¦èª¿æ•´æ•¸é‡ï¼ˆç°¡åŒ–ï¼šå…ˆé—œé–‰å†é–‹æ–°å€‰ï¼Œæˆ–å¯ä»¥å¯¦ä½œéƒ¨åˆ†å¹³å€‰/åŠ å€‰ï¼‰
                                # é€™è£¡æ¡ç”¨ç°¡å–®ç­–ç•¥ï¼šå¦‚æœå·®ç•°å¤§æ–¼ 10%ï¼Œå‰‡é‡æ–°é–‹å€‰
                                if abs(diff) / current_qty_signed > 0.1:
                                    # å…ˆé—œé–‰ç¾æœ‰å€‰ä½
                                    if current_position:
                                        try:
                                            close_order = close_futures_position(
                                                symbol=symbol,
                                                position_side=current_position.side,
                                                qty=current_position.qty,
                                                position_id=current_position.id
                                            )
                                            # ä½¿ç”¨çµ±ä¸€çš„å‡½æ•¸å–å¾— exit_priceï¼ˆå„ªå…ˆä½¿ç”¨ avgPriceï¼‰
                                            exit_price = get_exit_price_from_order(close_order, symbol)
                                            current_position.status = "CLOSED"
                                            current_position.closed_at = datetime.now(timezone.utc)
                                            current_position.exit_price = exit_price
                                            current_position.exit_reason = "tv_rebalance"
                                            db.commit()
                                        except Exception as e:
                                            logger.exception(f"Bot {bot.id} èª¿æ•´å¤šå€‰æ™‚é—œé–‰èˆŠå€‰å¤±æ•—: {e}")
                                    
                                    # é–‹æ–°å¤šå€‰
                                    try:
                                        # æ ¼å¼åŒ–æ•¸é‡ä»¥ç¬¦åˆ Binance ç²¾åº¦è¦æ±‚
                                        formatted_qty = format_quantity(symbol, target_qty)
                                        formatted_qty_float = float(formatted_qty)
                                        
                                        timestamp = int(time.time() * 1000)
                                        client_order_id = f"bot_{bot.id}_{timestamp}"
                                        order = client.futures_create_order(
                                            symbol=symbol,
                                            side="BUY",
                                            type="MARKET",
                                            quantity=formatted_qty,
                                            newClientOrderId=client_order_id
                                        )
                                        
                                        entry_price = float(order.get("avgPrice", 0)) or get_mark_price(symbol) or 0.0
                                        
                                        # è¨­å®š trail_callbackï¼ˆåªä½¿ç”¨ bot çš„è¦†å¯«å€¼ï¼Œå¦å‰‡è¨­ç‚º None ä»¥ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼‰
                                        trail_callback = None
                                        if bot.use_dynamic_stop and bot.trailing_callback_percent is not None:
                                            trail_callback = bot.trailing_callback_percent / 100.0
                                        # å¦å‰‡è¨­ç‚º Noneï¼Œè®“ä½ç½®ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼ˆä¸å„²å­˜å…¨å±€å€¼ä½œç‚ºè¦†å¯«ï¼‰
                                        
                                        position = Position(
                                            bot_id=bot.id,
                                            tv_signal_log_id=log.id,
                                            symbol=symbol,
                                            side="LONG",
                                            qty=formatted_qty_float,
                                            status="OPEN",
                                            binance_order_id=int(order.get("orderId")) if order.get("orderId") else None,
                                            client_order_id=order.get("clientOrderId"),
                                            entry_price=entry_price,
                                            trail_callback=trail_callback,
                                            highest_price=entry_price if trail_callback else None,
                                        )
                                        db.add(position)
                                        db.commit()
                                        db.refresh(position)
                                        
                                        results.append(f"bot={bot.id}, position_id={position.id}, result=rebalance_long")
                                        logger.info(f"Bot {bot.id} æˆåŠŸèª¿æ•´å¤šå€‰è‡³ {target_qty}")
                                    except Exception as e:
                                        error_msg = f"bot={bot.id}, error=rebalance_long_failed: {str(e)}"
                                        logger.exception(f"Bot {bot.id} èª¿æ•´å¤šå€‰å¤±æ•—: {e}")
                                        results.append(error_msg)
                                else:
                                    results.append(f"bot={bot.id}, result=long_qty_diff_small, diff={diff}")
                                    logger.info(f"Bot {bot.id} å¤šå€‰æ•¸é‡å·®ç•°å°æ–¼ 10%ï¼Œè·³éèª¿æ•´")
                        elif current_qty_signed < 0:
                            # ç•¶å‰æ˜¯ç©ºå€‰ï¼Œéœ€è¦åè½‰ç‚ºå¤šå€‰
                            # å…ˆé—œé–‰ç©ºå€‰
                            if current_position:
                                try:
                                    close_order = close_futures_position(
                                        symbol=symbol,
                                        position_side=current_position.side,
                                        qty=current_position.qty,
                                        position_id=current_position.id
                                    )
                                    exit_price = get_mark_price(symbol) if not close_order.get("avgPrice") else float(close_order["avgPrice"])
                                    current_position.status = "CLOSED"
                                    current_position.closed_at = datetime.now(timezone.utc)
                                    current_position.exit_price = exit_price
                                    current_position.exit_reason = "tv_reverse_to_long"
                                    db.commit()
                                except Exception as e:
                                    logger.exception(f"Bot {bot.id} åè½‰å€‰ä½æ™‚é—œé–‰ç©ºå€‰å¤±æ•—: {e}")
                            
                            # é–‹æ–°å¤šå€‰
                            try:
                                # æ ¼å¼åŒ–æ•¸é‡ä»¥ç¬¦åˆ Binance ç²¾åº¦è¦æ±‚
                                formatted_qty = format_quantity(symbol, target_qty)
                                formatted_qty_float = float(formatted_qty)
                                
                                timestamp = int(time.time() * 1000)
                                client_order_id = f"bot_{bot.id}_{timestamp}"
                                order = client.futures_create_order(
                                    symbol=symbol,
                                    side="BUY",
                                    type="MARKET",
                                    quantity=formatted_qty,
                                    newClientOrderId=client_order_id
                                )
                                
                                entry_price = float(order.get("avgPrice", 0)) or get_mark_price(symbol) or 0.0
                                
                                # è¨­å®š trail_callbackï¼ˆåªä½¿ç”¨ bot çš„è¦†å¯«å€¼ï¼Œå¦å‰‡è¨­ç‚º None ä»¥ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼‰
                                trail_callback = None
                                if bot.use_dynamic_stop and bot.trailing_callback_percent is not None:
                                    trail_callback = bot.trailing_callback_percent / 100.0
                                # å¦å‰‡è¨­ç‚º Noneï¼Œè®“ä½ç½®ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼ˆä¸å„²å­˜å…¨å±€å€¼ä½œç‚ºè¦†å¯«ï¼‰
                                
                                position = Position(
                                    bot_id=bot.id,
                                    tv_signal_log_id=log.id,
                                    symbol=symbol,
                                    side="LONG",
                                    qty=formatted_qty_float,
                                    status="OPEN",
                                    binance_order_id=int(order.get("orderId")) if order.get("orderId") else None,
                                    client_order_id=order.get("clientOrderId"),
                                    entry_price=entry_price,
                                    trail_callback=trail_callback,
                                    highest_price=entry_price if trail_callback else None,
                                )
                                db.add(position)
                                db.commit()
                                db.refresh(position)
                                
                                results.append(f"bot={bot.id}, position_id={position.id}, result=reverse_short_to_long")
                                logger.info(f"Bot {bot.id} æˆåŠŸåè½‰ç©ºå€‰ç‚ºå¤šå€‰ {target_qty}")
                            except Exception as e:
                                error_msg = f"bot={bot.id}, error=reverse_to_long_failed: {str(e)}"
                                logger.exception(f"Bot {bot.id} åè½‰ç‚ºå¤šå€‰å¤±æ•—: {e}")
                                results.append(error_msg)
                        else:
                            # ç•¶å‰ç„¡å€‰ä½ï¼Œé–‹æ–°å¤šå€‰
                            try:
                                # æ ¼å¼åŒ–æ•¸é‡ä»¥ç¬¦åˆ Binance ç²¾åº¦è¦æ±‚
                                formatted_qty = format_quantity(symbol, target_qty)
                                formatted_qty_float = float(formatted_qty)
                                
                                timestamp = int(time.time() * 1000)
                                client_order_id = f"bot_{bot.id}_{timestamp}"
                                order = client.futures_create_order(
                                    symbol=symbol,
                                    side="BUY",
                                    type="MARKET",
                                    quantity=formatted_qty,
                                    newClientOrderId=client_order_id
                                )
                                
                                entry_price = float(order.get("avgPrice", 0)) or get_mark_price(symbol) or 0.0
                                
                                # è¨­å®š trail_callbackï¼ˆåªä½¿ç”¨ bot çš„è¦†å¯«å€¼ï¼Œå¦å‰‡è¨­ç‚º None ä»¥ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼‰
                                trail_callback = None
                                if bot.use_dynamic_stop and bot.trailing_callback_percent is not None:
                                    trail_callback = bot.trailing_callback_percent / 100.0
                                # å¦å‰‡è¨­ç‚º Noneï¼Œè®“ä½ç½®ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼ˆä¸å„²å­˜å…¨å±€å€¼ä½œç‚ºè¦†å¯«ï¼‰
                                
                                position = Position(
                                    bot_id=bot.id,
                                    tv_signal_log_id=log.id,
                                    symbol=symbol,
                                    side="LONG",
                                    qty=formatted_qty_float,
                                    status="OPEN",
                                    binance_order_id=int(order.get("orderId")) if order.get("orderId") else None,
                                    client_order_id=order.get("clientOrderId"),
                                    entry_price=entry_price,
                                    trail_callback=trail_callback,
                                    highest_price=entry_price if trail_callback else None,
                                )
                                db.add(position)
                                db.commit()
                                db.refresh(position)
                                
                                results.append(f"bot={bot.id}, position_id={position.id}, result=open_long")
                                logger.info(f"Bot {bot.id} æˆåŠŸé–‹å¤šå€‰ {target_qty}")
                            except Exception as e:
                                error_msg = f"bot={bot.id}, error=open_long_failed: {str(e)}"
                                logger.exception(f"Bot {bot.id} é–‹å¤šå€‰å¤±æ•—: {e}")
                                results.append(error_msg)
                    
                    # Case C: ç›®æ¨™ç‚ºç©ºå€‰ (target < 0)
                    else:
                        # å¿½ç•¥ position_size çš„æ•¸é‡å€¼ï¼Œæ”¹ç‚ºä½¿ç”¨ bot è¨­å®šè¨ˆç®—æ•¸é‡
                        # ä½¿ç”¨ bot.max_invest_usdt è¨ˆç®—ï¼Œå¦‚æœæœªè¨­å®šå‰‡ä½¿ç”¨ bot.qty
                        target_qty = calculate_qty_from_max_invest(bot, symbol, target_qty=None)
                        
                        if current_qty_signed < 0:
                            # å·²ç¶“æ˜¯ç©ºå€‰
                            diff = target_qty - abs(current_qty_signed)
                            if abs(diff) < EPS:
                                results.append(f"bot={bot.id}, result=short_qty_already_match, qty={abs(current_qty_signed)}")
                                logger.info(f"Bot {bot.id} ç›®æ¨™ç©ºå€‰æ•¸é‡å·²åŒ¹é…ï¼Œç„¡éœ€æ“ä½œ")
                            else:
                                # éœ€è¦èª¿æ•´ï¼ˆç°¡åŒ–ï¼šå·®ç•°å¤§æ–¼ 10% å‰‡é‡æ–°é–‹å€‰ï¼‰
                                if abs(diff) / abs(current_qty_signed) > 0.1:
                                    if current_position:
                                        try:
                                            close_order = close_futures_position(
                                                symbol=symbol,
                                                position_side=current_position.side,
                                                qty=current_position.qty,
                                                position_id=current_position.id
                                            )
                                            # ä½¿ç”¨çµ±ä¸€çš„å‡½æ•¸å–å¾— exit_priceï¼ˆå„ªå…ˆä½¿ç”¨ avgPriceï¼‰
                                            exit_price = get_exit_price_from_order(close_order, symbol)
                                            current_position.status = "CLOSED"
                                            current_position.closed_at = datetime.now(timezone.utc)
                                            current_position.exit_price = exit_price
                                            current_position.exit_reason = "tv_rebalance"
                                            db.commit()
                                        except Exception as e:
                                            logger.exception(f"Bot {bot.id} èª¿æ•´ç©ºå€‰æ™‚é—œé–‰èˆŠå€‰å¤±æ•—: {e}")
                                    
                                    try:
                                        # æ ¼å¼åŒ–æ•¸é‡ä»¥ç¬¦åˆ Binance ç²¾åº¦è¦æ±‚
                                        formatted_qty = format_quantity(symbol, target_qty)
                                        formatted_qty_float = float(formatted_qty)
                                        
                                        timestamp = int(time.time() * 1000)
                                        client_order_id = f"bot_{bot.id}_{timestamp}"
                                        order = client.futures_create_order(
                                            symbol=symbol,
                                            side="SELL",
                                            type="MARKET",
                                            quantity=formatted_qty,
                                            newClientOrderId=client_order_id
                                        )
                                        
                                        entry_price = float(order.get("avgPrice", 0)) or get_mark_price(symbol) or 0.0
                                        
                                        # è¨­å®š trail_callbackï¼ˆåªä½¿ç”¨ bot çš„è¦†å¯«å€¼ï¼Œå¦å‰‡è¨­ç‚º None ä»¥ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼‰
                                        trail_callback = None
                                        if bot.use_dynamic_stop and bot.trailing_callback_percent is not None:
                                            trail_callback = bot.trailing_callback_percent / 100.0
                                        # å¦å‰‡è¨­ç‚º Noneï¼Œè®“ä½ç½®ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼ˆä¸å„²å­˜å…¨å±€å€¼ä½œç‚ºè¦†å¯«ï¼‰
                                        
                                        position = Position(
                                            bot_id=bot.id,
                                            tv_signal_log_id=log.id,
                                            symbol=symbol,
                                            side="SHORT",
                                            qty=formatted_qty_float,
                                            status="OPEN",
                                            binance_order_id=int(order.get("orderId")) if order.get("orderId") else None,
                                            client_order_id=order.get("clientOrderId"),
                                            entry_price=entry_price,
                                            trail_callback=trail_callback,
                                            highest_price=entry_price if trail_callback else None,
                                        )
                                        db.add(position)
                                        db.commit()
                                        db.refresh(position)
                                        
                                        results.append(f"bot={bot.id}, position_id={position.id}, result=rebalance_short")
                                        logger.info(f"Bot {bot.id} æˆåŠŸèª¿æ•´ç©ºå€‰è‡³ {target_qty}")
                                    except Exception as e:
                                        error_msg = f"bot={bot.id}, error=rebalance_short_failed: {str(e)}"
                                        logger.exception(f"Bot {bot.id} èª¿æ•´ç©ºå€‰å¤±æ•—: {e}")
                                        results.append(error_msg)
                                else:
                                    results.append(f"bot={bot.id}, result=short_qty_diff_small, diff={diff}")
                                    logger.info(f"Bot {bot.id} ç©ºå€‰æ•¸é‡å·®ç•°å°æ–¼ 10%ï¼Œè·³éèª¿æ•´")
                        elif current_qty_signed > 0:
                            # ç•¶å‰æ˜¯å¤šå€‰ï¼Œéœ€è¦åè½‰ç‚ºç©ºå€‰
                            if current_position:
                                try:
                                    close_order = close_futures_position(
                                        symbol=symbol,
                                        position_side=current_position.side,
                                        qty=current_position.qty,
                                        position_id=current_position.id
                                    )
                                    exit_price = get_mark_price(symbol) if not close_order.get("avgPrice") else float(close_order["avgPrice"])
                                    current_position.status = "CLOSED"
                                    current_position.closed_at = datetime.now(timezone.utc)
                                    current_position.exit_price = exit_price
                                    current_position.exit_reason = "tv_reverse_to_short"
                                    db.commit()
                                except Exception as e:
                                    logger.exception(f"Bot {bot.id} åè½‰å€‰ä½æ™‚é—œé–‰å¤šå€‰å¤±æ•—: {e}")
                            
                            try:
                                # æ ¼å¼åŒ–æ•¸é‡ä»¥ç¬¦åˆ Binance ç²¾åº¦è¦æ±‚
                                formatted_qty = format_quantity(symbol, target_qty)
                                formatted_qty_float = float(formatted_qty)
                                
                                timestamp = int(time.time() * 1000)
                                client_order_id = f"bot_{bot.id}_{timestamp}"
                                order = client.futures_create_order(
                                    symbol=symbol,
                                    side="SELL",
                                    type="MARKET",
                                    quantity=formatted_qty,
                                    newClientOrderId=client_order_id
                                )
                                
                                entry_price = float(order.get("avgPrice", 0)) or get_mark_price(symbol) or 0.0
                                
                                # è¨­å®š trail_callbackï¼ˆåªä½¿ç”¨ bot çš„è¦†å¯«å€¼ï¼Œå¦å‰‡è¨­ç‚º None ä»¥ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼‰
                                trail_callback = None
                                if bot.use_dynamic_stop and bot.trailing_callback_percent is not None:
                                    trail_callback = bot.trailing_callback_percent / 100.0
                                # å¦å‰‡è¨­ç‚º Noneï¼Œè®“ä½ç½®ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼ˆä¸å„²å­˜å…¨å±€å€¼ä½œç‚ºè¦†å¯«ï¼‰
                                
                                position = Position(
                                    bot_id=bot.id,
                                    tv_signal_log_id=log.id,
                                    symbol=symbol,
                                    side="SHORT",
                                    qty=formatted_qty_float,
                                    status="OPEN",
                                    binance_order_id=int(order.get("orderId")) if order.get("orderId") else None,
                                    client_order_id=order.get("clientOrderId"),
                                    entry_price=entry_price,
                                    trail_callback=trail_callback,
                                    highest_price=entry_price if trail_callback else None,
                                )
                                db.add(position)
                                db.commit()
                                db.refresh(position)
                                
                                results.append(f"bot={bot.id}, position_id={position.id}, result=reverse_long_to_short")
                                logger.info(f"Bot {bot.id} æˆåŠŸåè½‰å¤šå€‰ç‚ºç©ºå€‰ {target_qty}")
                            except Exception as e:
                                error_msg = f"bot={bot.id}, error=reverse_to_short_failed: {str(e)}"
                                logger.exception(f"Bot {bot.id} åè½‰ç‚ºç©ºå€‰å¤±æ•—: {e}")
                                results.append(error_msg)
                        else:
                            # ç•¶å‰ç„¡å€‰ä½ï¼Œé–‹æ–°ç©ºå€‰
                            try:
                                # æ ¼å¼åŒ–æ•¸é‡ä»¥ç¬¦åˆ Binance ç²¾åº¦è¦æ±‚
                                formatted_qty = format_quantity(symbol, target_qty)
                                formatted_qty_float = float(formatted_qty)
                                
                                timestamp = int(time.time() * 1000)
                                client_order_id = f"bot_{bot.id}_{timestamp}"
                                order = client.futures_create_order(
                                    symbol=symbol,
                                    side="SELL",
                                    type="MARKET",
                                    quantity=formatted_qty,
                                    newClientOrderId=client_order_id
                                )
                                
                                entry_price = float(order.get("avgPrice", 0)) or get_mark_price(symbol) or 0.0
                                
                                # è¨­å®š trail_callbackï¼ˆåªä½¿ç”¨ bot çš„è¦†å¯«å€¼ï¼Œå¦å‰‡è¨­ç‚º None ä»¥ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼‰
                                trail_callback = None
                                if bot.use_dynamic_stop and bot.trailing_callback_percent is not None:
                                    trail_callback = bot.trailing_callback_percent / 100.0
                                # å¦å‰‡è¨­ç‚º Noneï¼Œè®“ä½ç½®ä½¿ç”¨å…¨å±€é…ç½®å‹•æ…‹ï¼ˆä¸å„²å­˜å…¨å±€å€¼ä½œç‚ºè¦†å¯«ï¼‰
                                
                                position = Position(
                                    bot_id=bot.id,
                                    tv_signal_log_id=log.id,
                                    symbol=symbol,
                                    side="SHORT",
                                    qty=formatted_qty_float,
                                    status="OPEN",
                                    binance_order_id=int(order.get("orderId")) if order.get("orderId") else None,
                                    client_order_id=order.get("clientOrderId"),
                                    entry_price=entry_price,
                                    trail_callback=trail_callback,
                                    highest_price=entry_price if trail_callback else None,
                                )
                                db.add(position)
                                db.commit()
                                db.refresh(position)
                                
                                results.append(f"bot={bot.id}, position_id={position.id}, result=open_short")
                                logger.info(f"Bot {bot.id} æˆåŠŸé–‹ç©ºå€‰ {target_qty}")
                            except Exception as e:
                                error_msg = f"bot={bot.id}, error=open_short_failed: {str(e)}"
                                logger.exception(f"Bot {bot.id} é–‹ç©ºå€‰å¤±æ•—: {e}")
                                results.append(error_msg)
                
            except Exception as e:
                error_msg = f"bot={bot.id}, error={str(e)}"
                logger.exception(f"Bot {bot.id} ({bot.name}) è™•ç†å¤±æ•—: {e}")
                results.append(error_msg)
                # ç¹¼çºŒè™•ç†ä¸‹ä¸€å€‹ bot
        
        # 4) æ›´æ–° log çš„ processed å’Œ process_resultï¼ˆç¸½æ˜¯åŸ·è¡Œï¼‰
        log.processed = True
        log.process_result = "; ".join(results)
        db.commit()
        
        return {
            "success": True,
            "message": "è¨‚å–®å»ºç«‹æˆåŠŸ",
            "signal_log_id": log.id,
            "results": results
        }
    
    except HTTPException as http_err:
        # å¦‚æœæ˜¯ HTTPExceptionï¼Œä¸”æœ‰ logï¼Œå˜—è©¦æ›´æ–° log
        if log:
            try:
                log.processed = True
                log.process_result = f"HTTPException: {str(http_err.detail)}"
                db.commit()
            except:
                pass
        raise
    except Exception as e:
        # å…¶ä»–éŒ¯èª¤ï¼Œå¦‚æœæœ‰ logï¼Œæ›´æ–° log
        if log:
            try:
                log.processed = True
                log.process_result = f"ERROR: {str(e)}"
                db.commit()
            except:
                pass
        logger.exception(f"è™•ç† TradingView webhook æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        raise HTTPException(status_code=500, detail=f"è™•ç†å¤±æ•—: {str(e)}")




@app.get("/positions", response_model=List[PositionOut])
async def get_positions(
    user: dict = Depends(require_admin_user),
    symbol: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    æŸ¥è©¢æ‰€æœ‰å€‰ä½è¨˜éŒ„
    
    æ­¤ç«¯é»åƒ…é™å·²ç™»å…¥ä¸”é€šéç®¡ç†å“¡é©—è­‰çš„ä½¿ç”¨è€…ï¼ˆGoogle OAuth + ADMIN_GOOGLE_EMAILï¼‰ä½¿ç”¨ã€‚
    æ”¯æ´ä¾äº¤æ˜“å°ã€ç‹€æ…‹ã€æ—¥æœŸç¯„åœç¯©é¸ã€‚
    
    Args:
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
        symbol: äº¤æ˜“å°ç¯©é¸ï¼ˆå¯é¸ï¼‰
        status: ç‹€æ…‹ç¯©é¸ï¼ˆå¯é¸ï¼‰
        start_date: é–‹å§‹æ—¥æœŸï¼ˆYYYY-MM-DDæ ¼å¼ï¼Œå¯é¸ï¼‰
        end_date: çµæŸæ—¥æœŸï¼ˆYYYY-MM-DDæ ¼å¼ï¼Œå¯é¸ï¼‰
        db: è³‡æ–™åº« Session
    
    Returns:
        List[PositionOut]: å€‰ä½è¨˜éŒ„åˆ—è¡¨
    """
    query = db.query(Position)
    
    if symbol:
        query = query.filter(Position.symbol == symbol.upper())
    
    if status:
        query = query.filter(Position.status == status.upper())
    
    if start_date:
        try:
            # Parse date string and create datetime at start of day in UTC
            start_datetime = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            # Filter positions where created_at is on or after the start date
            # Compare datetime directly (SQLAlchemy handles timezone-aware comparisons)
            query = query.filter(Position.created_at >= start_datetime)
            logger.debug(f"Date filter: start_date={start_date}, start_datetime={start_datetime}")
        except ValueError as e:
            logger.warning(f"Invalid start_date format: {start_date}, error: {e}")
    
    if end_date:
        try:
            # Parse date string - end_date is inclusive, so include the entire day
            # Add 1 day and use < operator to include the entire end_date
            end_datetime = datetime.strptime(end_date, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
            # Filter positions where created_at is before the start of the next day
            query = query.filter(Position.created_at < end_datetime)
            logger.debug(f"Date filter: end_date={end_date}, end_datetime={end_datetime}")
        except ValueError as e:
            logger.warning(f"Invalid end_date format: {end_date}, error: {e}")
    
    positions = query.order_by(Position.created_at.desc()).all()
    
    # è¨ˆç®—æ¯å€‹ position çš„å¯¦éš›ä½¿ç”¨çš„å€¼å’Œä¾†æºæ¨™è¨˜
    result = []
    for pos in positions:
        pos_dict = pos.to_dict()
        
        # å–å¾—å°æ‡‰æ–¹å‘çš„å…¨å±€è¨­å®š
        side_config = TRAILING_CONFIG.get_config_for_side(pos.side)
        
        # è¨ˆç®—å¯¦éš›ä½¿ç”¨çš„å€¼å’Œä¾†æºæ¨™è¨˜
        # Profit Threshold
        profit_threshold_value = None
        profit_threshold_source = None
        if pos.dyn_profit_threshold_pct is not None:
            profit_threshold_value = pos.dyn_profit_threshold_pct
            profit_threshold_source = "override"  # position çš„è¨­å®šè¦–ç‚º override
        elif side_config.profit_threshold_pct is not None:
            profit_threshold_value = side_config.profit_threshold_pct
            profit_threshold_source = "global"
        else:
            profit_threshold_value = DYN_PROFIT_THRESHOLD_PCT
            profit_threshold_source = "default"
        
        # Lock Ratio
        lock_ratio_value = None
        lock_ratio_source = None
        if pos.trail_callback is not None:
            lock_ratio_value = pos.trail_callback
            lock_ratio_source = "override"  # position çš„è¨­å®šè¦–ç‚º override
        elif side_config.lock_ratio is not None:
            lock_ratio_value = side_config.lock_ratio
            lock_ratio_source = "global"
        else:
            lock_ratio_value = DYN_LOCK_RATIO_DEFAULT
            lock_ratio_source = "default"
        
        # Base SL%
        base_sl_value = None
        base_sl_source = None
        if pos.base_stop_loss_pct is not None:
            base_sl_value = pos.base_stop_loss_pct
            base_sl_source = "override"  # position çš„è¨­å®šè¦–ç‚º override
        elif side_config.base_sl_pct is not None:
            base_sl_value = side_config.base_sl_pct
            base_sl_source = "global"
        else:
            base_sl_value = DYN_BASE_SL_PCT
            base_sl_source = "default"
        
        # è¨ˆç®—åœæç‹€æ…‹ï¼ˆåƒ…å° OPEN ç‹€æ…‹çš„å€‰ä½ï¼‰
        stop_mode = None
        base_stop_price = None
        dynamic_stop_price = None
        if pos.status == "OPEN" and pos.entry_price and pos.entry_price > 0:
            try:
                # ç²å–ç•¶å‰æ¨™è¨˜åƒ¹æ ¼
                current_mark_price = get_mark_price(pos.symbol)
                if current_mark_price and current_mark_price > 0:
                    # è¨ˆç®— unrealized_pnl_pctï¼ˆPnL%ï¼‰
                    calculated_unrealized_pnl_pct = None
                    if pos.entry_price > 0 and pos.qty > 0:
                        # ç²å– leverage
                        leverage = 20  # é»˜èªæ æ¡¿
                        if pos.bot_id:
                            try:
                                from models import BotConfig
                                bot = db.query(BotConfig).filter(BotConfig.id == pos.bot_id).first()
                                if bot and bot.leverage:
                                    leverage = bot.leverage
                            except Exception:
                                pass
                        # è¨ˆç®— unrealized PnL
                        if pos.side == "LONG":
                            unrealized_pnl_amount = (current_mark_price - pos.entry_price) * pos.qty
                        else:  # SHORT
                            unrealized_pnl_amount = (pos.entry_price - current_mark_price) * pos.qty
                        # è¨ˆç®—ä¿è­‰é‡‘
                        notional = pos.entry_price * pos.qty
                        if notional > 0 and leverage > 0:
                            margin = notional / leverage
                            if margin > 0:
                                calculated_unrealized_pnl_pct = (unrealized_pnl_amount / margin) * 100.0
                    
                    # ä½¿ç”¨ compute_stop_state è¨ˆç®—åœæç‹€æ…‹ï¼ˆå‚³å…¥ leverage å’Œ qtyï¼‰
                    stop_state = compute_stop_state(pos, current_mark_price, calculated_unrealized_pnl_pct, leverage, pos.qty)
                    stop_mode = stop_state.stop_mode if stop_state.stop_mode != "none" else None
                    base_stop_price = round(stop_state.base_stop_price, 4) if stop_state.base_stop_price is not None and stop_state.base_stop_price > 0 else None
                    dynamic_stop_price = round(stop_state.dynamic_stop_price, 4) if stop_state.dynamic_stop_price is not None and stop_state.dynamic_stop_price > 0 else None
            except Exception as e:
                logger.debug(f"è¨ˆç®—å€‰ä½ {pos.id} çš„åœæç‹€æ…‹å¤±æ•—: {e}")
        
        # æ·»åŠ é¡å¤–å­—æ®µ
        pos_dict.update({
            "profit_threshold_value": profit_threshold_value,
            "profit_threshold_source": profit_threshold_source,
            "lock_ratio_value": lock_ratio_value,
            "lock_ratio_source": lock_ratio_source,
            "base_sl_value": base_sl_value,
            "base_sl_source": base_sl_source,
            "stop_mode": stop_mode,
            "base_stop_price": base_stop_price,
            "dynamic_stop_price": dynamic_stop_price,
        })
        
        result.append(PositionOut(**pos_dict))
    
    return result


def compute_realized_pnl(position: Position, db: Session = None) -> tuple:
    """
    è¨ˆç®—å€‰ä½çš„å·²å¯¦ç¾ç›ˆè™§å’Œç›ˆè™§ç™¾åˆ†æ¯”
    
    Args:
        position: Position æ¨¡å‹å¯¦ä¾‹
        db: å¯é¸çš„è³‡æ–™åº« Sessionï¼Œç”¨æ–¼æŸ¥è©¢ Bot çš„ leverage
    
    Returns:
        tuple: (realized_pnl, pnl_pct)
    """
    if not position.entry_price or not position.qty:
        return 0.0, 0.0
    
    # å¦‚æœ exit_price ç‚º 0 æˆ– Noneï¼Œå˜—è©¦å¾ Binance æŸ¥è©¢
    exit_price = position.exit_price
    if not exit_price or exit_price <= 0:
        if position.status == "CLOSED" and position.binance_order_id:
            try:
                client = get_client()
                order_detail = client.futures_get_order(
                    symbol=position.symbol,
                    orderId=position.binance_order_id
                )
                if order_detail.get("avgPrice"):
                    exit_price = float(order_detail["avgPrice"])
                    if exit_price > 0:
                        # æ›´æ–°è³‡æ–™åº«ä¸­çš„ exit_priceï¼ˆéœ€è¦åœ¨å¤–éƒ¨å‚³å…¥ db sessionï¼‰
                        logger.info(f"å¾ Binance æŸ¥è©¢åˆ°å€‰ä½ {position.id} çš„ exit_price: {exit_price}")
                        # æ³¨æ„ï¼šé€™è£¡ä¸ç›´æ¥æ›´æ–° DBï¼Œå› ç‚ºæ²’æœ‰ db sessionï¼Œè®“èª¿ç”¨è€…è™•ç†
                        position.exit_price = exit_price  # æš«æ™‚æ›´æ–°ç‰©ä»¶ï¼Œä½†ä¸å¯«å…¥ DB
            except Exception as e:
                logger.debug(f"æŸ¥è©¢å€‰ä½ {position.id} çš„è¨‚å–®è©³æƒ…å¤±æ•—: {e}")
    
    if not exit_price or exit_price <= 0:
        # å¦‚æœé‚„æ˜¯æ²’æœ‰ exit_priceï¼Œç„¡æ³•è¨ˆç®— PnL
        return 0.0, 0.0
    
    try:
        entry = float(position.entry_price)
        exit = float(exit_price)  # ä½¿ç”¨è™•ç†éçš„ exit_price
        qty = float(position.qty)
        
        if entry <= 0 or qty <= 0:
            return 0.0, 0.0
        
        # è¨ˆç®—å·²å¯¦ç¾ç›ˆè™§
        if position.side == "LONG":
            realized = (exit - entry) * qty
        else:  # SHORT
            realized = (entry - exit) * qty
        
        # è¨ˆç®—ç›ˆè™§ç™¾åˆ†æ¯”ï¼ˆåŸºæ–¼ä¿è­‰é‡‘ï¼Œè€Œéåç¾©åƒ¹å€¼ï¼‰
        # PnL% = (Realized PnL / ä¿è­‰é‡‘) * 100
        # ä¿è­‰é‡‘ = åç¾©åƒ¹å€¼ / æ æ¡¿ = (Entry Price * Qty) / Leverage
        entry_notional = entry * qty
        
        # å˜—è©¦å¾é—œè¯çš„ Bot å–å¾— leverageï¼Œå¦‚æœæ²’æœ‰å‰‡ä½¿ç”¨é è¨­å€¼ 20ï¼ˆå¸¸è¦‹çš„æ æ¡¿å€æ•¸ï¼‰
        leverage = 20  # é è¨­æ æ¡¿å€æ•¸
        if position.bot_id and db is not None:
            try:
                from models import BotConfig
                bot = db.query(BotConfig).filter(BotConfig.id == position.bot_id).first()
                if bot and bot.leverage:
                    leverage = bot.leverage
            except Exception as e:
                logger.debug(f"ç„¡æ³•å–å¾—å€‰ä½ {position.id} çš„ Bot leverage: {e}")
                # ä½¿ç”¨é è¨­å€¼ 20
        
        if entry_notional > 0 and leverage > 0:
            margin = entry_notional / leverage
            if margin > 0:
                pnl_pct = (realized / margin) * 100.0
            else:
                pnl_pct = 0.0
        else:
            pnl_pct = 0.0
        
        return realized, pnl_pct
    except (ValueError, TypeError) as e:
        logger.warning(f"è¨ˆç®—å€‰ä½ {position.id} çš„ PnL æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        return 0.0, 0.0


@app.get("/bot-positions/stats")
async def get_bot_positions_stats(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    æŸ¥è©¢ Bot/Manual Positions çµ±è¨ˆæ•¸æ“šï¼ˆå·²å¹³å€‰å€‰ä½ï¼‰ï¼Œæ”¯æ´æ™‚é–“å€é–“ç¯©é¸
    
    æ­¤ç«¯é»åƒ…é™å·²ç™»å…¥ä¸”é€šéç®¡ç†å“¡é©—è­‰çš„ä½¿ç”¨è€…ä½¿ç”¨ã€‚
    é è¨­æŸ¥è©¢æœ€è¿‘ 7 å¤©çš„å·²å¹³å€‰å€‰ä½ï¼ˆåŒ…å«æ‰€æœ‰ bot å‰µå»ºå’Œæ‰‹å‹•å‰µå»ºçš„å€‰ä½ï¼‰ã€‚
    
    Args:
        start_date: é–‹å§‹æ—¥æœŸï¼ˆå¯é¸ï¼Œé è¨­ç‚º 7 å¤©å‰ï¼‰
        end_date: çµæŸæ—¥æœŸï¼ˆå¯é¸ï¼Œé è¨­ç‚ºä»Šå¤©ï¼‰
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Š
        db: è³‡æ–™åº« Session
    
    Returns:
        dict: åŒ…å« stats çµ±è¨ˆè³‡è¨Š
    """
    # è¨­å®šé è¨­æ—¥æœŸç¯„åœï¼ˆæœ€è¿‘ 7 å¤©ï¼‰
    if end_date is None:
        end_date = date.today()
    
    if start_date is None:
        start_date = end_date - timedelta(days=7)
    
    # ç¢ºä¿ start_date <= end_date
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    
    # æŸ¥è©¢å·²å¹³å€‰çš„å€‰ä½
    start_datetime = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_datetime = datetime.combine(end_date + timedelta(days=1), datetime.min.time()).replace(tzinfo=timezone.utc)
    
    query = (
        db.query(Position)
        .filter(Position.status == "CLOSED")
        .filter(Position.closed_at.isnot(None))
        .filter(Position.closed_at >= start_datetime)
        .filter(Position.closed_at < end_datetime)
    )
    
    positions = query.all()
    
    # è¨ˆç®—æ¯ä¸€ç­† Position çš„ PnLï¼Œä¸¦æ›´æ–° exit_price ç‚º 0 çš„å€‰ä½
    pnl_list = []
    valid_positions = []  # æœ‰æœ‰æ•ˆ exit_price çš„å€‰ä½
    for pos in positions:
        # å¦‚æœ exit_price ç‚º 0ï¼Œå˜—è©¦å¾ Binance æŸ¥è©¢ä¸¦æ›´æ–°
        if (pos.status == "CLOSED" and 
            (not pos.exit_price or pos.exit_price <= 0) and 
            pos.binance_order_id):
            try:
                client = get_client()
                order_detail = client.futures_get_order(
                    symbol=pos.symbol,
                    orderId=pos.binance_order_id
                )
                if order_detail.get("avgPrice"):
                    exit_price = float(order_detail["avgPrice"])
                    if exit_price > 0:
                        pos.exit_price = exit_price
                        db.add(pos)
                        db.commit()
                        logger.info(f"æ›´æ–°å€‰ä½ {pos.id} çš„ exit_price: {exit_price}")
            except Exception as e:
                logger.debug(f"æŸ¥è©¢å€‰ä½ {pos.id} çš„è¨‚å–®è©³æƒ…å¤±æ•—: {e}")
        
        realized, pnl_pct = compute_realized_pnl(pos, db)
        # åªè¨ˆç®—æœ‰æœ‰æ•ˆ exit_price çš„å€‰ä½
        if pos.exit_price and pos.exit_price > 0:
            pnl_list.append(realized)
            valid_positions.append(pos)
    
    # è¨ˆç®—çµ±è¨ˆæ•¸æ“šï¼ˆåªè¨ˆç®—æœ‰æœ‰æ•ˆ exit_price çš„å€‰ä½ï¼‰
    wins = [pnl for pnl in pnl_list if pnl > 0]
    losses = [pnl for pnl in pnl_list if pnl < 0]
    
    win_count = len(wins)
    loss_count = len(losses)
    total_trades = len(valid_positions)  # ä½¿ç”¨æœ‰æ•ˆå€‰ä½æ•¸é‡
    
    win_rate = (win_count / total_trades * 100.0) if total_trades > 0 else 0.0
    
    profit_sum = sum(wins)
    loss_sum = sum(losses)  # é€™æœƒæ˜¯è² æ•¸
    
    # PnL Ratio = (å¹³å‡ç›ˆåˆ© / å¹³å‡è™§æ)ï¼Œè€Œä¸æ˜¯ (ç¸½ç›ˆåˆ© / ç¸½è™§æ)
    average_profit = profit_sum / win_count if win_count > 0 else 0.0
    average_loss = abs(loss_sum) / loss_count if loss_count > 0 else 0.0
    
    if average_loss > 0:
        pnl_ratio = average_profit / average_loss
    else:
        pnl_ratio = None
    
    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "win_count": win_count,
        "loss_count": loss_count,
        "win_rate": round(win_rate, 2),
        "profit_sum": round(profit_sum, 4),
        "loss_sum": round(loss_sum, 4),
        "pnl_ratio": round(pnl_ratio, 4) if pnl_ratio is not None else None,
        "total_trades": total_trades,
    }


@app.get("/bot-positions/export")
async def export_bot_positions_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    åŒ¯å‡º Bot/Manual Positions ç‚º Excel æª”æ¡ˆï¼ˆå·²å¹³å€‰å€‰ä½ï¼‰
    
    æ­¤ç«¯é»åƒ…é™å·²ç™»å…¥ä¸”é€šéç®¡ç†å“¡é©—è­‰çš„ä½¿ç”¨è€…ä½¿ç”¨ã€‚
    åŒ…å«æ‰€æœ‰ bot å‰µå»ºå’Œæ‰‹å‹•å‰µå»ºçš„å€‰ä½ã€‚
    
    Args:
        start_date: é–‹å§‹æ—¥æœŸï¼ˆå¯é¸ï¼Œé è¨­ç‚º 7 å¤©å‰ï¼‰
        end_date: çµæŸæ—¥æœŸï¼ˆå¯é¸ï¼Œé è¨­ç‚ºä»Šå¤©ï¼‰
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Š
        db: è³‡æ–™åº« Session
    
    Returns:
        StreamingResponse: Excel æª”æ¡ˆä¸‹è¼‰
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # è¨­å®šé è¨­æ—¥æœŸç¯„åœï¼ˆèˆ‡ /bot-positions/stats ä¸€è‡´ï¼‰
        if end_date is None:
            end_date = date.today()
        
        if start_date is None:
            start_date = end_date - timedelta(days=7)
        
        # ç¢ºä¿ start_date <= end_date
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        
        # æŸ¥è©¢å·²å¹³å€‰çš„å€‰ä½
        start_datetime = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
        end_datetime = datetime.combine(end_date + timedelta(days=1), datetime.min.time()).replace(tzinfo=timezone.utc)
        
        query = (
            db.query(Position)
            .filter(Position.status == "CLOSED")
            .filter(Position.closed_at.isnot(None))
            .filter(Position.closed_at >= start_datetime)
            .filter(Position.closed_at < end_datetime)
            .order_by(Position.closed_at.desc())
        )
        
        positions = query.all()
        
        # å»ºç«‹ Excel å·¥ä½œç°¿
        wb = Workbook()
        ws = wb.active
        ws.title = "Positions"
        
        # è¨­å®šæ¨™é¡Œæ¨£å¼
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # å¯«å…¥æ¨™é¡Œè¡Œ
        headers = [
            "ID", "Symbol", "Side", "Qty", "Entry Price", "Exit Price",
            "Realized PnL", "PnL%", "Exit Reason", "Created At", "Closed At"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # åˆå§‹åŒ– row_idxï¼ˆå¦‚æœæ²’æœ‰è³‡æ–™ï¼Œè‡³å°‘å¾æ¨™é¡Œè¡Œä¹‹å¾Œé–‹å§‹ï¼‰
        row_idx = 1
        
        # å¯«å…¥è³‡æ–™è¡Œ
        for row_idx, pos in enumerate(positions, start=2):
            realized, pnl_pct = compute_realized_pnl(pos, db)
            
            ws.cell(row=row_idx, column=1, value=pos.id)
            ws.cell(row=row_idx, column=2, value=pos.symbol)
            ws.cell(row=row_idx, column=3, value=pos.side)
            ws.cell(row=row_idx, column=4, value=pos.qty)
            ws.cell(row=row_idx, column=5, value=pos.entry_price)
            ws.cell(row=row_idx, column=6, value=pos.exit_price)
            ws.cell(row=row_idx, column=7, value=round(realized, 4))
            ws.cell(row=row_idx, column=8, value=round(pnl_pct, 2))
            ws.cell(row=row_idx, column=9, value=pos.exit_reason or "")
            ws.cell(row=row_idx, column=10, value=pos.created_at.isoformat() if pos.created_at else "")
            ws.cell(row=row_idx, column=11, value=pos.closed_at.isoformat() if pos.closed_at else "")
            
            # æ ¹æ“š PnL è¨­å®šé¡è‰²
            if realized > 0:
                ws.cell(row=row_idx, column=7).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif realized < 0:
                ws.cell(row=row_idx, column=7).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # è¨ˆç®—çµ±è¨ˆæ•¸æ“š
        wins = [p for p in positions if compute_realized_pnl(p)[0] > 0]
        losses = [p for p in positions if compute_realized_pnl(p)[0] < 0]
        
        win_count = len(wins)
        loss_count = len(losses)
        total_trades = win_count + loss_count
        win_rate = (win_count / total_trades * 100.0) if total_trades > 0 else 0.0
        profit_sum = sum(compute_realized_pnl(p)[0] for p in wins)
        loss_sum = sum(compute_realized_pnl(p)[0] for p in losses)
        # PnL Ratio = (å¹³å‡ç›ˆåˆ© / å¹³å‡è™§æ)ï¼Œè€Œä¸æ˜¯ (ç¸½ç›ˆåˆ© / ç¸½è™§æ)
        average_profit = profit_sum / win_count if win_count > 0 else 0.0
        average_loss = abs(loss_sum) / loss_count if loss_count > 0 else 0.0
        pnl_ratio = average_profit / average_loss if average_loss > 0 else None
        
        # åœ¨è³‡æ–™ä¸‹æ–¹å¯«å…¥çµ±è¨ˆè³‡è¨Š
        # å¦‚æœæ²’æœ‰è³‡æ–™ï¼Œrow_idx ä»ç„¶æ˜¯ 1ï¼ˆæ¨™é¡Œè¡Œï¼‰ï¼Œæ‰€ä»¥çµ±è¨ˆå¾ç¬¬ 3 è¡Œé–‹å§‹
        # å¦‚æœæœ‰è³‡æ–™ï¼Œrow_idx æ˜¯æœ€å¾Œä¸€ç­†è³‡æ–™çš„è¡Œè™Ÿï¼Œçµ±è¨ˆå¾ row_idx + 2 é–‹å§‹
        stats_row = row_idx + 2
        ws.cell(row=stats_row, column=1, value="Stats").font = Font(bold=True)
        stats_row += 1
        
        stats_data = [
            ("Start Date", start_date.isoformat()),
            ("End Date", end_date.isoformat()),
            ("Win Count", win_count),
            ("Loss Count", loss_count),
            ("Total Trades", total_trades),
            ("Win Rate (%)", round(win_rate, 2)),
            ("Profit Sum", round(profit_sum, 4)),
            ("Loss Sum", round(loss_sum, 4)),
            ("PnL Ratio", round(pnl_ratio, 4) if pnl_ratio is not None else "N/A"),
        ]
        
        for stat_row_idx, (label, value) in enumerate(stats_data, start=stats_row):
            ws.cell(row=stat_row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=stat_row_idx, column=2, value=value)
        
        # èª¿æ•´æ¬„ä½å¯¬åº¦
        column_widths = [8, 12, 8, 10, 12, 12, 12, 10, 15, 20, 20]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        
        # å°‡å·¥ä½œç°¿å¯«å…¥è¨˜æ†¶é«”
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # ç”¢ç”Ÿæª”æ¡ˆåç¨±
        start_str = start_date.strftime("%Y%m%d")
        end_str = end_date.strftime("%Y%m%d")
        filename = f"bot_positions_{start_str}_{end_str}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl æœªå®‰è£ï¼Œè«‹åŸ·è¡Œ: pip install openpyxl"
        )
    except Exception as e:
        logger.exception(f"åŒ¯å‡º Excel æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        raise HTTPException(status_code=500, detail=f"åŒ¯å‡ºå¤±æ•—: {str(e)}")


@app.get("/binance/open-positions")
async def get_binance_open_positions(
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    å–å¾—ç›®å‰ Binance Futures å¸³æˆ¶çš„æ‰€æœ‰æœªå¹³å€‰éƒ¨ä½ï¼ˆä¸å¯«å…¥æœ¬åœ° DBï¼‰ã€‚
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    æ³¨æ„ï¼šæ­¤ endpoint æœƒä½¿ç”¨èˆ‡ Bot ç›¸åŒçš„ Binance é€£ç·šè¨­å®šï¼ˆæ¸¬è©¦ç¶²/æ­£å¼ç¶²ï¼‰ã€‚
    å¦‚æœ USE_TESTNET=1ï¼Œå‰‡æœƒæŸ¥è©¢ Binance Futures Testnet çš„å€‰ä½ã€‚
    
    æ¯å€‹å€‰ä½æœƒåŒ…å«ï¼š
    - åŸºæœ¬è³‡è¨Šï¼ˆsymbol, position_amt, entry_price, mark_price, etc.ï¼‰
    - unrealized_pnl_pct: æœªå¯¦ç¾ç›ˆè™§ç™¾åˆ†æ¯”
    - stop_mode: åœææ¨¡å¼ï¼ˆ"dynamic", "base", "none"ï¼‰
    - base_stop_price: åŸºç¤åœæåƒ¹æ ¼ï¼ˆå¦‚æœ‰ï¼‰
    - dynamic_stop_price: å‹•æ…‹åœæåƒ¹æ ¼ï¼ˆå¦‚æœ‰ï¼‰
    
    Args:
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
        db: è³‡æ–™åº« Sessionï¼ˆç”¨æ–¼æŸ¥æ‰¾å°æ‡‰çš„æœ¬åœ° Positionï¼‰
    
    Returns:
        List[dict]: Binance Futures æœªå¹³å€‰éƒ¨ä½åˆ—è¡¨ï¼ˆåŒ…å« PnL% å’Œåœæè³‡è¨Šï¼‰
    
    Raises:
        HTTPException: ç•¶ Binance API å‘¼å«å¤±æ•—æ™‚
            - 400: API Key/Secret æœªè¨­å®š
            - 500: å…¶ä»– Binance API éŒ¯èª¤
    """
    try:
        # å˜—è©¦å–å¾— Binance client
        client = get_client()
        
        # ä½¿ç”¨ USDT-M Futures position info
        raw_positions = client.futures_position_information()
        
        positions = []
        for item in raw_positions:
            # åªä¿ç•™æŒå€‰ä¸ç‚º 0 çš„éƒ¨ä½
            try:
                position_amt = float(item.get("positionAmt", "0") or 0)
            except (ValueError, TypeError):
                position_amt = 0.0
            
            if position_amt == 0:
                # ç•¶ position é—œé–‰æ™‚ï¼ˆposition_amt == 0ï¼‰ï¼Œæ¸…ç†è¿½è¹¤è¨˜éŒ„
                symbol_for_cleanup = item.get("symbol", "")
                if symbol_for_cleanup:
                    # æ¸…ç†å°æ‡‰çš„è¿½è¹¤è¨˜éŒ„
                    for side_cleanup in ["LONG", "SHORT"]:
                        tracking_key_cleanup = f"{symbol_for_cleanup}|{side_cleanup}"
                        if tracking_key_cleanup in _non_bot_position_tracking:
                            del _non_bot_position_tracking[tracking_key_cleanup]
                            logger.debug(f"æ¸…ç†é bot å€‰ä½è¿½è¹¤è¨˜éŒ„: {tracking_key_cleanup}")
                continue
            
            # è§£æå…¶ä»–æ¬„ä½ï¼ˆå¿…é ˆåœ¨ position_amt != 0 ä¹‹å¾Œï¼‰
            
            # è§£æå…¶ä»–æ¬„ä½
            try:
                entry_price = float(item.get("entryPrice", "0") or 0)
                mark_price = float(item.get("markPrice", "0") or 0)
                unrealized_pnl = float(item.get("unRealizedProfit", "0") or 0)
                leverage = int(float(item.get("leverage", "0") or 0))
                isolated_wallet = float(item.get("isolatedWallet", "0") or 0)
                update_time = int(item.get("updateTime", 0) or 0)
            except (ValueError, TypeError) as e:
                logger.warning(f"è§£æ Binance position æ¬„ä½å¤±æ•—: {item.get('symbol', 'unknown')}, éŒ¯èª¤: {e}")
                continue
            
            margin_type = item.get("marginType", "")
            symbol = item.get("symbol", "")
            
            # è¨ˆç®— unrealized PnL ç™¾åˆ†æ¯”ï¼ˆåŸºæ–¼ä¿è­‰é‡‘ï¼Œè€Œéåç¾©åƒ¹å€¼ï¼‰
            # PnL% = (Unrealized PnL / ä¿è­‰é‡‘) * 100
            # ä¿è­‰é‡‘ = åç¾©åƒ¹å€¼ / æ æ¡¿ = (Position Size * Entry Price) / Leverage
            unrealized_pnl_pct = None
            if entry_price > 0 and abs(position_amt) > 0 and leverage > 0:
                notional = abs(position_amt) * entry_price
                if notional > 0:
                    # è¨ˆç®—ä¿è­‰é‡‘
                    margin = notional / leverage
                    if margin > 0:
                        # PnL% = (Unrealized PnL / ä¿è­‰é‡‘) * 100
                        unrealized_pnl_pct = (unrealized_pnl / margin) * 100.0
            # å¦‚æœç„¡æ³•è¨ˆç®—ï¼ˆä¾‹å¦‚ leverage = 0ï¼‰ï¼Œunrealized_pnl_pct ä¿æŒç‚º None
            
            # è¨ˆç®—åœæè³‡è¨Šï¼šæŸ¥æ‰¾å°æ‡‰çš„æœ¬åœ° Position
            stop_mode = "none"
            base_stop_price = None
            dynamic_stop_price = None
            
            # æ±ºå®šæœ¬åœ° Position çš„ side
            side_local = "LONG" if position_amt > 0 else "SHORT"
            
            # æŸ¥æ‰¾åŒ¹é…çš„æœ¬åœ° Positionï¼ˆæœ€æ–°çš„ OPEN å€‰ä½ï¼‰
            # æ³¨æ„ï¼šé€™å€‹æŸ¥è©¢åœ¨ try å¡Šå¤–åŸ·è¡Œï¼Œç¢ºä¿ local_pos åœ¨å¾ŒçºŒä»£ç¢¼ä¸­å¯ç”¨
            # ä½¿ç”¨å¤§å°å¯«ä¸æ•æ„Ÿçš„åŒ¹é…ï¼ˆsymbol æ‡‰è©²éƒ½æ˜¯å¤§å¯«ï¼Œä½†ç‚ºäº†å®‰å…¨èµ·è¦‹ï¼‰
            local_pos = (
                db.query(Position)
                .filter(
                    Position.symbol == symbol.upper(),
                    Position.side == side_local,
                    Position.status == "OPEN",
                )
                .order_by(Position.id.desc())
                .first()
            )
            
            # èª¿è©¦æ—¥èªŒï¼šè¨˜éŒ„åŒ¹é…çµæœ
            if local_pos:
                logger.debug(
                    f"Binance position {symbol} ({side_local}) åŒ¹é…åˆ°æœ¬åœ° Position ID={local_pos.id}, "
                    f"entry_price={local_pos.entry_price}, highest_price={local_pos.highest_price}"
                )
            else:
                logger.debug(
                    f"Binance position {symbol} ({side_local}) æ²’æœ‰åŒ¹é…çš„æœ¬åœ° Positionï¼ˆå¯èƒ½æ˜¯æ‰‹å‹•é–‹å€‰ï¼‰"
                )
            
            try:
                
                # æ¸…ç†å·²é—œé–‰çš„ position çš„è¿½è¹¤è¨˜éŒ„
                # å¦‚æœæ‰¾åˆ°äº†æœ¬åœ° Positionï¼Œèªªæ˜é€™æ˜¯ç”± bot å‰µå»ºçš„ï¼Œä¸éœ€è¦è¿½è¹¤
                # å¦‚æœæ‰¾ä¸åˆ°æœ¬åœ° Positionï¼Œæˆ‘å€‘æœƒä½¿ç”¨è¨˜æ†¶é«”è¿½è¹¤ä¾†ç¶­æŒ highest_price
                
                if local_pos:
                    # æ‰¾åˆ°åŒ¹é…çš„æœ¬åœ° Positionï¼ˆbot å‰µå»ºçš„å€‰ä½ï¼‰
                    logger.debug(
                        f"Binance Live Position {symbol} ({side_local}) åŒ¹é…åˆ°æœ¬åœ° Position ID={local_pos.id}"
                    )
                    # æª¢æŸ¥æ˜¯å¦æœ‰åœæé…ç½®è¦†å¯«å€¼ï¼ˆBinance Live Positions çš„è¦†å¯«å„ªå…ˆæ–¼æœ¬åœ° Positionï¼‰
                    override_key = f"{symbol}|{side_local}"
                    overrides = _binance_position_stop_overrides.get(override_key, {})
                    
                    # å¦‚æœæœ¬åœ° Position çš„ entry_price ç„¡æ•ˆï¼Œä½¿ç”¨ Binance çš„ entryPrice ä½œç‚º fallback
                    if local_pos.entry_price <= 0 and entry_price > 0:
                        logger.warning(
                            f"æœ¬åœ°å€‰ä½ {local_pos.id} ({symbol}) entry_price={local_pos.entry_price} ç„¡æ•ˆï¼Œ"
                            f"ä½¿ç”¨ Binance entryPrice={entry_price} ä½œç‚º fallback è¨ˆç®—åœæ"
                        )
                        # å»ºç«‹ä¸€å€‹è‡¨æ™‚çš„ Position ç‰©ä»¶ç”¨æ–¼è¨ˆç®—
                        from copy import copy
                        temp_pos = copy(local_pos)
                        temp_pos.entry_price = entry_price
                        # å¦‚æœ highest_price ä¹Ÿç„¡æ•ˆï¼Œä½¿ç”¨ç•¶å‰ mark_price
                        if temp_pos.highest_price is None or temp_pos.highest_price <= 0:
                            temp_pos.highest_price = mark_price
                        # æ‡‰ç”¨è¦†å¯«å€¼ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
                        if overrides.get("dyn_profit_threshold_pct") is not None:
                            temp_pos.dyn_profit_threshold_pct = overrides["dyn_profit_threshold_pct"]
                        if overrides.get("base_stop_loss_pct") is not None:
                            temp_pos.base_stop_loss_pct = overrides["base_stop_loss_pct"]
                        if overrides.get("trail_callback") is not None:
                            temp_pos.trail_callback = overrides["trail_callback"]
                        # è¨ˆç®— unrealized_pnl_pctï¼ˆPnL%ï¼‰ç”¨æ–¼åˆ¤æ–·æ˜¯å¦é€²å…¥ dynamic mode
                        # å°æ–¼ bot å‰µå»ºçš„ positionï¼Œéœ€è¦è¨ˆç®— PnL%
                        calculated_unrealized_pnl_pct = None
                        if entry_price > 0 and abs(position_amt) > 0 and leverage > 0:
                            notional = abs(position_amt) * entry_price
                            if notional > 0:
                                margin = notional / leverage
                                if margin > 0:
                                    calculated_unrealized_pnl_pct = (unrealized_pnl / margin) * 100.0
                        stop_state = compute_stop_state(temp_pos, mark_price, calculated_unrealized_pnl_pct, leverage, abs(position_amt))
                    else:
                        # æ‡‰ç”¨è¦†å¯«å€¼ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
                        if overrides.get("dyn_profit_threshold_pct") is not None:
                            local_pos.dyn_profit_threshold_pct = overrides["dyn_profit_threshold_pct"]
                        if overrides.get("base_stop_loss_pct") is not None:
                            local_pos.base_stop_loss_pct = overrides["base_stop_loss_pct"]
                        if overrides.get("trail_callback") is not None:
                            local_pos.trail_callback = overrides["trail_callback"]
                        # è¨ˆç®— unrealized_pnl_pctï¼ˆPnL%ï¼‰ç”¨æ–¼åˆ¤æ–·æ˜¯å¦é€²å…¥ dynamic mode
                        # å°æ–¼ bot å‰µå»ºçš„ positionï¼Œéœ€è¦è¨ˆç®— PnL%
                        calculated_unrealized_pnl_pct = None
                        if entry_price > 0 and abs(position_amt) > 0 and leverage > 0:
                            notional = abs(position_amt) * entry_price
                            if notional > 0:
                                margin = notional / leverage
                                if margin > 0:
                                    calculated_unrealized_pnl_pct = (unrealized_pnl / margin) * 100.0
                        # ä½¿ç”¨ compute_stop_state è¨ˆç®—åœæç‹€æ…‹ï¼ˆå‚³å…¥ leverage å’Œ qtyï¼‰
                        stop_state = compute_stop_state(local_pos, mark_price, calculated_unrealized_pnl_pct, leverage, abs(position_amt))
                    stop_mode = stop_state.stop_mode
                    base_stop_price = stop_state.base_stop_price
                    dynamic_stop_price = stop_state.dynamic_stop_price
                else:
                    # é bot å‰µå»ºçš„ positionï¼šä½¿ç”¨ Binance è³‡æ–™å»ºç«‹è‡¨æ™‚ Position ç‰©ä»¶ä¾†è¨ˆç®—åœæ
                    # é—œéµï¼šä½¿ç”¨è¨˜æ†¶é«”ä¸­çš„è¿½è¹¤è¨˜éŒ„ä¾†ç¶­æŒæ­·å²æœ€é«˜/æœ€ä½åƒ¹æ ¼
                    # é€™æ¨£å³ä½¿ç•¶å‰åƒ¹æ ¼ä¸‹è·Œï¼Œdynamic stop ä¹Ÿèƒ½ä¿æŒç©©å®š
                    logger.debug(
                        f"Binance Live Position {symbol} ({side_local}) æ²’æœ‰åŒ¹é…çš„æœ¬åœ° Positionï¼ˆé bot å‰µå»ºï¼‰"
                    )
                    
                    # å»ºç«‹è¿½è¹¤ keyï¼šä½¿ç”¨ symbol å’Œ side ä¾†å”¯ä¸€æ¨™è­˜ä¸€å€‹ position
                    tracking_key = f"{symbol}|{side_local}"
                    
                    # æª¢æŸ¥æ˜¯å¦æœ‰åœæé…ç½®è¦†å¯«å€¼ï¼ˆåœ¨æª¢æŸ¥è¿½è¹¤è¨˜éŒ„ä¹‹å‰ï¼‰
                    override_key = f"{symbol}|{side_local}"
                    overrides = _binance_position_stop_overrides.get(override_key, {})
                    
                    # æª¢æŸ¥æ˜¯å¦å·²æœ‰è¿½è¹¤è¨˜éŒ„
                    if tracking_key in _non_bot_position_tracking:
                        tracked = _non_bot_position_tracking[tracking_key]
                        tracked_entry = tracked.get("entry_price")
                        tracked_highest = tracked.get("highest_price")
                        tracked_side = tracked.get("side")
                        
                        # å¦‚æœ entry_price æ”¹è®Šï¼ˆå¯èƒ½æ˜¯åŒä¸€å€‹ symbol ä½†ä¸åŒçš„ positionï¼‰ï¼Œé‡ç½®è¿½è¹¤
                        # ä½¿ç”¨ç›¸å°èª¤å·®è€Œä¸æ˜¯çµ•å°èª¤å·®ï¼Œé¿å…å°æ•¸é»ç²¾åº¦å•é¡Œ
                        if tracked_entry is None or (abs(tracked_entry - entry_price) / max(abs(tracked_entry), abs(entry_price), 1.0)) > 0.001:
                            logger.debug(
                                f"é bot å€‰ä½ {symbol} ({side_local}) entry_price æ”¹è®Šï¼š"
                                f"èˆŠ={tracked_entry}, æ–°={entry_price}ï¼Œé‡ç½®è¿½è¹¤è¨˜éŒ„"
                            )
                            tracked_entry = entry_price
                            tracked_highest = None
                            tracked_side = side_local
                    else:
                        # é¦–æ¬¡çœ‹åˆ°é€™å€‹ positionï¼Œåˆå§‹åŒ–è¿½è¹¤
                        tracked_entry = entry_price
                        tracked_highest = None
                        tracked_side = side_local
                    
                    # æ›´æ–°æ­·å²æœ€é«˜/æœ€ä½åƒ¹æ ¼ï¼ˆåªèƒ½ä¸Šå‡/ä¸‹é™ï¼Œä¸èƒ½å›é€€ï¼‰
                    if side_local == "LONG":
                        # LONGï¼šhighest_price åªèƒ½ä¸Šå‡ï¼Œä¸èƒ½ä¸‹é™
                        if tracked_highest is None:
                            tracked_highest = max(mark_price, entry_price) if entry_price > 0 else mark_price
                        else:
                            # åªèƒ½æ›´æ–°ç‚ºæ›´é«˜çš„åƒ¹æ ¼
                            tracked_highest = max(tracked_highest, mark_price)
                    else:
                        # SHORTï¼šhighest_price æ¬„ä½å¯¦éš›å­˜å„²çš„æ˜¯æœ€ä½åƒ¹æ ¼ï¼Œåªèƒ½ä¸‹é™ï¼Œä¸èƒ½ä¸Šå‡
                        if tracked_highest is None:
                            tracked_highest = min(mark_price, entry_price) if entry_price > 0 else mark_price
                        else:
                            # åªèƒ½æ›´æ–°ç‚ºæ›´ä½çš„åƒ¹æ ¼
                            tracked_highest = min(tracked_highest, mark_price)
                    
                    # ç¢ºä¿ tracked_highest ä¸ç‚º Noneï¼ˆä½¿ç”¨ç•¶å‰ mark_price ä½œç‚º fallbackï¼‰
                    if tracked_highest is None:
                        tracked_highest = mark_price
                    
                    # æ›´æ–°è¿½è¹¤è¨˜éŒ„
                    _non_bot_position_tracking[tracking_key] = {
                        "entry_price": tracked_entry,
                        "highest_price": tracked_highest,
                        "side": tracked_side
                    }
                    
                    # å»ºç«‹è‡¨æ™‚ Position ç‰©ä»¶
                    class TempPosition:
                        def __init__(self, entry_price, side, highest_price=None):
                            self.entry_price = entry_price
                            self.side = side
                            self.highest_price = highest_price  # LONG: æœ€é«˜åƒ¹, SHORT: æœ€ä½åƒ¹
                            # ä½¿ç”¨è¦†å¯«å€¼ï¼ˆå¦‚æœå­˜åœ¨ï¼‰ï¼Œå¦å‰‡ä½¿ç”¨ Noneï¼ˆæœƒä½¿ç”¨å…¨å±€é…ç½®ï¼‰
                            self.trail_callback = overrides.get("trail_callback")
                            self.dyn_profit_threshold_pct = overrides.get("dyn_profit_threshold_pct")
                            self.base_stop_loss_pct = overrides.get("base_stop_loss_pct")
                    
                    temp_pos = TempPosition(
                        entry_price=tracked_entry if tracked_entry else entry_price,  # ä½¿ç”¨è¿½è¹¤çš„ entry_priceï¼ˆæ›´æº–ç¢ºï¼‰
                        side=side_local,
                        highest_price=tracked_highest  # ä½¿ç”¨è¿½è¹¤çš„æ­·å²æœ€é«˜/æœ€ä½åƒ¹æ ¼
                    )
                    # ä½¿ç”¨å·²è¨ˆç®—çš„ unrealized_pnl_pctï¼ˆPnL%ï¼‰ä¾†åˆ¤æ–·æ˜¯å¦é€²å…¥ dynamic modeï¼ˆå‚³å…¥ leverage å’Œ qtyï¼‰
                    stop_state = compute_stop_state(temp_pos, mark_price, unrealized_pnl_pct, leverage, abs(position_amt))
                    stop_mode = stop_state.stop_mode
                    base_stop_price = stop_state.base_stop_price
                    dynamic_stop_price = stop_state.dynamic_stop_price
                    
                    # èª¿è©¦æ—¥èªŒï¼šè¨˜éŒ„é bot å‰µå»ºçš„ position çš„åœæè¨ˆç®—çµæœ
                    # ä½¿ç”¨ info ç´šåˆ¥ä»¥ä¾¿è¿½è¹¤å•é¡Œ
                    # è¨ˆç®— profit_pct ç”¨æ–¼èª¿è©¦
                    profit_pct_debug = ((tracked_highest - tracked_entry) / tracked_entry * 100.0) if (tracked_entry and tracked_entry > 0 and tracked_highest) else 0.0
                    # å®‰å…¨åœ°æ ¼å¼åŒ–å¯èƒ½ç‚º None çš„å€¼
                    tracked_highest_str = f"{tracked_highest:.4f}" if tracked_highest else "None"
                    base_stop_str = f"{base_stop_price:.4f}" if base_stop_price else "None"
                    dynamic_stop_str = f"{dynamic_stop_price:.4f}" if dynamic_stop_price else "None"
                    logger.info(
                        f"é bot å‰µå»ºçš„å€‰ä½ {symbol} ({side_local}) åœæè¨ˆç®—ï¼š"
                        f"entry={entry_price:.4f}, mark={mark_price:.4f}, "
                        f"tracked_highest={tracked_highest_str}, "
                        f"profit_pct={profit_pct_debug:.2f}%, "
                        f"trail_callback={temp_pos.trail_callback}, "
                        f"dyn_profit_threshold_pct={temp_pos.dyn_profit_threshold_pct}, "
                        f"overrides={overrides}, "
                        f"stop_mode={stop_mode}, "
                        f"base_stop={base_stop_str}, "
                        f"dynamic_stop={dynamic_stop_str}"
                    )
                
            except Exception as e:
                # å¦‚æœè¨ˆç®—åœæç‹€æ…‹å¤±æ•—ï¼Œè¨˜éŒ„è­¦å‘Šå’Œå®Œæ•´çš„å †ç–Šè¿½è¹¤
                logger.warning(f"è¨ˆç®—å€‰ä½ {symbol} ({side_local}) åœæç‹€æ…‹æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
                # å³ä½¿è¨ˆç®—å¤±æ•—ï¼Œä¹Ÿå˜—è©¦ä½¿ç”¨åŸºæœ¬é…ç½®è¨ˆç®— base stopï¼ˆå¦‚æœæœ‰é…ç½®ï¼‰
                try:
                    # ä½¿ç”¨å°æ‡‰æ–¹å‘çš„å…¨å±€é…ç½®å˜—è©¦è¨ˆç®— base stop
                    side_config = TRAILING_CONFIG.get_config_for_side(side_local)
                    base_sl_pct = side_config.base_sl_pct if side_config.base_sl_pct is not None else DYN_BASE_SL_PCT
                    if base_sl_pct > 0 and entry_price > 0:
                        if side_local == "LONG":
                            base_stop_price = entry_price * (1 - base_sl_pct / 100.0)
                            stop_mode = "base"
                        else:  # SHORT
                            base_stop_price = entry_price * (1 + base_sl_pct / 100.0)
                            stop_mode = "base"
                        dynamic_stop_price = None
                    else:
                        stop_mode = "none"
                        base_stop_price = None
                        dynamic_stop_price = None
                except Exception as e2:
                    logger.warning(f"è¨ˆç®— base stop ä¹Ÿå¤±æ•—: {e2}")
                    stop_mode = "none"
                    base_stop_price = None
                    dynamic_stop_price = None
            
            # æª¢æŸ¥æ˜¯å¦æœ‰åœæé…ç½®è¦†å¯«å€¼ï¼ˆç”¨æ–¼å‰ç«¯é¡¯ç¤ºï¼‰
            # override_key æ‡‰è©²å·²ç¶“åœ¨ try å¡Šä¸­å®šç¾©ï¼ˆå…©å€‹åˆ†æ”¯éƒ½æœƒå®šç¾©ï¼‰
            if 'override_key' not in locals():
                override_key = f"{symbol}|{side_local}"
            overrides = _binance_position_stop_overrides.get(override_key, {})
            
            # ç¢ºä¿ local_pos å·²å®šç¾©ï¼ˆåœ¨ try å¡Šå¤–å·²å®šç¾©ï¼Œä½†å¦‚æœ try å¡Šå¤±æ•—ï¼Œéœ€è¦é‡æ–°æŸ¥æ‰¾ï¼‰
            if 'local_pos' not in locals():
                # å¦‚æœæ²’æœ‰åœ¨ try å¡Šä¸­å®šç¾©ï¼Œé‡æ–°æŸ¥æ‰¾
                local_pos = (
                    db.query(Position)
                    .filter(
                        Position.symbol == symbol,
                        Position.side == side_local,
                        Position.status == "OPEN",
                    )
                    .order_by(Position.id.desc())
                    .first()
                )
            
            # è¨ˆç®—å¯¦éš›ä½¿ç”¨çš„å€¼å’Œä¾†æºæ¨™è¨˜
            # å„ªå…ˆé †åºï¼šoverride (æ‰‹å‹•è¨­å®š) > local_pos (bot position è¨­å®šï¼Œä¹Ÿè¦–ç‚º override) > global (å…¨å±€è¨­å®š) > default (é»˜èªå€¼)
            # ä¾†æºæ¨™è¨˜ï¼š
            #   "override" - æ‰‹å‹•å¡«çš„åœæè¨­å®šï¼ˆé»ƒè‰²ï¼‰
            #   "global" - å…¨å±€è¨­å®šï¼ˆè—è‰²ï¼‰
            #   "default" - é»˜èªå€¼ï¼ˆç°è‰²ï¼‰
            
            # å–å¾—å°æ‡‰æ–¹å‘çš„å…¨å±€è¨­å®š
            side_config = TRAILING_CONFIG.get_config_for_side(side_local)
            
            # Profit Threshold (dyn_profit_threshold_pct)
            profit_threshold_value = None
            profit_threshold_source = None
            if overrides.get("dyn_profit_threshold_pct") is not None:
                profit_threshold_value = overrides["dyn_profit_threshold_pct"]
                profit_threshold_source = "override"
            elif local_pos and local_pos.dyn_profit_threshold_pct is not None:
                profit_threshold_value = local_pos.dyn_profit_threshold_pct
                profit_threshold_source = "override"  # bot position çš„è¨­å®šä¹Ÿè¦–ç‚º override
            elif side_config.profit_threshold_pct is not None:
                profit_threshold_value = side_config.profit_threshold_pct
                profit_threshold_source = "global"
            else:
                profit_threshold_value = DYN_PROFIT_THRESHOLD_PCT
                profit_threshold_source = "default"
            
            # Lock Ratio (trail_callback)
            lock_ratio_value = None
            lock_ratio_source = None
            if overrides.get("trail_callback") is not None:
                lock_ratio_value = overrides["trail_callback"]
                lock_ratio_source = "override"
            elif local_pos and local_pos.trail_callback is not None:
                lock_ratio_value = local_pos.trail_callback
                lock_ratio_source = "override"  # bot position çš„è¨­å®šä¹Ÿè¦–ç‚º override
            elif side_config.lock_ratio is not None:
                lock_ratio_value = side_config.lock_ratio
                lock_ratio_source = "global"
            else:
                lock_ratio_value = DYN_LOCK_RATIO_DEFAULT
                lock_ratio_source = "default"
            
            # Base SL% (base_stop_loss_pct)
            base_sl_value = None
            base_sl_source = None
            if overrides.get("base_stop_loss_pct") is not None:
                base_sl_value = overrides["base_stop_loss_pct"]
                base_sl_source = "override"
            elif local_pos and local_pos.base_stop_loss_pct is not None:
                base_sl_value = local_pos.base_stop_loss_pct
                base_sl_source = "override"  # bot position çš„è¨­å®šä¹Ÿè¦–ç‚º override
            elif side_config.base_sl_pct is not None:
                base_sl_value = side_config.base_sl_pct
                base_sl_source = "global"
            else:
                base_sl_value = DYN_BASE_SL_PCT
                base_sl_source = "default"
            
            # ç¢ºä¿æ‰€æœ‰æ¬„ä½éƒ½åŒ…å«åœ¨å›æ‡‰ä¸­ï¼Œç„¡è³‡æ–™æ™‚ä½¿ç”¨ null
            positions.append({
                "symbol": symbol,
                "position_amt": position_amt,
                "entry_price": entry_price,
                "mark_price": mark_price,
                "unrealized_pnl": unrealized_pnl,
                "unrealized_pnl_pct": round(unrealized_pnl_pct, 4) if unrealized_pnl_pct is not None else None,  # ä¿ç•™ 4 ä½å°æ•¸ï¼Œç„¡è³‡æ–™æ™‚ç‚º null
                "leverage": leverage,
                "margin_type": margin_type,
                "isolated_wallet": isolated_wallet,
                "update_time": update_time,
                "stop_mode": stop_mode if stop_mode != "none" else None,  # "dynamic", "base", æˆ– Noneï¼ˆå‰ç«¯æœƒé¡¯ç¤ºç‚º â€”ï¼‰
                "base_stop_price": round(base_stop_price, 4) if base_stop_price is not None and base_stop_price > 0 else None,  # ç„¡è³‡æ–™æ™‚ç‚º nullï¼Œä¿ç•™ 4 ä½å°æ•¸
                "dynamic_stop_price": round(dynamic_stop_price, 4) if dynamic_stop_price is not None and dynamic_stop_price > 0 else None,  # ç„¡è³‡æ–™æ™‚ç‚º nullï¼Œä¿ç•™ 4 ä½å°æ•¸
                # æ·»åŠ åœæé…ç½®è¦†å¯«å€¼ï¼ˆç”¨æ–¼å‰ç«¯é¡¯ç¤ºï¼Œä¿ç•™åŸå§‹è¦†å¯«å€¼ï¼‰
                "dyn_profit_threshold_pct": overrides.get("dyn_profit_threshold_pct"),
                "base_stop_loss_pct": overrides.get("base_stop_loss_pct"),
                "trail_callback": overrides.get("trail_callback"),
                # æ·»åŠ å¯¦éš›ä½¿ç”¨çš„å€¼å’Œä¾†æºæ¨™è¨˜ï¼ˆç”¨æ–¼å‰ç«¯é¡¯ç¤ºå’Œé¡è‰²æ¨™è¨˜ï¼‰
                "profit_threshold_value": profit_threshold_value,
                "profit_threshold_source": profit_threshold_source,  # "override", "local", "global", "default"
                "lock_ratio_value": lock_ratio_value,
                "lock_ratio_source": lock_ratio_source,  # "override", "local", "global", "default"
                "base_sl_value": base_sl_value,
                "base_sl_source": base_sl_source,  # "override", "local", "global", "default"
                # æ·»åŠ æ¨™è­˜ï¼šæ˜¯å¦ç‚º bot å‰µå»ºçš„å€‰ä½ï¼ˆç”¨æ–¼å‰ç«¯é¡¯ç¤ºå’Œå€åˆ†ï¼‰
                "is_bot_position": local_pos is not None,
                "bot_position_id": local_pos.id if local_pos else None,
            })
        
        logger.info(
            f"å–å¾— Binance open positions: {len(positions)} ç­† "
            f"(åŒ…å« bot å‰µå»ºçš„å€‰ä½å’Œæ‰‹å‹•å‰µå»ºçš„å€‰ä½)"
        )
        return positions
        
    except ValueError as e:
        # API Key/Secret æœªè¨­å®š
        error_msg = str(e)
        logger.error(f"Binance API è¨­å®šéŒ¯èª¤: {error_msg}")
        raise HTTPException(
            status_code=400,
            detail=f"Binance API æœªè¨­å®š: {error_msg}ã€‚è«‹åœ¨ç’°å¢ƒè®Šæ•¸ä¸­è¨­å®š BINANCE_API_KEY å’Œ BINANCE_API_SECRETã€‚"
        )
    except Exception as e:
        # å…¶ä»– Binance API éŒ¯èª¤
        logger.exception("å–å¾— Binance open positions å¤±æ•—")
        raise HTTPException(
            status_code=500,
            detail=f"Binance API éŒ¯èª¤: {str(e)}"
        )


@app.post("/binance/positions/close")
async def close_binance_live_position(
    payload: BinanceCloseRequest,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    é—œé–‰ Binance Live Positionï¼ˆä¸å±¬æ–¼ bot ç®¡ç†çš„å€‰ä½ï¼‰ã€‚
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        payload: é—œå€‰è«‹æ±‚ï¼ˆåŒ…å« symbol å’Œ position_sideï¼‰
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: é—œå€‰è¨‚å–®è³‡è¨Š
    
    Raises:
        HTTPException: ç•¶é—œå€‰å¤±æ•—æ™‚
    """
    try:
        client = get_client()
        symbol = payload.symbol.upper()
        position_side = payload.position_side.upper()
        
        # å–å¾—æœ€æ–°çš„å€‰ä½è³‡è¨Š
        positions = client.futures_position_information(symbol=symbol)
        
        if not positions:
            raise HTTPException(
                status_code=404,
                detail=f"æ‰¾ä¸åˆ°äº¤æ˜“å° {symbol} çš„å€‰ä½è³‡è¨Š"
            )
        
        # æ‰¾åˆ°è©²äº¤æ˜“å°çš„å€‰ä½ï¼ˆé€šå¸¸åªæœ‰ä¸€å€‹ï¼‰
        position_info = positions[0]
        
        try:
            position_amt = float(position_info.get("positionAmt", "0") or 0)
        except (ValueError, TypeError):
            position_amt = 0.0
        
        # æª¢æŸ¥å€‰ä½æ˜¯å¦ç‚º 0
        if position_amt == 0:
            raise HTTPException(
                status_code=400,
                detail=f"{symbol} ç›®å‰æ²’æœ‰æœªå¹³å€‰éƒ¨ä½"
            )
        
        # é©—è­‰å€‰ä½æ–¹å‘
        actual_side = "LONG" if position_amt > 0 else "SHORT"
        if actual_side != position_side:
            raise HTTPException(
                status_code=400,
                detail=f"å€‰ä½æ–¹å‘ä¸åŒ¹é…ï¼šå¯¦éš›ç‚º {actual_side}ï¼Œè«‹æ±‚ç‚º {position_side}"
            )
        
        # ä½¿ç”¨ close_futures_position çš„é‚è¼¯ï¼Œä½†ä¸éœ€è¦ position_id
        # ç”±æ–¼ close_futures_position éœ€è¦ position_idï¼Œæˆ‘å€‘ç›´æ¥åœ¨é€™è£¡å¯¦ä½œ
        if position_side == "LONG":
            side = "SELL"  # å¹³å¤šï¼šè³£å‡º
        elif position_side == "SHORT":
            side = "BUY"   # å¹³ç©ºï¼šè²·å…¥
        else:
            raise HTTPException(
                status_code=400,
                detail=f"ä¸æ”¯æ´çš„ position_side: {position_side}ï¼Œå¿…é ˆæ˜¯ LONG æˆ– SHORT"
            )
        
        qty = abs(position_amt)
        
        # ç”¢ç”Ÿè‡ªè¨‚çš„ client order ID
        timestamp = int(time.time() * 1000)
        client_order_id = f"TVBOT_CLOSE_LIVE_{timestamp}"
        
        logger.info(f"é—œé–‰ Binance Live Position: {symbol} {position_side}ï¼Œæ•¸é‡: {qty}, ä¸‹å–®æ–¹å‘: {side}")
        
        # å»ºç«‹å¸‚åƒ¹å–®é—œå€‰
        order = client.futures_create_order(
            symbol=symbol,
            side=side,
            type="MARKET",
            quantity=qty,
            reduceOnly=True,  # è¨­å®šç‚ºåªæ¸›å€‰ï¼Œç¢ºä¿æ˜¯å¹³å€‰å–®
            newClientOrderId=client_order_id
        )
        
        logger.info(f"æˆåŠŸé—œé–‰ Binance Live Position: {symbol}ï¼Œè¨‚å–®ID: {order.get('orderId')}")
        
        # å–å¾—å¹³å€‰åƒ¹æ ¼
        exit_price = get_exit_price_from_order(order, symbol)
        
        # å–å¾— entry_priceï¼ˆå¾ Binance position infoï¼‰
        entry_price = float(position_info.get("entryPrice", "0") or 0)
        if entry_price <= 0:
            # å¦‚æœ entry_price ç„¡æ•ˆï¼Œå˜—è©¦å¾è¿½è¹¤è¨˜éŒ„å–å¾—
            tracking_key = f"{symbol}|{position_side}"
            if tracking_key in _non_bot_position_tracking:
                tracked_entry = _non_bot_position_tracking[tracking_key].get("entry_price")
                if tracked_entry and tracked_entry > 0:
                    entry_price = tracked_entry
        
        # æŸ¥æ‰¾ç¾æœ‰çš„ OPEN å€‰ä½ï¼ˆå„ªå…ˆæ›´æ–°ç¾æœ‰å€‰ä½ï¼Œè€Œä¸æ˜¯å»ºç«‹æ–°å€‰ä½ï¼‰
        existing_position = (
            db.query(Position)
            .filter(
                Position.symbol == symbol.upper(),
                Position.side == position_side,
                Position.status == "OPEN",
            )
            .order_by(Position.id.desc())
            .first()
        )
        
        try:
            if existing_position:
                # æ›´æ–°ç¾æœ‰å€‰ä½
                existing_position.status = "CLOSED"
                existing_position.closed_at = datetime.now(timezone.utc)
                existing_position.exit_price = exit_price
                existing_position.exit_reason = "manual_close"
                # æ›´æ–°è¨‚å–®è³‡è¨Šï¼ˆå¦‚æœæœ‰çš„è©±ï¼‰
                if order.get("orderId"):
                    existing_position.binance_order_id = int(order.get("orderId"))
                if order.get("clientOrderId"):
                    existing_position.client_order_id = order.get("clientOrderId")
                # å¦‚æœ entry_price æœ‰æ•ˆä¸”ç¾æœ‰å€‰ä½çš„ entry_price ç„¡æ•ˆï¼Œæ›´æ–°å®ƒ
                if entry_price > 0 and (existing_position.entry_price <= 0 or existing_position.entry_price is None):
                    existing_position.entry_price = entry_price
                
                db.commit()
                db.refresh(existing_position)
                
                logger.info(
                    f"å·²æ›´æ–°ç¾æœ‰å€‰ä½ {symbol} ({position_side}) "
                    f"(position_id={existing_position.id}, exit_reason=manual_close, exit_price={exit_price})"
                )
                position = existing_position
            else:
                # æ²’æœ‰ç¾æœ‰å€‰ä½ï¼Œå»ºç«‹æ–°çš„ Position è¨˜éŒ„ï¼ˆç”¨æ–¼çµ±è¨ˆè¨ˆç®—ï¼‰
                position = Position(
                    bot_id=None,  # é bot å‰µå»ºçš„å€‰ä½
                    tv_signal_log_id=None,  # é bot å‰µå»ºçš„å€‰ä½
                    symbol=symbol.upper(),
                    side=position_side,
                    qty=qty,
                    entry_price=entry_price if entry_price > 0 else exit_price,  # å¦‚æœ entry_price ç„¡æ•ˆï¼Œä½¿ç”¨ exit_price ä½œç‚º fallback
                    exit_price=exit_price,
                    status="CLOSED",
                    closed_at=datetime.now(timezone.utc),
                    exit_reason="manual_close",  # æ‰‹å‹•é—œé–‰
                    binance_order_id=int(order.get("orderId")) if order.get("orderId") else None,
                    client_order_id=order.get("clientOrderId"),
                )
                
                db.add(position)
                db.commit()
                db.refresh(position)
                
                logger.info(
                    f"é bot å‰µå»ºå€‰ä½ {symbol} ({position_side}) å·²å»ºç«‹è³‡æ–™åº«è¨˜éŒ„ "
                    f"(position_id={position.id}, exit_reason=manual_close, exit_price={exit_price})"
                )
            
            # æ¸…ç†è¿½è¹¤è¨˜éŒ„
            tracking_key = f"{symbol}|{position_side}"
            if tracking_key in _non_bot_position_tracking:
                del _non_bot_position_tracking[tracking_key]
                logger.debug(f"æ¸…ç†é bot å€‰ä½è¿½è¹¤è¨˜éŒ„: {tracking_key}")
            
        except Exception as e:
            logger.error(f"æ›´æ–°/å»ºç«‹å€‰ä½ {symbol} ({position_side}) è³‡æ–™åº«è¨˜éŒ„å¤±æ•—: {e}")
            db.rollback()
            raise
        
        # è¿”å›é—œéµè³‡è¨Š
        return {
            "success": True,
            "symbol": symbol,
            "position_side": position_side,
            "order_id": order.get("orderId"),
            "client_order_id": order.get("clientOrderId"),
            "executed_qty": float(order.get("executedQty", 0) or 0),
            "avg_price": float(order.get("avgPrice", 0) or 0),
            "status": order.get("status"),
        }
        
    except HTTPException:
        raise
    except ValueError as e:
        logger.error(f"Binance API è¨­å®šéŒ¯èª¤: {e}")
        raise HTTPException(
            status_code=400,
            detail=f"Binance API æœªè¨­å®š: {str(e)}ã€‚è«‹åœ¨ç’°å¢ƒè®Šæ•¸ä¸­è¨­å®š BINANCE_API_KEY å’Œ BINANCE_API_SECRETã€‚"
        )
    except Exception as e:
        logger.exception("é—œé–‰ Binance Live Position å¤±æ•—")
        raise HTTPException(
            status_code=500,
            detail=f"é—œé–‰å€‰ä½å¤±æ•—: {str(e)}"
        )


@app.get("/settings/trailing", response_model=TrailingConfig)
async def get_trailing_settings(
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾— Trailing Stop å…¨åŸŸè¨­å®šã€‚
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        TrailingConfig: ç›®å‰çš„ Trailing è¨­å®š
    """
    return TRAILING_CONFIG


@app.post("/settings/trailing", response_model=TrailingConfig)
async def update_trailing_settings(
    payload: TrailingConfigUpdate,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    æ›´æ–° Trailing Stop å…¨åŸŸè¨­å®šã€‚
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    ç•¶ lock_ratio æ›´æ–°æ™‚ï¼Œæœƒè‡ªå‹•æ›´æ–°æ‰€æœ‰ OPEN ç‹€æ…‹çš„å€‰ä½ä¸­ï¼Œ
    åŸæœ¬ä½¿ç”¨èˆŠçš„å…¨å±€ lock_ratio çš„å€‰ä½ï¼ˆtrail_callback ç­‰æ–¼èˆŠå€¼ï¼‰ï¼Œ
    è®“å®ƒå€‘ä½¿ç”¨æ–°çš„ lock_ratioã€‚
    
    Args:
        payload: è¦æ›´æ–°çš„è¨­å®šï¼ˆåªæ›´æ–°æä¾›çš„æ¬„ä½ï¼‰
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
        db: è³‡æ–™åº« Session
    
    Returns:
        TrailingConfig: æ›´æ–°å¾Œçš„ Trailing è¨­å®š
    
    Raises:
        HTTPException: ç•¶è¨­å®šå€¼ç„¡æ•ˆæ™‚
    """
    global TRAILING_CONFIG
    
    # ä¿å­˜èˆŠçš„è¨­å®šå€¼ï¼ˆç”¨æ–¼æ›´æ–°ä½¿ç”¨èˆŠå…¨å±€å€¼çš„å€‰ä½ï¼‰
    old_long_config = TRAILING_CONFIG.long_config
    old_short_config = TRAILING_CONFIG.short_config
    
    # ä½¿ç”¨ dict() æ–¹æ³•ï¼ˆPydantic v1/v2 å…¼å®¹ï¼‰
    if hasattr(TRAILING_CONFIG, 'model_dump'):
        updated = TRAILING_CONFIG.model_dump()
    else:
        updated = TRAILING_CONFIG.dict()
    
    if hasattr(payload, 'model_dump'):
        data = payload.model_dump(exclude_unset=True)
    else:
        data = payload.dict(exclude_unset=True)
    
    # Force trailing_enabled and auto_close_enabled to always be True
    data['trailing_enabled'] = True
    data['auto_close_enabled'] = True
    
    # è™•ç†å‘å¾Œå…¼å®¹ï¼šå¦‚æœæä¾›äº†èˆŠæ ¼å¼çš„è¨­å®šï¼ˆprofit_threshold_pct, lock_ratio, base_sl_pctï¼‰ï¼Œ
    # åŒæ™‚æ›´æ–° LONG å’Œ SHORT çš„è¨­å®š
    if "profit_threshold_pct" in data or "lock_ratio" in data or "base_sl_pct" in data:
        if "long_config" not in data:
            if hasattr(TRAILING_CONFIG.long_config, 'model_dump'):
                data['long_config'] = TRAILING_CONFIG.long_config.model_dump()
            else:
                data['long_config'] = TRAILING_CONFIG.long_config.dict()
        if "short_config" not in data:
            if hasattr(TRAILING_CONFIG.short_config, 'model_dump'):
                data['short_config'] = TRAILING_CONFIG.short_config.model_dump()
            else:
                data['short_config'] = TRAILING_CONFIG.short_config.dict()
        
        # åŒæ™‚æ›´æ–° LONG å’Œ SHORT
        if "profit_threshold_pct" in data:
            data['long_config']['profit_threshold_pct'] = data['profit_threshold_pct']
            data['short_config']['profit_threshold_pct'] = data['profit_threshold_pct']
            del data['profit_threshold_pct']
        
        if "lock_ratio" in data:
            data['long_config']['lock_ratio'] = data['lock_ratio']
            data['short_config']['lock_ratio'] = data['lock_ratio']
            del data['lock_ratio']
        
        if "base_sl_pct" in data:
            data['long_config']['base_sl_pct'] = data['base_sl_pct']
            data['short_config']['base_sl_pct'] = data['base_sl_pct']
            del data['base_sl_pct']
    
    # è™•ç† LONG è¨­å®š
    if "long_config" in data:
        long_data = data['long_config']
        if not isinstance(long_data, dict):
            if hasattr(long_data, 'model_dump'):
                long_data = long_data.model_dump(exclude_unset=True)
            else:
                long_data = long_data.dict(exclude_unset=True)
        
        # é©—è­‰ LONG è¨­å®š
        if "lock_ratio" in long_data:
            if long_data["lock_ratio"] is not None and long_data["lock_ratio"] < 0:
                raise HTTPException(status_code=400, detail="LONG lock_ratio ä¸èƒ½å°æ–¼ 0")
            if long_data["lock_ratio"] is not None and long_data["lock_ratio"] > 1:
                logger.warning(f"LONG lock_ratio > 1ï¼ˆå€¼={long_data['lock_ratio']}ï¼‰ï¼Œå·²å¼·åˆ¶èª¿æ•´ç‚º 1.0")
                long_data["lock_ratio"] = 1.0
        
        if "profit_threshold_pct" in long_data:
            if long_data["profit_threshold_pct"] is not None and long_data["profit_threshold_pct"] < 0:
                raise HTTPException(status_code=400, detail="LONG profit_threshold_pct ä¸èƒ½å°æ–¼ 0")
        
        if "base_sl_pct" in long_data:
            if long_data["base_sl_pct"] is not None and long_data["base_sl_pct"] < 0:
                raise HTTPException(status_code=400, detail="LONG base_sl_pct ä¸èƒ½å°æ–¼ 0")
        
        # æ›´æ–° LONG è¨­å®š
        if 'long_config' not in updated:
            updated['long_config'] = {}
        updated['long_config'].update(long_data)
        data['long_config'] = updated['long_config']
    
    # è™•ç† SHORT è¨­å®š
    if "short_config" in data:
        short_data = data['short_config']
        if not isinstance(short_data, dict):
            if hasattr(short_data, 'model_dump'):
                short_data = short_data.model_dump(exclude_unset=True)
            else:
                short_data = short_data.dict(exclude_unset=True)
        
        # é©—è­‰ SHORT è¨­å®š
        if "lock_ratio" in short_data:
            if short_data["lock_ratio"] is not None and short_data["lock_ratio"] < 0:
                raise HTTPException(status_code=400, detail="SHORT lock_ratio ä¸èƒ½å°æ–¼ 0")
            if short_data["lock_ratio"] is not None and short_data["lock_ratio"] > 1:
                logger.warning(f"SHORT lock_ratio > 1ï¼ˆå€¼={short_data['lock_ratio']}ï¼‰ï¼Œå·²å¼·åˆ¶èª¿æ•´ç‚º 1.0")
                short_data["lock_ratio"] = 1.0
        
        if "profit_threshold_pct" in short_data:
            if short_data["profit_threshold_pct"] is not None and short_data["profit_threshold_pct"] < 0:
                raise HTTPException(status_code=400, detail="SHORT profit_threshold_pct ä¸èƒ½å°æ–¼ 0")
        
        if "base_sl_pct" in short_data:
            if short_data["base_sl_pct"] is not None and short_data["base_sl_pct"] < 0:
                raise HTTPException(status_code=400, detail="SHORT base_sl_pct ä¸èƒ½å°æ–¼ 0")
        
        # æ›´æ–° SHORT è¨­å®š
        if 'short_config' not in updated:
            updated['short_config'] = {}
        updated['short_config'].update(short_data)
        data['short_config'] = updated['short_config']
    
    # æ›´æ–°å…¶ä»–æ¬„ä½ï¼ˆtrailing_enabled, auto_close_enabledï¼‰
    for key in ['trailing_enabled', 'auto_close_enabled']:
        if key in data:
            updated[key] = data[key]
    
    # é‡æ–°å»ºç«‹ TrailingConfig ç‰©ä»¶
    TRAILING_CONFIG = TrailingConfig(**updated)
    
    # è™•ç†è¨­å®šæ›´æ–°å¾Œçš„å€‰ä½åŒæ­¥
    # ç•¶ä»»ä½• LONG æˆ– SHORT è¨­å®šè¢«æ›´æ–°æ™‚ï¼Œæ¸…é™¤å°æ‡‰ OPEN å€‰ä½çš„æ‰€æœ‰è¦†å¯«å€¼ï¼ˆè®“å®ƒå€‘ä½¿ç”¨æ–°çš„å…¨å±€é…ç½®ï¼‰
    for side_name, side_key, old_config in [("LONG", "long_config", old_long_config), 
                                           ("SHORT", "short_config", old_short_config)]:
        # æª¢æŸ¥è©² side_config æ˜¯å¦åœ¨æ›´æ–°æ•¸æ“šä¸­ï¼ˆè¡¨ç¤ºç”¨æˆ¶æ›´æ–°äº†é€™å€‹æ–¹å‘çš„è¨­å®šï¼‰
        side_config_key = f"{side_key}"
        side_config_updated = side_config_key in data
        
        if side_config_updated:
            # ç²å–æ›´æ–°å¾Œçš„é…ç½®
            new_config = getattr(TRAILING_CONFIG, side_key)
            
            # æª¢æŸ¥æ˜¯å¦æœ‰ä»»ä½•è¨­å®šè¢«æ”¹è®Š
            config_changed = (
                old_config.lock_ratio != new_config.lock_ratio or
                old_config.profit_threshold_pct != new_config.profit_threshold_pct or
                old_config.base_sl_pct != new_config.base_sl_pct
            )
            
            if config_changed:
                try:
                    # æ‰¾å‡ºæ‰€æœ‰ OPEN ç‹€æ…‹ä¸”å°æ‡‰æ–¹å‘çš„å€‰ä½
                    positions_to_update = (
                        db.query(Position)
                        .filter(
                            Position.status == "OPEN",
                            Position.side == side_name
                        )
                        .all()
                    )
                    
                    if positions_to_update:
                        cleared_trail_callback = 0
                        cleared_profit_threshold = 0
                        cleared_base_sl = 0
                        
                        for position in positions_to_update:
                            # æ¸…é™¤æ‰€æœ‰è¦†å¯«å€¼ï¼Œè®“å€‰ä½ä½¿ç”¨æ–°çš„å…¨å±€é…ç½®
                            if position.trail_callback is not None:
                                position.trail_callback = None
                                cleared_trail_callback += 1
                            
                            if position.dyn_profit_threshold_pct is not None:
                                position.dyn_profit_threshold_pct = None
                                cleared_profit_threshold += 1
                            
                            if position.base_stop_loss_pct is not None:
                                position.base_stop_loss_pct = None
                                cleared_base_sl += 1
                        
                        db.commit()
                        
                        changes = []
                        if old_config.lock_ratio != new_config.lock_ratio:
                            changes.append(f"lock_ratio: {old_config.lock_ratio}->{new_config.lock_ratio}")
                        if old_config.profit_threshold_pct != new_config.profit_threshold_pct:
                            changes.append(f"profit_threshold: {old_config.profit_threshold_pct}->{new_config.profit_threshold_pct}")
                        if old_config.base_sl_pct != new_config.base_sl_pct:
                            changes.append(f"base_sl: {old_config.base_sl_pct}->{new_config.base_sl_pct}")
                        
                        logger.info(
                            f"å·²æ¸…é™¤ {len(positions_to_update)} å€‹ {side_name} OPEN å€‰ä½çš„è¦†å¯«å€¼ï¼Œè®“å®ƒå€‘ä½¿ç”¨æ–°çš„å…¨å±€é…ç½® "
                            f"(è®Šæ›´: {', '.join(changes)}ï¼›æ¸…é™¤: trail_callback={cleared_trail_callback}, "
                            f"dyn_profit_threshold_pct={cleared_profit_threshold}, base_stop_loss_pct={cleared_base_sl})"
                        )
                except Exception as e:
                    logger.error(f"æ¸…é™¤ {side_name} å€‰ä½çš„è¦†å¯«å€¼æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}", exc_info=True)
                    db.rollback()
    
    # ä½¿ç”¨å…¼å®¹çš„åºåˆ—åŒ–æ–¹æ³•
    if hasattr(TRAILING_CONFIG, 'model_dump_json'):
        config_json = TRAILING_CONFIG.model_dump_json()
    else:
        import json
        config_json = json.dumps(TRAILING_CONFIG.dict())
    
    logger.info(f"æ›´æ–° Trailing è¨­å®š: {config_json}")
    return TRAILING_CONFIG


@app.get("/signals", response_model=List[dict])
async def list_signals(
    limit: int = 50,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾— TradingView Signal æ—¥èªŒåˆ—è¡¨
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        limit: è¿”å›çš„æœ€å¤§æ•¸é‡ï¼Œé è¨­ 50
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        List[dict]: Signal æ—¥èªŒåˆ—è¡¨
    """
    q = (
        db.query(TradingViewSignalLog)
        .order_by(TradingViewSignalLog.id.desc())
        .limit(limit)
    )
    rows = q.all()
    return [
        {
            "id": r.id,
            "bot_key": r.bot_key,
            "signal_id": r.signal_id,
            "symbol": r.symbol,
            "side": r.side,
            "qty": r.qty,
            "position_size": r.position_size,
            "received_at": r.received_at.isoformat() if r.received_at else None,
            "processed": r.processed,
            "process_result": r.process_result,
            "raw_payload": r.raw_payload,
        }
        for r in rows
    ]


@app.get("/signals/{signal_id}", response_model=dict)
async def get_signal_detail(
    signal_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾—å–®ç­† TradingView Signal è©³ç´°è³‡æ–™ï¼ˆåŒ…å« raw_payloadï¼‰
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        signal_id: Signal Log ID
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: Signal è©³ç´°è³‡æ–™ï¼ŒåŒ…å« raw_payload
    
    Raises:
        HTTPException: ç•¶ Signal ä¸å­˜åœ¨æ™‚
    """
    signal = db.query(TradingViewSignalLog).filter(TradingViewSignalLog.id == signal_id).first()
    if not signal:
        raise HTTPException(status_code=404, detail="Signal not found")
    
    return {
        "id": signal.id,
        "bot_key": signal.bot_key,
        "signal_id": signal.signal_id,
        "symbol": signal.symbol,
        "side": signal.side,
        "qty": signal.qty,
        "position_size": signal.position_size,
        "raw_body": signal.raw_body,
        "raw_payload": signal.raw_payload,
        "received_at": signal.received_at.isoformat() if signal.received_at else None,
        "processed": signal.processed,
        "process_result": signal.process_result,
    }


# ==================== Signal Config CRUD APIs ====================

@app.get("/signal-configs", response_model=List[TVSignalConfigOut])
async def list_signal_configs(
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾—æ‰€æœ‰ Signal Config åˆ—è¡¨
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        List[TVSignalConfigOut]: Signal Config åˆ—è¡¨
    """
    configs = db.query(TVSignalConfig).order_by(TVSignalConfig.id.desc()).all()
    return [
        TVSignalConfigOut.model_validate(c) if hasattr(TVSignalConfigOut, 'model_validate') else TVSignalConfigOut.from_orm(c)
        for c in configs
    ]


@app.post("/signal-configs", response_model=TVSignalConfigOut)
async def create_signal_config(
    config: TVSignalConfigCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å»ºç«‹æ–°çš„ Signal Config
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        config: Signal Config è³‡æ–™
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        TVSignalConfigOut: å»ºç«‹çš„ Signal Config
    
    Raises:
        HTTPException: ç•¶ signal_key å·²å­˜åœ¨æ™‚
    """
    # æª¢æŸ¥ signal_key æ˜¯å¦å·²å­˜åœ¨
    existing = db.query(TVSignalConfig).filter(TVSignalConfig.signal_key == config.signal_key).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"signal_key '{config.signal_key}' å·²å­˜åœ¨")
    
    # å»ºç«‹
    db_config = TVSignalConfig(
        name=config.name,
        signal_key=config.signal_key,
        description=config.description,
        symbol_hint=config.symbol_hint,
        timeframe_hint=config.timeframe_hint,
        enabled=config.enabled,
    )
    
    db.add(db_config)
    db.commit()
    db.refresh(db_config)
    
    logger.info(f"å»ºç«‹ Signal Config: {db_config.id} ({db_config.name}, signal_key={db_config.signal_key})")
    
    return TVSignalConfigOut.model_validate(db_config) if hasattr(TVSignalConfigOut, 'model_validate') else TVSignalConfigOut.from_orm(db_config)


@app.get("/signal-configs/{signal_id}", response_model=TVSignalConfigOut)
async def get_signal_config(
    signal_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾—å–®ä¸€ Signal Config
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        signal_id: Signal Config ID
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        TVSignalConfigOut: Signal Config
    
    Raises:
        HTTPException: ç•¶ Signal Config ä¸å­˜åœ¨æ™‚
    """
    config = db.query(TVSignalConfig).filter(TVSignalConfig.id == signal_id).first()
    if not config:
        raise HTTPException(status_code=404, detail=f"Signal Config {signal_id} ä¸å­˜åœ¨")
    
    return TVSignalConfigOut.model_validate(config) if hasattr(TVSignalConfigOut, 'model_validate') else TVSignalConfigOut.from_orm(config)


@app.put("/signal-configs/{signal_id}", response_model=TVSignalConfigOut)
async def update_signal_config(
    signal_id: int,
    config_update: TVSignalConfigUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    æ›´æ–° Signal Config
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        signal_id: Signal Config ID
        config_update: æ›´æ–°è³‡æ–™
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        TVSignalConfigOut: æ›´æ–°å¾Œçš„ Signal Config
    
    Raises:
        HTTPException: ç•¶ Signal Config ä¸å­˜åœ¨æˆ– signal_key å·²å­˜åœ¨æ™‚
    """
    config = db.query(TVSignalConfig).filter(TVSignalConfig.id == signal_id).first()
    if not config:
        raise HTTPException(status_code=404, detail=f"Signal Config {signal_id} ä¸å­˜åœ¨")
    
    # å¦‚æœæ›´æ–° signal_keyï¼Œæª¢æŸ¥æ˜¯å¦é‡è¤‡
    if config_update.signal_key is not None and config_update.signal_key != config.signal_key:
        existing = db.query(TVSignalConfig).filter(TVSignalConfig.signal_key == config_update.signal_key).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"signal_key '{config_update.signal_key}' å·²å­˜åœ¨")
    
    # å–å¾—æ›´æ–°è³‡æ–™
    update_data = config_update.model_dump(exclude_unset=True) if hasattr(config_update, 'model_dump') else config_update.dict(exclude_unset=True)
    
    # æ›´æ–°æ¬„ä½
    for key, value in update_data.items():
        if value is not None:
            setattr(config, key, value)
    
    db.commit()
    db.refresh(config)
    
    logger.info(f"æ›´æ–° Signal Config: {config.id} ({config.name})")
    
    return TVSignalConfigOut.model_validate(config) if hasattr(TVSignalConfigOut, 'model_validate') else TVSignalConfigOut.from_orm(config)


@app.delete("/signal-configs/{signal_id}")
async def delete_signal_config(
    signal_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    åˆªé™¤ Signal Config
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    æ³¨æ„ï¼šå¦‚æœè©² Signal Config ä¸‹å·²æœ‰é—œè¯çš„ Botsï¼Œå°‡ç„¡æ³•åˆªé™¤ï¼ˆéœ€å…ˆç§»é™¤é—œè¯æˆ–åˆªé™¤ Botsï¼‰ã€‚
    æˆ–è€…ï¼Œå¯ä»¥é¸æ“‡å°‡ enabled è¨­ç‚º Falseï¼ˆè»Ÿåˆªé™¤ï¼‰è€ŒéçœŸæ­£åˆªé™¤ã€‚
    
    Args:
        signal_id: Signal Config ID
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: åˆªé™¤çµæœ
    
    Raises:
        HTTPException: ç•¶ Signal Config ä¸å­˜åœ¨æˆ–ä»æœ‰é—œè¯çš„ Bots æ™‚
    """
    config = db.query(TVSignalConfig).filter(TVSignalConfig.id == signal_id).first()
    if not config:
        raise HTTPException(status_code=404, detail=f"Signal Config {signal_id} ä¸å­˜åœ¨")
    
    # æª¢æŸ¥æ˜¯å¦æœ‰é—œè¯çš„ Bots
    bots_count = db.query(BotConfig).filter(BotConfig.signal_id == signal_id).count()
    if bots_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"ç„¡æ³•åˆªé™¤ Signal Config {signal_id}ï¼šä»æœ‰ {bots_count} å€‹ Bot é—œè¯åˆ°æ­¤ Signalã€‚è«‹å…ˆç§»é™¤ Bot é—œè¯æˆ–å°‡ Signal è¨­ç‚º disabledã€‚"
        )
    
    db.delete(config)
    db.commit()
    
    logger.info(f"åˆªé™¤ Signal Config: {signal_id} ({config.name})")
    
    return {"success": True, "message": f"Signal Config {signal_id} å·²åˆªé™¤"}


@app.get("/bots", response_model=List[BotConfigOut])
async def list_bots(
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾—æ‰€æœ‰ Bot è¨­å®š
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        List[BotConfigOut]: Bot è¨­å®šåˆ—è¡¨ï¼ˆåŒ…å«é—œè¯çš„ Signal Config è³‡è¨Šï¼‰
    """
    bots = db.query(BotConfig).order_by(BotConfig.id.desc()).all()
    result = []
    for bot in bots:
        # è¼‰å…¥é—œè¯çš„ signalï¼ˆå¦‚æœæœ‰çš„è©±ï¼‰
        signal_config_obj = None
        if bot.signal_id:
            signal_config_obj = db.query(TVSignalConfig).filter(TVSignalConfig.id == bot.signal_id).first()
        
        # æ‰‹å‹•æ§‹å»º BotConfigOutï¼Œé¿å… relationship åºåˆ—åŒ–å•é¡Œ
        bot_out_dict = {
            "id": bot.id,
            "name": bot.name,
            "bot_key": bot.bot_key,
            "enabled": bot.enabled,
            "symbol": bot.symbol,
            "use_signal_side": bot.use_signal_side,
            "fixed_side": bot.fixed_side,
            "qty": bot.qty,
            "max_invest_usdt": bot.max_invest_usdt,
            "leverage": bot.leverage,
            "use_dynamic_stop": bot.use_dynamic_stop,
            "trailing_callback_percent": bot.trailing_callback_percent,
            "base_stop_loss_pct": bot.base_stop_loss_pct,
            "signal_id": bot.signal_id,
            "created_at": bot.created_at,
            "updated_at": bot.updated_at,
        }
        bot_out = BotConfigOut(**bot_out_dict)
        
        if signal_config_obj:
            signal_dict = {
                "id": signal_config_obj.id,
                "name": signal_config_obj.name,
                "signal_key": signal_config_obj.signal_key,
                "description": signal_config_obj.description,
                "symbol_hint": signal_config_obj.symbol_hint,
                "timeframe_hint": signal_config_obj.timeframe_hint,
                "enabled": signal_config_obj.enabled,
                "created_at": signal_config_obj.created_at,
                "updated_at": signal_config_obj.updated_at,
            }
            bot_out.signal = TVSignalConfigOut(**signal_dict)
        
        result.append(bot_out)
    return result


@app.post("/bots", response_model=BotConfigOut)
async def create_bot(
    bot: BotConfigCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å»ºç«‹æ–°çš„ Bot è¨­å®š
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        bot: Bot è¨­å®šè³‡æ–™
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        BotConfigOut: å»ºç«‹çš„ Bot è¨­å®š
    
    Raises:
        HTTPException: ç•¶è¨­å®šå€¼ç„¡æ•ˆæ™‚
    """
    # é©—è­‰
    # å¦‚æœè¨­å®šäº† max_invest_usdtï¼Œé©—è­‰å®ƒå¿…é ˆå¤§æ–¼ 0
    if bot.max_invest_usdt is not None:
        if bot.max_invest_usdt <= 0:
            raise HTTPException(status_code=400, detail="max_invest_usdt å¿…é ˆå¤§æ–¼ 0")
    # å¦‚æœæ²’æœ‰è¨­å®š max_invest_usdtï¼Œå‰‡ qty å¿…é ˆå¤§æ–¼ 0
    if bot.max_invest_usdt is None:
        if bot.qty <= 0:
            raise HTTPException(status_code=400, detail="qty å¿…é ˆå¤§æ–¼ 0ï¼ˆç•¶ max_invest_usdt æœªè¨­å®šæ™‚ï¼‰")
    
    if bot.trailing_callback_percent is not None:
        if bot.trailing_callback_percent < 0 or bot.trailing_callback_percent > 100:
            raise HTTPException(status_code=400, detail="trailing_callback_percent å¿…é ˆåœ¨ 0~100 ä¹‹é–“")
    
    # å¦‚æœ use_signal_side=Trueï¼Œå‰‡è‡ªå‹•å°‡ fixed_side=None
    if bot.use_signal_side:
        bot.fixed_side = None
    
    # å¦‚æœ fixed_side æœ‰å€¼ï¼Œè½‰æˆå¤§å¯«
    if bot.fixed_side:
        bot.fixed_side = bot.fixed_side.upper()
        if bot.fixed_side not in ["BUY", "SELL"]:
            raise HTTPException(status_code=400, detail="fixed_side å¿…é ˆæ˜¯ BUY æˆ– SELL")
    
    # æª¢æŸ¥ bot_key æ˜¯å¦å·²å­˜åœ¨
    existing = db.query(BotConfig).filter(BotConfig.bot_key == bot.bot_key).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"bot_key '{bot.bot_key}' å·²å­˜åœ¨")
    
    # å¦‚æœæä¾›äº† signal_idï¼Œé©—è­‰ Signal Config å­˜åœ¨ä¸” enabled
    if bot.signal_id is not None:
        signal_config = db.query(TVSignalConfig).filter(TVSignalConfig.id == bot.signal_id).first()
        if not signal_config:
            raise HTTPException(status_code=404, detail=f"æ‰¾ä¸åˆ° signal_id={bot.signal_id} çš„ Signal Config")
        if not signal_config.enabled:
            raise HTTPException(status_code=400, detail=f"Signal Config {bot.signal_id} æœªå•Ÿç”¨ï¼Œç„¡æ³•å»ºç«‹é—œè¯çš„ Bot")
    
    # å»ºç«‹
    db_bot = BotConfig(
        name=bot.name,
        bot_key=bot.bot_key,
        enabled=bot.enabled,
        symbol=bot.symbol.upper(),
        use_signal_side=bot.use_signal_side,
        fixed_side=bot.fixed_side,
        qty=bot.qty,
        max_invest_usdt=bot.max_invest_usdt,
        leverage=bot.leverage,
        use_dynamic_stop=bot.use_dynamic_stop,
        trailing_callback_percent=bot.trailing_callback_percent,
        base_stop_loss_pct=bot.base_stop_loss_pct,
        signal_id=bot.signal_id,
    )
    
    try:
        db.add(db_bot)
        db.commit()
        db.refresh(db_bot)
        logger.info(f"å»ºç«‹ Bot è¨­å®š: {db_bot.id} ({db_bot.name}, bot_key={db_bot.bot_key}, signal_id={db_bot.signal_id}, max_invest_usdt={db_bot.max_invest_usdt})")
    except Exception as e:
        db.rollback()
        logger.error(f"å»ºç«‹ Bot æ™‚ç™¼ç”Ÿè³‡æ–™åº«éŒ¯èª¤: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"å»ºç«‹ Bot æ™‚ç™¼ç”Ÿè³‡æ–™åº«éŒ¯èª¤: {str(e)}")
    
    # è¼‰å…¥é—œè¯çš„ signalï¼ˆå¦‚æœæœ‰çš„è©±ï¼‰
    signal_config_obj = None
    if db_bot.signal_id:
        signal_config_obj = db.query(TVSignalConfig).filter(TVSignalConfig.id == db_bot.signal_id).first()
    
    # æ§‹å»ºå›æ‡‰ï¼ŒåŒ…å« signal è³‡è¨Š
    # ç›´æ¥æ‰‹å‹•æ§‹å»ºï¼Œé¿å… relationship åºåˆ—åŒ–å•é¡Œ
    try:
        bot_out_dict = {
            "id": db_bot.id,
            "name": db_bot.name,
            "bot_key": db_bot.bot_key,
            "enabled": db_bot.enabled,
            "symbol": db_bot.symbol,
            "use_signal_side": db_bot.use_signal_side,
            "fixed_side": db_bot.fixed_side,
            "qty": db_bot.qty,
            "max_invest_usdt": db_bot.max_invest_usdt,
            "leverage": db_bot.leverage,
            "use_dynamic_stop": db_bot.use_dynamic_stop,
            "trailing_callback_percent": db_bot.trailing_callback_percent,
            "base_stop_loss_pct": db_bot.base_stop_loss_pct,
            "signal_id": db_bot.signal_id,
            "created_at": db_bot.created_at,
            "updated_at": db_bot.updated_at,
        }
        bot_out = BotConfigOut(**bot_out_dict)
        
        # å¦‚æœ signal å­˜åœ¨ï¼Œæ·»åŠ åˆ°å›æ‡‰ä¸­
        if signal_config_obj:
            signal_dict = {
                "id": signal_config_obj.id,
                "name": signal_config_obj.name,
                "signal_key": signal_config_obj.signal_key,
                "description": signal_config_obj.description,
                "symbol_hint": signal_config_obj.symbol_hint,
                "timeframe_hint": signal_config_obj.timeframe_hint,
                "enabled": signal_config_obj.enabled,
                "created_at": signal_config_obj.created_at,
                "updated_at": signal_config_obj.updated_at,
            }
            bot_out.signal = TVSignalConfigOut(**signal_dict)
    except Exception as e:
        logger.error(f"æ§‹å»º BotConfigOut å›æ‡‰æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"å»ºç«‹ Bot æˆåŠŸï¼Œä½†åºåˆ—åŒ–å›æ‡‰æ™‚ç™¼ç”ŸéŒ¯èª¤: {str(e)}")
    
    return bot_out


@app.get("/bots/{bot_id}", response_model=BotConfigOut)
async def get_bot(
    bot_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å–å¾—å–®ä¸€ Bot è¨­å®š
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        bot_id: Bot ID
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        BotConfigOut: Bot è¨­å®š
    
    Raises:
        HTTPException: ç•¶ Bot ä¸å­˜åœ¨æ™‚
    """
    bot = db.query(BotConfig).filter(BotConfig.id == bot_id).first()
    if not bot:
        raise HTTPException(status_code=404, detail=f"Bot {bot_id} ä¸å­˜åœ¨")
    
    # è¼‰å…¥é—œè¯çš„ signalï¼ˆå¦‚æœæœ‰çš„è©±ï¼‰
    signal_config_obj = None
    if bot.signal_id:
        signal_config_obj = db.query(TVSignalConfig).filter(TVSignalConfig.id == bot.signal_id).first()
    
    # æ§‹å»ºå›æ‡‰ï¼ŒåŒ…å« signal è³‡è¨Š
    bot_out_dict = {
        "id": bot.id,
        "name": bot.name,
        "bot_key": bot.bot_key,
        "enabled": bot.enabled,
        "symbol": bot.symbol,
        "use_signal_side": bot.use_signal_side,
        "fixed_side": bot.fixed_side,
        "qty": bot.qty,
        "max_invest_usdt": bot.max_invest_usdt,
        "leverage": bot.leverage,
        "use_dynamic_stop": bot.use_dynamic_stop,
        "trailing_callback_percent": bot.trailing_callback_percent,
        "base_stop_loss_pct": bot.base_stop_loss_pct,
        "signal_id": bot.signal_id,
        "created_at": bot.created_at,
        "updated_at": bot.updated_at,
    }
    bot_out = BotConfigOut(**bot_out_dict)
    
    if signal_config_obj:
        signal_dict = {
            "id": signal_config_obj.id,
            "name": signal_config_obj.name,
            "signal_key": signal_config_obj.signal_key,
            "description": signal_config_obj.description,
            "symbol_hint": signal_config_obj.symbol_hint,
            "timeframe_hint": signal_config_obj.timeframe_hint,
            "enabled": signal_config_obj.enabled,
            "created_at": signal_config_obj.created_at,
            "updated_at": signal_config_obj.updated_at,
        }
        bot_out.signal = TVSignalConfigOut(**signal_dict)
    
    return bot_out


@app.put("/bots/{bot_id}", response_model=BotConfigOut)
async def update_bot(
    bot_id: int,
    bot_update: BotConfigUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    æ›´æ–° Bot è¨­å®š
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        bot_id: Bot ID
        bot_update: è¦æ›´æ–°çš„æ¬„ä½
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        BotConfigOut: æ›´æ–°å¾Œçš„ Bot è¨­å®š
    
    Raises:
        HTTPException: ç•¶ Bot ä¸å­˜åœ¨æˆ–è¨­å®šå€¼ç„¡æ•ˆæ™‚
    """
    bot = db.query(BotConfig).filter(BotConfig.id == bot_id).first()
    if not bot:
        raise HTTPException(status_code=404, detail=f"Bot {bot_id} ä¸å­˜åœ¨")
    
    # å–å¾—æ›´æ–°è³‡æ–™
    update_data = bot_update.model_dump(exclude_unset=True) if hasattr(bot_update, 'model_dump') else bot_update.dict(exclude_unset=True)
    
    # é©—è­‰ï¼šå¦‚æœæ›´æ–° max_invest_usdtï¼Œéœ€è¦æª¢æŸ¥å¯†ç¢¼
    if "max_invest_usdt" in update_data:
        new_max_invest = update_data["max_invest_usdt"]
        old_max_invest = bot.max_invest_usdt
        
        # æª¢æŸ¥æ˜¯å¦çœŸçš„æ”¹è®Šäº†ï¼ˆå¦‚æœæ˜¯æ–°å€¼èˆ‡èˆŠå€¼ç›¸åŒï¼Œä¸éœ€è¦å¯†ç¢¼ï¼‰
        # è™•ç† None/null å’Œæ•¸å€¼æ¯”è¼ƒ
        has_changed = False
        if old_max_invest is None and new_max_invest is not None:
            has_changed = True
        elif old_max_invest is not None and new_max_invest is None:
            has_changed = True
        elif old_max_invest is not None and new_max_invest is not None:
            # æ•¸å€¼æ¯”è¼ƒï¼Œä½¿ç”¨å°çš„å®¹å·®å€¼ä¾†è™•ç†æµ®é»æ•¸ç²¾åº¦å•é¡Œ
            has_changed = abs(float(old_max_invest) - float(new_max_invest)) > 0.0001
        
        if has_changed:
            # max_invest_usdt è¢«æ”¹è®Šäº†ï¼Œéœ€è¦å¯†ç¢¼é©—è­‰
            required_password = os.getenv("MAX_INVEST_PASSWORD", "")
            if not required_password:
                logger.warning("MAX_INVEST_PASSWORD æœªè¨­å®šåœ¨ç’°å¢ƒè®Šæ•¸ä¸­ï¼Œè·³éå¯†ç¢¼é©—è­‰")
            else:
                provided_password = update_data.get("max_invest_password")
                if not provided_password or provided_password != required_password:
                    logger.warning(f"Bot {bot_id} æ›´æ–° max_invest_usdt æ™‚å¯†ç¢¼é©—è­‰å¤±æ•—")
                    raise HTTPException(
                        status_code=403,
                        detail="æ›´æ–° Max Invest USDT éœ€è¦å¯†ç¢¼é©—è­‰ï¼Œå¯†ç¢¼ä¸æ­£ç¢º"
                    )
        
        if new_max_invest is not None and new_max_invest <= 0:
            raise HTTPException(status_code=400, detail="max_invest_usdt å¿…é ˆå¤§æ–¼ 0")
    if "qty" in update_data and update_data["qty"] is not None and update_data["qty"] <= 0:
        raise HTTPException(status_code=400, detail="qty å¿…é ˆå¤§æ–¼ 0ï¼ˆç•¶ max_invest_usdt æœªè¨­å®šæ™‚ï¼‰")
    
    if "trailing_callback_percent" in update_data and update_data["trailing_callback_percent"] is not None:
        if update_data["trailing_callback_percent"] < 0 or update_data["trailing_callback_percent"] > 100:
            raise HTTPException(status_code=400, detail="trailing_callback_percent å¿…é ˆåœ¨ 0~100 ä¹‹é–“")
    
    # å¦‚æœ use_signal_side=Trueï¼Œå‰‡è‡ªå‹•å°‡ fixed_side=None
    if update_data.get("use_signal_side") is True:
        update_data["fixed_side"] = None
    
    # å¦‚æœ fixed_side æœ‰å€¼ï¼Œè½‰æˆå¤§å¯«
    if "fixed_side" in update_data and update_data["fixed_side"]:
        update_data["fixed_side"] = update_data["fixed_side"].upper()
        if update_data["fixed_side"] not in ["BUY", "SELL"]:
            raise HTTPException(status_code=400, detail="fixed_side å¿…é ˆæ˜¯ BUY æˆ– SELL")
    
    # å¦‚æœæ›´æ–° signal_idï¼Œé©—è­‰ Signal Config å­˜åœ¨ä¸” enabled
    if "signal_id" in update_data and update_data["signal_id"] is not None:
        signal_config = db.query(TVSignalConfig).filter(TVSignalConfig.id == update_data["signal_id"]).first()
        if not signal_config:
            raise HTTPException(status_code=404, detail=f"æ‰¾ä¸åˆ° signal_id={update_data['signal_id']} çš„ Signal Config")
        if not signal_config.enabled:
            raise HTTPException(status_code=400, detail=f"Signal Config {update_data['signal_id']} æœªå•Ÿç”¨ï¼Œç„¡æ³•é—œè¯åˆ°æ­¤ Bot")
    
    # æ›´æ–°
    for key, value in update_data.items():
        if value is not None:
            setattr(bot, key, value)
    
    # è‡ªå‹•æ›´æ–° updated_atï¼ˆSQLAlchemy çš„ onupdate æœƒè™•ç†ï¼Œä½†æˆ‘å€‘ç¢ºä¿ä¸€ä¸‹ï¼‰
    from datetime import datetime, timezone
    bot.updated_at = datetime.now(timezone.utc)
    
    db.commit()
    db.refresh(bot)
    
    # è¼‰å…¥é—œè¯çš„ signalï¼ˆå¦‚æœæœ‰çš„è©±ï¼‰
    signal_config_obj = None
    if bot.signal_id:
        signal_config_obj = db.query(TVSignalConfig).filter(TVSignalConfig.id == bot.signal_id).first()
    
    logger.info(f"æ›´æ–° Bot è¨­å®š: {bot.id} ({bot.name}, signal_id={bot.signal_id})")
    
    # æ§‹å»ºå›æ‡‰ï¼ŒåŒ…å« signal è³‡è¨Š
    bot_out_dict = {
        "id": bot.id,
        "name": bot.name,
        "bot_key": bot.bot_key,
        "enabled": bot.enabled,
        "symbol": bot.symbol,
        "use_signal_side": bot.use_signal_side,
        "fixed_side": bot.fixed_side,
        "qty": bot.qty,
        "max_invest_usdt": bot.max_invest_usdt,
        "leverage": bot.leverage,
        "use_dynamic_stop": bot.use_dynamic_stop,
        "trailing_callback_percent": bot.trailing_callback_percent,
        "base_stop_loss_pct": bot.base_stop_loss_pct,
        "signal_id": bot.signal_id,
        "created_at": bot.created_at,
        "updated_at": bot.updated_at,
    }
    bot_out = BotConfigOut(**bot_out_dict)
    
    if signal_config_obj:
        signal_dict = {
            "id": signal_config_obj.id,
            "name": signal_config_obj.name,
            "signal_key": signal_config_obj.signal_key,
            "description": signal_config_obj.description,
            "symbol_hint": signal_config_obj.symbol_hint,
            "timeframe_hint": signal_config_obj.timeframe_hint,
            "enabled": signal_config_obj.enabled,
            "created_at": signal_config_obj.created_at,
            "updated_at": signal_config_obj.updated_at,
        }
        bot_out.signal = TVSignalConfigOut(**signal_dict)
    
    return bot_out


@app.post("/bots/{bot_id}/enable")
async def enable_bot(
    bot_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    å•Ÿç”¨ Bot
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        bot_id: Bot ID
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: æ“ä½œçµæœ
    
    Raises:
        HTTPException: ç•¶ Bot ä¸å­˜åœ¨æ™‚
    """
    bot = db.query(BotConfig).filter(BotConfig.id == bot_id).first()
    if not bot:
        raise HTTPException(status_code=404, detail=f"Bot {bot_id} ä¸å­˜åœ¨")
    
    bot.enabled = True
    bot.updated_at = datetime.now(timezone.utc)
    db.commit()
    
    logger.info(f"å•Ÿç”¨ Bot: {bot.id} ({bot.name})")
    
    return {"status": "enabled", "bot_id": bot.id}


@app.post("/bots/{bot_id}/disable")
async def disable_bot(
    bot_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    åœç”¨ Bot
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        bot_id: Bot ID
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: æ“ä½œçµæœ
    
    Raises:
        HTTPException: ç•¶ Bot ä¸å­˜åœ¨æ™‚
    """
    bot = db.query(BotConfig).filter(BotConfig.id == bot_id).first()
    if not bot:
        raise HTTPException(status_code=404, detail=f"Bot {bot_id} ä¸å­˜åœ¨")
    
    bot.enabled = False
    bot.updated_at = datetime.now(timezone.utc)
    db.commit()
    
    logger.info(f"åœç”¨ Bot: {bot.id} ({bot.name})")
    
    return {"status": "disabled", "bot_id": bot.id}


@app.delete("/bots/{bot_id}")
async def delete_bot(
    bot_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user)
):
    """
    åˆªé™¤ Bot
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    æ³¨æ„ï¼šåˆªé™¤ Bot å‰æœƒæª¢æŸ¥æ˜¯å¦æœ‰é—œè¯çš„ OPEN å€‰ä½ã€‚å¦‚æœæœ‰ï¼Œå»ºè­°å…ˆé—œé–‰å€‰ä½å†åˆªé™¤ Botã€‚
    
    Args:
        bot_id: Bot ID
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: åˆªé™¤çµæœ
    
    Raises:
        HTTPException: ç•¶ Bot ä¸å­˜åœ¨æ™‚
    """
    bot = db.query(BotConfig).filter(BotConfig.id == bot_id).first()
    if not bot:
        raise HTTPException(status_code=404, detail=f"Bot {bot_id} ä¸å­˜åœ¨")
    
    # æª¢æŸ¥æ˜¯å¦æœ‰é—œè¯çš„ OPEN å€‰ä½
    open_positions_count = db.query(Position).filter(
        Position.bot_id == bot_id,
        Position.status == "OPEN"
    ).count()
    
    if open_positions_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"ç„¡æ³•åˆªé™¤ Bot {bot_id}ï¼šä»æœ‰ {open_positions_count} å€‹ OPEN å€‰ä½é—œè¯åˆ°æ­¤ Botã€‚è«‹å…ˆé—œé–‰å€‰ä½å†åˆªé™¤ Botã€‚"
        )
    
    bot_name = bot.name
    db.delete(bot)
    db.commit()
    
    logger.info(f"åˆªé™¤ Bot: {bot_id} ({bot_name})")
    
    return {"success": True, "message": f"Bot {bot_id} ({bot_name}) å·²åˆªé™¤"}


class BulkUpdateInvestAmountRequest(BaseModel):
    """æ‰¹é‡æ›´æ–°æŠ•è³‡é‡‘é¡çš„è«‹æ±‚æ ¼å¼"""
    max_invest_usdt: float = Field(..., gt=0, description="æ–°çš„æŠ•è³‡é‡‘é¡ï¼ˆUSDTï¼‰ï¼Œå¿…é ˆå¤§æ–¼ 0")
    bot_ids: Optional[List[int]] = Field(None, description="å¯é¸çš„ Bot ID åˆ—è¡¨ï¼Œå¦‚æœæä¾›å‰‡åƒ…æ›´æ–°é€™äº› Botï¼Œå¦å‰‡æ›´æ–°æ‰€æœ‰ Bot")
    max_invest_password: str = Field(..., description="æ›´æ–° max_invest_usdt æ™‚éœ€è¦çš„å¯†ç¢¼")


@app.post("/bots/bulk-update-invest-amount")
async def bulk_update_invest_amount(
    request: BulkUpdateInvestAmountRequest,
    user: dict = Depends(require_admin_user)
):
    """
    æ‰¹é‡èª¿æ•´æ‰€æœ‰ Bot çš„æŠ•è³‡é‡‘é¡ï¼ˆmax_invest_usdtï¼‰

    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    å¯ä»¥ä¸€æ¬¡æ›´æ–°æ‰€æœ‰ Botï¼Œæˆ–åƒ…æ›´æ–°æŒ‡å®šçš„ Bot IDsã€‚
    éœ€è¦æä¾›å¯†ç¢¼é©—è­‰ã€‚

    Args:
        request: åŒ…å« max_invest_usdtã€å¯é¸çš„ bot_ids å’Œå¿…éœ€çš„å¯†ç¢¼
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰

    Returns:
        dict: åŒ…å«ä»¥ä¸‹æ¬„ä½çš„å­—å…¸ï¼š
            - success: æ˜¯å¦æˆåŠŸ
            - updated_count: æ›´æ–°çš„ Bot æ•¸é‡
            - bot_ids: å·²æ›´æ–°çš„ Bot ID åˆ—è¡¨
            - message: æ“ä½œçµæœè¨Šæ¯

    Raises:
        HTTPException: ç•¶ max_invest_usdt ç„¡æ•ˆæˆ–å¯†ç¢¼ä¸æ­£ç¢ºæ™‚
    """
    # é©—è­‰å¯†ç¢¼
    required_password = os.getenv("MAX_INVEST_PASSWORD", "")
    if not required_password:
        logger.warning("MAX_INVEST_PASSWORD æœªè¨­å®šåœ¨ç’°å¢ƒè®Šæ•¸ä¸­ï¼Œè·³éå¯†ç¢¼é©—è­‰")
    else:
        if not request.max_invest_password or request.max_invest_password != required_password:
            logger.warning("æ‰¹é‡æ›´æ–° max_invest_usdt æ™‚å¯†ç¢¼é©—è­‰å¤±æ•—")
            raise HTTPException(
                status_code=403,
                detail="æ›´æ–° Max Invest USDT éœ€è¦å¯†ç¢¼é©—è­‰ï¼Œå¯†ç¢¼ä¸æ­£ç¢º"
            )
    
    try:
        result = update_all_bots_invest_amount(
            max_invest_usdt=request.max_invest_usdt,
            bot_ids=request.bot_ids
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"æ‰¹é‡æ›´æ–° Bot æŠ•è³‡é‡‘é¡å¤±æ•—: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"æ‰¹é‡æ›´æ–°å¤±æ•—: {str(e)}")


# ==================== Position Stop Config API ====================

class PositionStopConfigUpdate(BaseModel):
    """æ›´æ–°å€‰ä½åœæé…ç½®çš„è«‹æ±‚æ ¼å¼"""
    dyn_profit_threshold_pct: Optional[float] = Field(None, description="PnL% é–€æª»ç™¾åˆ†æ¯”è¦†å¯«ï¼ˆä¾‹å¦‚ 1.0 è¡¨ç¤º 1%ï¼‰ï¼Œnull è¡¨ç¤ºä½¿ç”¨å…¨å±€é…ç½®")
    base_stop_loss_pct: Optional[float] = Field(None, description="åŸºç¤åœæç™¾åˆ†æ¯”è¦†å¯«ï¼ˆä¾‹å¦‚ 0.5 è¡¨ç¤º 0.5%ï¼‰ï¼Œnull è¡¨ç¤ºä½¿ç”¨å…¨å±€é…ç½®")
    trail_callback: Optional[float] = Field(None, description="é–åˆ©æ¯”ä¾‹è¦†å¯«ï¼ˆ0~1ï¼‰ï¼Œ0 è¡¨ç¤ºåƒ…ä½¿ç”¨ base stopï¼Œnull è¡¨ç¤ºä½¿ç”¨å…¨å±€é…ç½®")
    clear_overrides: bool = Field(False, description="å¦‚æœç‚º trueï¼Œæ¸…é™¤æ‰€æœ‰è¦†å¯«å€¼ï¼ˆè¨­ç‚º nullï¼‰")


class PositionMechanismConfigUpdate(BaseModel):
    """æ›´æ–°å€‰ä½åœæ/æ­¢ç›ˆæ©Ÿåˆ¶å•Ÿç”¨ç‹€æ…‹çš„è«‹æ±‚æ ¼å¼"""
    bot_stop_loss_enabled: Optional[bool] = Field(None, description="æ˜¯å¦å•Ÿç”¨ Bot å…§å»ºçš„åœææ©Ÿåˆ¶ï¼ˆdynamic stop / base stopï¼‰")
    tv_signal_close_enabled: Optional[bool] = Field(None, description="æ˜¯å¦å•Ÿç”¨ TradingView è¨Šè™Ÿé—œå€‰æ©Ÿåˆ¶ï¼ˆposition_size=0ï¼‰")


class PositionMechanismConfigUpdate(BaseModel):
    """æ›´æ–°å€‰ä½åœæ/æ­¢ç›ˆæ©Ÿåˆ¶å•Ÿç”¨ç‹€æ…‹çš„è«‹æ±‚æ ¼å¼"""
    bot_stop_loss_enabled: Optional[bool] = Field(None, description="æ˜¯å¦å•Ÿç”¨ Bot å…§å»ºçš„åœææ©Ÿåˆ¶ï¼ˆdynamic stop / base stopï¼‰")
    tv_signal_close_enabled: Optional[bool] = Field(None, description="æ˜¯å¦å•Ÿç”¨ TradingView è¨Šè™Ÿé—œå€‰æ©Ÿåˆ¶ï¼ˆposition_size=0ï¼‰")


@app.patch("/positions/{pos_id}/stop-config", response_model=PositionOut)
async def update_position_stop_config(
    pos_id: int,
    update: PositionStopConfigUpdate,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    æ›´æ–°å€‰ä½çš„åœæé…ç½®è¦†å¯«å€¼
    
    å…è¨±ç‚ºæ¯ç­†å€‰ä½è¨­å®šç¨ç«‹çš„åœæé…ç½®ï¼Œè¦†å¯«å…¨å±€é…ç½®ï¼š
    - dyn_profit_threshold_pct: PnL% é–€æª»ç™¾åˆ†æ¯”
    - base_stop_loss_pct: åŸºç¤åœæç™¾åˆ†æ¯”
    - trail_callback: é–åˆ©æ¯”ä¾‹ï¼ˆ0~1ï¼‰ï¼Œ0 è¡¨ç¤ºåƒ…ä½¿ç”¨ base stop
    
    å¦‚æœ clear_overrides=Trueï¼Œæœƒæ¸…é™¤æ‰€æœ‰è¦†å¯«å€¼ï¼ˆè¨­ç‚º nullï¼‰ï¼Œæ¢å¾©ä½¿ç”¨å…¨å±€é…ç½®ã€‚
    
    Args:
        pos_id: å€‰ä½ ID
        update: æ›´æ–°è«‹æ±‚
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Š
        db: è³‡æ–™åº« Session
    
    Returns:
        PositionOut: æ›´æ–°å¾Œçš„å€‰ä½è³‡è¨Š
    
    Raises:
        HTTPException: ç•¶å€‰ä½ä¸å­˜åœ¨æ™‚
    """
    position = db.query(Position).filter(Position.id == pos_id).first()
    if not position:
        raise HTTPException(status_code=404, detail="Position not found")
    
    if update.clear_overrides:
        position.dyn_profit_threshold_pct = None
        position.base_stop_loss_pct = None
        # trail_callback ä¿ç•™åŸå€¼ï¼Œé™¤éæ˜ç¢ºæŒ‡å®š
        if update.trail_callback is not None:
            position.trail_callback = update.trail_callback
    else:
        if update.dyn_profit_threshold_pct is not None:
            if update.dyn_profit_threshold_pct < 0:
                logger.warning(
                    f"å€‰ä½ {position.id} dyn_profit_threshold_pct < 0 ({update.dyn_profit_threshold_pct})ï¼Œå·²è¨­ç‚º 0"
                )
                position.dyn_profit_threshold_pct = 0.0
            else:
                position.dyn_profit_threshold_pct = update.dyn_profit_threshold_pct
        
        if update.base_stop_loss_pct is not None:
            if update.base_stop_loss_pct < 0:
                logger.warning(
                    f"å€‰ä½ {position.id} base_stop_loss_pct < 0 ({update.base_stop_loss_pct})ï¼Œå·²è¨­ç‚º 0"
                )
                position.base_stop_loss_pct = 0.0
            else:
                position.base_stop_loss_pct = update.base_stop_loss_pct
        
        if update.trail_callback is not None:
            # clamp to [0, 1] and handle 0 meaning "base stop only"
            val = update.trail_callback
            if val < 0:
                logger.warning(
                    f"å€‰ä½ {position.id} trail_callback < 0 ({val})ï¼Œå·²è¨­ç‚º 0 (base-stop only)"
                )
                val = 0.0
            elif val > 1:
                logger.warning(
                    f"å€‰ä½ {position.id} trail_callback > 1 ({val})ï¼Œå·²èª¿æ•´ç‚º 1.0"
                )
                val = 1.0
            position.trail_callback = val
    
    db.commit()
    db.refresh(position)
    
    logger.info(
        f"å€‰ä½ {position.id} ({position.symbol}) åœæé…ç½®å·²æ›´æ–°ï¼š"
        f"dyn_profit_threshold_pct={position.dyn_profit_threshold_pct}, "
        f"base_stop_loss_pct={position.base_stop_loss_pct}, "
        f"trail_callback={position.trail_callback}"
    )
    
    # è¨ˆç®—å¯¦éš›ä½¿ç”¨çš„å€¼å’Œä¾†æºæ¨™è¨˜ï¼ˆèˆ‡ get_positions ä¸­çš„é‚è¼¯ä¸€è‡´ï¼‰
    pos_dict = position.to_dict()
    
    # å–å¾—å°æ‡‰æ–¹å‘çš„å…¨å±€è¨­å®š
    side_config = TRAILING_CONFIG.get_config_for_side(position.side)
    
    # Profit Threshold
    profit_threshold_value = None
    profit_threshold_source = None
    if position.dyn_profit_threshold_pct is not None:
        profit_threshold_value = position.dyn_profit_threshold_pct
        profit_threshold_source = "override"
    elif side_config.profit_threshold_pct is not None:
        profit_threshold_value = side_config.profit_threshold_pct
        profit_threshold_source = "global"
    else:
        profit_threshold_value = DYN_PROFIT_THRESHOLD_PCT
        profit_threshold_source = "default"
    
    # Lock Ratio
    lock_ratio_value = None
    lock_ratio_source = None
    if position.trail_callback is not None:
        lock_ratio_value = position.trail_callback
        lock_ratio_source = "override"
    elif side_config.lock_ratio is not None:
        lock_ratio_value = side_config.lock_ratio
        lock_ratio_source = "global"
    else:
        lock_ratio_value = DYN_LOCK_RATIO_DEFAULT
        lock_ratio_source = "default"
    
    # Base SL%
    base_sl_value = None
    base_sl_source = None
    if position.base_stop_loss_pct is not None:
        base_sl_value = position.base_stop_loss_pct
        base_sl_source = "override"
    elif side_config.base_sl_pct is not None:
        base_sl_value = side_config.base_sl_pct
        base_sl_source = "global"
    else:
        base_sl_value = DYN_BASE_SL_PCT
        base_sl_source = "default"
    
    # æ·»åŠ é¡å¤–å­—æ®µ
    pos_dict.update({
        "profit_threshold_value": profit_threshold_value,
        "profit_threshold_source": profit_threshold_source,
        "lock_ratio_value": lock_ratio_value,
        "lock_ratio_source": lock_ratio_source,
        "base_sl_value": base_sl_value,
        "base_sl_source": base_sl_source,
    })
    
    # ä½¿ç”¨ from_orm æˆ– model_validate è¿”å› PositionOut
    return PositionOut(**pos_dict)


@app.patch("/positions/{pos_id}/mechanism-config", response_model=PositionOut)
async def update_position_mechanism_config(
    pos_id: int,
    update: PositionMechanismConfigUpdate,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    æ›´æ–°å€‰ä½çš„åœæ/æ­¢ç›ˆæ©Ÿåˆ¶å•Ÿç”¨ç‹€æ…‹
    
    å…è¨±ç‚ºæ¯ç­†å€‰ä½ç¨ç«‹æ§åˆ¶å…©ç¨®åœæ/æ­¢ç›ˆæ©Ÿåˆ¶ï¼š
    - bot_stop_loss_enabled: Bot å…§å»ºçš„åœææ©Ÿåˆ¶ï¼ˆdynamic stop / base stopï¼‰
    - tv_signal_close_enabled: TradingView è¨Šè™Ÿé—œå€‰æ©Ÿåˆ¶ï¼ˆposition_size=0ï¼‰
    
    é è¨­å€¼éƒ½æ˜¯ Trueï¼ˆå…©ç¨®æ©Ÿåˆ¶éƒ½å•Ÿç”¨ï¼‰ã€‚
    å¯ä»¥è¨­å®šç‚º False ä¾†åœç”¨ç‰¹å®šæ©Ÿåˆ¶ã€‚
    
    ç¯„ä¾‹ï¼š
    - bot_stop_loss_enabled=True, tv_signal_close_enabled=False: åªä½¿ç”¨ Bot åœæï¼Œå¿½ç•¥ TradingView é—œå€‰è¨Šè™Ÿ
    - bot_stop_loss_enabled=False, tv_signal_close_enabled=True: åªä½¿ç”¨ TradingView é—œå€‰è¨Šè™Ÿï¼Œä¸ä½¿ç”¨ Bot åœæ
    - bot_stop_loss_enabled=True, tv_signal_close_enabled=True: å…©ç¨®æ©Ÿåˆ¶éƒ½å•Ÿç”¨ï¼ˆé è¨­ï¼‰
    
    Args:
        pos_id: å€‰ä½ ID
        update: æ›´æ–°è«‹æ±‚
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Š
        db: è³‡æ–™åº« Session
    
    Returns:
        PositionOut: æ›´æ–°å¾Œçš„å€‰ä½è³‡è¨Š
    
    Raises:
        HTTPException: ç•¶å€‰ä½ä¸å­˜åœ¨æ™‚
    """
    position = db.query(Position).filter(Position.id == pos_id).first()
    if not position:
        raise HTTPException(status_code=404, detail="Position not found")
    
    if position.status != "OPEN":
        raise HTTPException(
            status_code=400, 
            detail=f"åªèƒ½æ›´æ–° OPEN ç‹€æ…‹çš„å€‰ä½ï¼Œç•¶å‰ç‹€æ…‹: {position.status}"
        )
    
    # æ›´æ–°æ¨™èªŒ
    if update.bot_stop_loss_enabled is not None:
        position.bot_stop_loss_enabled = update.bot_stop_loss_enabled
        logger.info(
            f"å€‰ä½ {position.id} ({position.symbol}) bot_stop_loss_enabled "
            f"å·²æ›´æ–°ç‚º {update.bot_stop_loss_enabled}"
        )
    
    if update.tv_signal_close_enabled is not None:
        position.tv_signal_close_enabled = update.tv_signal_close_enabled
        logger.info(
            f"å€‰ä½ {position.id} ({position.symbol}) tv_signal_close_enabled "
            f"å·²æ›´æ–°ç‚º {update.tv_signal_close_enabled}"
        )
    
    db.commit()
    db.refresh(position)
    
    logger.info(
        f"å€‰ä½ {position.id} ({position.symbol}) æ©Ÿåˆ¶é…ç½®å·²æ›´æ–°ï¼š"
        f"bot_stop_loss_enabled={position.bot_stop_loss_enabled}, "
        f"tv_signal_close_enabled={position.tv_signal_close_enabled}"
    )
    
    # ä½¿ç”¨ to_dict() ç„¶å¾Œè½‰æ›ç‚º PositionOutï¼Œç¢ºä¿åŒ…å«æ‰€æœ‰æ¬„ä½
    pos_dict = position.to_dict()
    return PositionOut(**pos_dict)


class BinancePositionStopConfigUpdate(BaseModel):
    """æ›´æ–° Binance Live Position åœæé…ç½®çš„è«‹æ±‚æ ¼å¼"""
    symbol: str = Field(..., description="äº¤æ˜“å°ï¼Œä¾‹å¦‚ï¼šBTCUSDT")
    position_side: str = Field(..., description="å€‰ä½æ–¹å‘ï¼šLONG æˆ– SHORT")
    dyn_profit_threshold_pct: Optional[float] = Field(None, description="PnL% é–€æª»ç™¾åˆ†æ¯”è¦†å¯«ï¼ˆä¾‹å¦‚ 1.0 è¡¨ç¤º 1%ï¼‰ï¼Œnull è¡¨ç¤ºä½¿ç”¨å…¨å±€é…ç½®")
    base_stop_loss_pct: Optional[float] = Field(None, description="åŸºç¤åœæç™¾åˆ†æ¯”è¦†å¯«ï¼ˆä¾‹å¦‚ 0.5 è¡¨ç¤º 0.5%ï¼‰ï¼Œnull è¡¨ç¤ºä½¿ç”¨å…¨å±€é…ç½®")
    trail_callback: Optional[float] = Field(None, description="é–åˆ©æ¯”ä¾‹è¦†å¯«ï¼ˆ0~1ï¼‰ï¼Œ0 è¡¨ç¤ºåƒ…ä½¿ç”¨ base stopï¼Œnull è¡¨ç¤ºä½¿ç”¨å…¨å±€é…ç½®")
    clear_overrides: bool = Field(False, description="å¦‚æœç‚º trueï¼Œæ¸…é™¤æ‰€æœ‰è¦†å¯«å€¼ï¼ˆè¨­ç‚º nullï¼‰")


class PortfolioTrailingConfigOut(BaseModel):
    """Portfolio Trailing Stop è¨­å®šæ¨¡å‹ï¼ˆAPI å›æ‡‰ç”¨ï¼‰"""
    enabled: bool = Field(False, description="æ˜¯å¦å•Ÿç”¨è‡ªå‹•è³£å‡º")
    target_pnl: Optional[float] = Field(None, description="ç›®æ¨™ PnLï¼ˆUSDTï¼‰ï¼Œç•¶é”åˆ°æ­¤å€¼æ™‚é–‹å§‹è¿½è¹¤")
    lock_ratio: Optional[float] = Field(None, description="Lock ratioï¼ˆ0~1ï¼‰ï¼Œå¦‚æœ None å‰‡ä½¿ç”¨å…¨å±€ lock_ratio")
    max_pnl_reached: Optional[float] = Field(None, description="å·²é”åˆ°çš„æœ€å¤§ PnLï¼ˆåªè®€ï¼‰")


class PortfolioTrailingConfigUpdate(BaseModel):
    """æ›´æ–° Portfolio Trailing Stop è¨­å®šçš„è«‹æ±‚æ ¼å¼"""
    enabled: Optional[bool] = None
    target_pnl: Optional[float] = None
    lock_ratio: Optional[float] = None


@app.patch("/binance/positions/stop-config", response_model=dict)
async def update_binance_position_stop_config(
    update: BinancePositionStopConfigUpdate,
    user: dict = Depends(require_admin_user)
):
    """
    æ›´æ–° Binance Live Position çš„åœæé…ç½®è¦†å¯«å€¼
    
    å…è¨±ç‚ºæ¯å€‹ Binance Live Position è¨­å®šç¨ç«‹çš„åœæé…ç½®ï¼Œè¦†å¯«å…¨å±€é…ç½®ã€‚
    é€™äº›è¦†å¯«å€¼å­˜å„²åœ¨è¨˜æ†¶é«”ä¸­ï¼Œæ‡‰ç”¨é‡å•Ÿå¾Œæœƒé‡ç½®ã€‚
    
    Args:
        update: æ›´æ–°è«‹æ±‚ï¼ˆåŒ…å« symbol, position_side å’Œè¦†å¯«å€¼ï¼‰
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Š
    
    Returns:
        dict: æ›´æ–°çµæœ
    
    Raises:
        HTTPException: ç•¶åƒæ•¸ç„¡æ•ˆæ™‚
    """
    if update.position_side.upper() not in ["LONG", "SHORT"]:
        raise HTTPException(status_code=400, detail="position_side å¿…é ˆæ˜¯ LONG æˆ– SHORT")
    
    override_key = f"{update.symbol.upper()}|{update.position_side.upper()}"
    
    if update.clear_overrides:
        # æ¸…é™¤è¦†å¯«å€¼
        if override_key in _binance_position_stop_overrides:
            del _binance_position_stop_overrides[override_key]
        logger.info(f"å·²æ¸…é™¤ Binance Live Position {override_key} çš„åœæé…ç½®è¦†å¯«")
    else:
        # æ›´æ–°è¦†å¯«å€¼
        overrides = _binance_position_stop_overrides.get(override_key, {})
        
        if update.dyn_profit_threshold_pct is not None:
            if update.dyn_profit_threshold_pct < 0:
                logger.warning(
                    f"Binance Live Position {override_key} dyn_profit_threshold_pct < 0 ({update.dyn_profit_threshold_pct})ï¼Œå·²è¨­ç‚º 0"
                )
                overrides["dyn_profit_threshold_pct"] = 0.0
            else:
                overrides["dyn_profit_threshold_pct"] = update.dyn_profit_threshold_pct
        
        if update.base_stop_loss_pct is not None:
            if update.base_stop_loss_pct < 0:
                logger.warning(
                    f"Binance Live Position {override_key} base_stop_loss_pct < 0 ({update.base_stop_loss_pct})ï¼Œå·²è¨­ç‚º 0"
                )
                overrides["base_stop_loss_pct"] = 0.0
            else:
                overrides["base_stop_loss_pct"] = update.base_stop_loss_pct
        
        if update.trail_callback is not None:
            # clamp to [0, 1] and handle 0 meaning "base stop only"
            val = update.trail_callback
            if val < 0:
                logger.warning(
                    f"Binance Live Position {override_key} trail_callback < 0 ({val})ï¼Œå·²è¨­ç‚º 0 (base-stop only)"
                )
                val = 0.0
            elif val > 1:
                logger.warning(
                    f"Binance Live Position {override_key} trail_callback > 1 ({val})ï¼Œå·²èª¿æ•´ç‚º 1.0"
                )
                val = 1.0
            overrides["trail_callback"] = val
        
        _binance_position_stop_overrides[override_key] = overrides
        
        logger.info(
            f"Binance Live Position {override_key} åœæé…ç½®å·²æ›´æ–°ï¼š"
            f"dyn_profit_threshold_pct={overrides.get('dyn_profit_threshold_pct')}, "
            f"base_stop_loss_pct={overrides.get('base_stop_loss_pct')}, "
            f"trail_callback={overrides.get('trail_callback')}"
        )
    
    # è¨ˆç®—å¯¦éš›ä½¿ç”¨çš„å€¼å’Œä¾†æºæ¨™è¨˜ï¼ˆèˆ‡ get_binance_open_positions ä¸­çš„é‚è¼¯ä¸€è‡´ï¼‰
    overrides = _binance_position_stop_overrides.get(override_key, {})
    
    # æŸ¥æ‰¾å°æ‡‰çš„æœ¬åœ° Positionï¼ˆå¦‚æœå­˜åœ¨ï¼‰
    from models import Position
    db = next(get_db())
    try:
        local_pos = (
            db.query(Position)
            .filter(
                Position.symbol == update.symbol.upper(),
                Position.side == update.position_side.upper(),
                Position.status == "OPEN",
            )
            .order_by(Position.id.desc())
            .first()
        )
    except Exception:
        local_pos = None
    finally:
        db.close()
    
    # å–å¾—å°æ‡‰æ–¹å‘çš„å…¨å±€è¨­å®š
    side_config = TRAILING_CONFIG.get_config_for_side(update.position_side.upper())
    
    # Profit Threshold
    profit_threshold_value = None
    profit_threshold_source = None
    if overrides.get("dyn_profit_threshold_pct") is not None:
        profit_threshold_value = overrides["dyn_profit_threshold_pct"]
        profit_threshold_source = "override"
    elif local_pos and local_pos.dyn_profit_threshold_pct is not None:
        profit_threshold_value = local_pos.dyn_profit_threshold_pct
        profit_threshold_source = "override"
    elif side_config.profit_threshold_pct is not None:
        profit_threshold_value = side_config.profit_threshold_pct
        profit_threshold_source = "global"
    else:
        profit_threshold_value = DYN_PROFIT_THRESHOLD_PCT
        profit_threshold_source = "default"
    
    # Lock Ratio
    lock_ratio_value = None
    lock_ratio_source = None
    if overrides.get("trail_callback") is not None:
        lock_ratio_value = overrides["trail_callback"]
        lock_ratio_source = "override"
    elif local_pos and local_pos.trail_callback is not None:
        lock_ratio_value = local_pos.trail_callback
        lock_ratio_source = "override"
    elif side_config.lock_ratio is not None:
        lock_ratio_value = side_config.lock_ratio
        lock_ratio_source = "global"
    else:
        lock_ratio_value = DYN_LOCK_RATIO_DEFAULT
        lock_ratio_source = "default"
    
    # Base SL%
    base_sl_value = None
    base_sl_source = None
    if overrides.get("base_stop_loss_pct") is not None:
        base_sl_value = overrides["base_stop_loss_pct"]
        base_sl_source = "override"
    elif local_pos and local_pos.base_stop_loss_pct is not None:
        base_sl_value = local_pos.base_stop_loss_pct
        base_sl_source = "override"
    elif side_config.base_sl_pct is not None:
        base_sl_value = side_config.base_sl_pct
        base_sl_source = "global"
    else:
        base_sl_value = DYN_BASE_SL_PCT
        base_sl_source = "default"
    
    return {
        "success": True,
        "symbol": update.symbol.upper(),
        "position_side": update.position_side.upper(),
        "overrides": overrides,
        # æ·»åŠ å¯¦éš›ä½¿ç”¨çš„å€¼å’Œä¾†æºæ¨™è¨˜ï¼ˆç”¨æ–¼å‰ç«¯é¡¯ç¤ºå’Œé¡è‰²æ¨™è¨˜ï¼‰
        "profit_threshold_value": profit_threshold_value,
        "profit_threshold_source": profit_threshold_source,
        "lock_ratio_value": lock_ratio_value,
        "lock_ratio_source": lock_ratio_source,
        "base_sl_value": base_sl_value,
        "base_sl_source": base_sl_source,
    }


@app.get("/binance/portfolio/summary")
async def get_binance_portfolio_summary(
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    å–å¾— Binance Live Positions çš„ Portfolio æ‘˜è¦ï¼ˆåˆ†åˆ¥è¿”å› LONG å’Œ SHORTï¼‰ã€‚
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Returns:
        dict: {
            "long": {
                "total_unrealized_pnl": float,  # LONG ç¸½æœªå¯¦ç¾ç›ˆè™§ï¼ˆUSDTï¼‰
                "position_count": int,          # LONG å€‰ä½æ•¸é‡
                "portfolio_trailing": {         # LONG Portfolio trailing ç‹€æ…‹
                    "enabled": bool,
                    "target_pnl": float | None,
                    "lock_ratio": float | None,
                    "max_pnl_reached": float | None
                }
            },
            "short": {
                "total_unrealized_pnl": float,  # SHORT ç¸½æœªå¯¦ç¾ç›ˆè™§ï¼ˆUSDTï¼‰
                "position_count": int,          # SHORT å€‰ä½æ•¸é‡
                "portfolio_trailing": {         # SHORT Portfolio trailing ç‹€æ…‹
                    "enabled": bool,
                    "target_pnl": float | None,
                    "lock_ratio": float | None,
                    "max_pnl_reached": float | None
                }
            }
        }
    
    Raises:
        HTTPException: ç•¶ Binance API å‘¼å«å¤±æ•—æ™‚
    """
    global _portfolio_trailing_runtime_state
    
    # å¾è³‡æ–™åº«è¼‰å…¥æŒä¹…åŒ–é…ç½®ï¼ˆid=1 for LONG, id=2 for SHORTï¼‰
    try:
        long_config = db.query(PortfolioTrailingConfig).filter(PortfolioTrailingConfig.id == 1).first()
        short_config = db.query(PortfolioTrailingConfig).filter(PortfolioTrailingConfig.id == 2).first()
        
        # å¦‚æœä¸å­˜åœ¨ï¼Œå‰µå»ºé è¨­é…ç½®
        if not long_config:
            try:
                long_config = PortfolioTrailingConfig(id=1, enabled=False, target_pnl=None, lock_ratio=None)
                db.add(long_config)
                db.commit()
                db.refresh(long_config)
            except Exception:
                db.rollback()
                long_config = None
        
        if not short_config:
            try:
                short_config = PortfolioTrailingConfig(id=2, enabled=False, target_pnl=None, lock_ratio=None)
                db.add(short_config)
                db.commit()
                db.refresh(short_config)
            except Exception:
                db.rollback()
                short_config = None
                
    except Exception as db_error:
        logger.error(f"æŸ¥è©¢ Portfolio Trailing Config å¤±æ•—: {db_error}", exc_info=True)
        long_config = None
        short_config = None
    
    try:
        client = get_client()
        raw_positions = client.futures_position_information()
        
        # åˆ†åˆ¥è¨ˆç®— LONG å’Œ SHORT çš„ PnL å’Œæ•¸é‡
        long_total_pnl = 0.0
        long_position_count = 0
        short_total_pnl = 0.0
        short_position_count = 0
        
        for item in raw_positions:
            try:
                position_amt = float(item.get("positionAmt", "0") or 0)
            except (ValueError, TypeError):
                position_amt = 0.0
            
            if position_amt == 0:
                continue
            
            try:
                unrealized_pnl = float(item.get("unRealizedProfit", "0") or 0)
                if position_amt > 0:
                    # LONG position
                    long_total_pnl += unrealized_pnl
                    long_position_count += 1
                else:
                    # SHORT position
                    short_total_pnl += unrealized_pnl
                    short_position_count += 1
            except (ValueError, TypeError):
                continue
        
        # è™•ç† LONG é…ç½®
        long_enabled = long_config.enabled if long_config else False
        long_target_pnl = long_config.target_pnl if long_config else None
        long_lock_ratio = long_config.lock_ratio if long_config else None
        long_max_pnl_reached = _portfolio_trailing_runtime_state.get("long", {}).get("max_pnl_reached")
        
        # ä½¿ç”¨å…¨å±€ lock_ratio å¦‚æœ portfolio lock_ratio ç‚º None
        long_effective_lock_ratio = long_lock_ratio
        if long_effective_lock_ratio is None:
            long_side_config = TRAILING_CONFIG.get_config_for_side("LONG")
            long_effective_lock_ratio = long_side_config.lock_ratio if long_side_config.lock_ratio is not None else DYN_LOCK_RATIO_DEFAULT
        
        # è™•ç† SHORT é…ç½®
        short_enabled = short_config.enabled if short_config else False
        short_target_pnl = short_config.target_pnl if short_config else None
        short_lock_ratio = short_config.lock_ratio if short_config else None
        short_max_pnl_reached = _portfolio_trailing_runtime_state.get("short", {}).get("max_pnl_reached")
        
        # ä½¿ç”¨å…¨å±€ lock_ratio å¦‚æœ portfolio lock_ratio ç‚º None
        short_effective_lock_ratio = short_lock_ratio
        if short_effective_lock_ratio is None:
            short_side_config = TRAILING_CONFIG.get_config_for_side("SHORT")
            short_effective_lock_ratio = short_side_config.lock_ratio if short_side_config.lock_ratio is not None else DYN_LOCK_RATIO_DEFAULT
        
        return {
            "long": {
                "total_unrealized_pnl": long_total_pnl,
                "position_count": long_position_count,
                "portfolio_trailing": {
                    "enabled": long_enabled,
                    "target_pnl": long_target_pnl,
                    "lock_ratio": long_lock_ratio,
                    "max_pnl_reached": long_max_pnl_reached,
                    "effective_lock_ratio": long_effective_lock_ratio
                }
            },
            "short": {
                "total_unrealized_pnl": short_total_pnl,
                "position_count": short_position_count,
                "portfolio_trailing": {
                    "enabled": short_enabled,
                    "target_pnl": short_target_pnl,
                    "lock_ratio": short_lock_ratio,
                    "max_pnl_reached": short_max_pnl_reached,
                    "effective_lock_ratio": short_effective_lock_ratio
                }
            }
        }
    except Exception as e:
        logger.exception("å–å¾— Portfolio Summary å¤±æ•—")
        raise HTTPException(
            status_code=500,
            detail=f"å–å¾— Portfolio Summary å¤±æ•—: {str(e)}"
        )


@app.post("/binance/positions/close-all")
async def close_all_binance_positions(
    side: Optional[str] = Query(None, description="Position side to close: 'long', 'short', or None for all"),
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    é—œé–‰ Binance Live Positionsï¼ˆå¯é¸æŒ‡å®š LONG æˆ– SHORTï¼‰ã€‚
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        side: Position side ('long', 'short'), å¦‚æœç‚º None å‰‡é—œé–‰æ‰€æœ‰å€‰ä½
    
    Returns:
        dict: {
            "success": bool,
            "closed_count": int,
            "errors": List[str]
        }
    
    Raises:
        HTTPException: ç•¶ Binance API å‘¼å«å¤±æ•—æ™‚
    """
    # é©—è­‰ side åƒæ•¸
    target_side = None
    if side:
        side_lower = side.lower()
        if side_lower not in ["long", "short"]:
            raise HTTPException(status_code=400, detail="side å¿…é ˆæ˜¯ 'long' æˆ– 'short'")
        target_side = side_lower.upper()  # "LONG" or "SHORT"
    
    try:
        client = get_client()
        raw_positions = client.futures_position_information()
        
        closed_count = 0
        errors = []
        db_positions_to_update = []  # æ”¶é›†éœ€è¦æ›´æ–°çš„ Position è¨˜éŒ„
        
        for item in raw_positions:
            try:
                position_amt = float(item.get("positionAmt", "0") or 0)
            except (ValueError, TypeError):
                continue
            
            if position_amt == 0:
                continue
            
            symbol = item.get("symbol", "")
            if not symbol:
                continue
            
            position_side = "LONG" if position_amt > 0 else "SHORT"
            
            # å¦‚æœæŒ‡å®šäº† sideï¼Œåªé—œé–‰è©²é¡åˆ¥çš„å€‰ä½
            if target_side and position_side != target_side:
                continue
            
            try:
                # ä½¿ç”¨ç¾æœ‰çš„é—œå€‰é‚è¼¯
                side = "SELL" if position_side == "LONG" else "BUY"
                qty = abs(position_amt)
                
                timestamp = int(time.time() * 1000)
                side_prefix = target_side if target_side else "ALL"
                client_order_id = f"TVBOT_CLOSE_{side_prefix}_{timestamp}_{closed_count}"
                
                logger.info(f"é—œé–‰{side_prefix if target_side else 'æ‰€æœ‰'}å€‰ä½: {symbol} {position_side}ï¼Œæ•¸é‡: {qty}")
                
                order = client.futures_create_order(
                    symbol=symbol,
                    side=side,
                    type="MARKET",
                    quantity=qty,
                    reduceOnly=True,
                    newClientOrderId=client_order_id
                )
                
                # å–å¾—å¹³å€‰åƒ¹æ ¼
                exit_price = get_exit_price_from_order(order, symbol)
                
                # æ›´æ–°è³‡æ–™åº«ä¸­çš„ Position è¨˜éŒ„ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
                # æŸ¥æ‰¾æ‰€æœ‰åŒ¹é…çš„ OPEN ç‹€æ…‹ Position è¨˜éŒ„
                matching_positions = (
                    db.query(Position)
                    .filter(
                        Position.symbol == symbol.upper(),
                        Position.side == position_side,
                        Position.status == "OPEN"
                    )
                    .all()
                )
                
                # è¨˜éŒ„éœ€è¦æ›´æ–°çš„ Position è¨˜éŒ„å’Œç›¸é—œè³‡è¨Š
                for db_position in matching_positions:
                    db_positions_to_update.append({
                        "position": db_position,
                        "exit_price": exit_price,
                        "order": order
                    })
                
                closed_count += 1
                logger.info(
                    f"æˆåŠŸé—œé–‰ {symbol} {position_side}ï¼Œè¨‚å–®ID: {order.get('orderId')}ï¼Œ"
                    f"æ‰¾åˆ° {len(matching_positions)} ç­†éœ€è¦æ›´æ–°çš„è³‡æ–™åº«è¨˜éŒ„"
                )
            except Exception as e:
                error_msg = f"{symbol} {position_side}: {str(e)}"
                errors.append(error_msg)
                logger.error(f"é—œé–‰ {symbol} {position_side} å¤±æ•—: {e}")
        
        # æ‰¹é‡æ›´æ–°æ‰€æœ‰ Position è¨˜éŒ„
        updated_count = 0
        for update_info in db_positions_to_update:
            try:
                db_position = update_info["position"]
                exit_price = update_info["exit_price"]
                order = update_info["order"]
                
                db_position.status = "CLOSED"
                db_position.closed_at = datetime.now(timezone.utc)
                db_position.exit_price = exit_price
                exit_reason = f"manual_close_{target_side.lower()}" if target_side else "manual_close_all"
                db_position.exit_reason = exit_reason
                # æ›´æ–°è¨‚å–® IDï¼ˆå¦‚æœå¯ç”¨ï¼‰
                if order.get("orderId"):
                    try:
                        db_position.binance_order_id = int(order["orderId"])
                    except (ValueError, TypeError):
                        pass
                if order.get("clientOrderId"):
                    db_position.client_order_id = order["clientOrderId"]
                
                updated_count += 1
                logger.info(
                    f"å·²æ›´æ–°è³‡æ–™åº« Position è¨˜éŒ„ {db_position.id} ({db_position.symbol} {db_position.side}) ç‹€æ…‹ç‚º CLOSED"
                )
            except Exception as db_update_error:
                logger.error(
                    f"æ›´æ–°è³‡æ–™åº« Position è¨˜éŒ„ {update_info['position'].id} å¤±æ•—: {db_update_error}"
                )
                # ç¹¼çºŒè™•ç†å…¶ä»–è¨˜éŒ„ï¼Œä¸å½±éŸ¿æ•´é«”æµç¨‹
        
        # æäº¤æ‰€æœ‰è³‡æ–™åº«è®Šæ›´
        try:
            if db_positions_to_update:
                db.commit()
                logger.info(f"æˆåŠŸæ›´æ–° {updated_count} ç­† Position è¨˜éŒ„ç‹€æ…‹ç‚º CLOSED")
        except Exception as commit_error:
            logger.error(f"æäº¤è³‡æ–™åº«è®Šæ›´å¤±æ•—: {commit_error}")
            db.rollback()
            errors.append(f"è³‡æ–™åº«æ›´æ–°å¤±æ•—: {str(commit_error)}")
        
        return {
            "success": len(errors) == 0,
            "closed_count": closed_count,
            "updated_db_records": updated_count,
            "errors": errors
        }
    except Exception as e:
        logger.exception("é—œé–‰æ‰€æœ‰å€‰ä½å¤±æ•—")
        raise HTTPException(
            status_code=500,
            detail=f"é—œé–‰æ‰€æœ‰å€‰ä½å¤±æ•—: {str(e)}"
        )


@app.get("/binance/portfolio/trailing", response_model=PortfolioTrailingConfigOut)
async def get_portfolio_trailing_config(
    side: str = Query("long", description="Position side: 'long' or 'short'"),
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    å–å¾— Portfolio Trailing Stop è¨­å®šï¼ˆåˆ†åˆ¥è™•ç† LONG å’Œ SHORTï¼‰ã€‚
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        side: Position side ('long' or 'short'), é è¨­ç‚º 'long'
    
    Returns:
        PortfolioTrailingConfigOut: ç›®å‰çš„ Portfolio Trailing è¨­å®š
    """
    global _portfolio_trailing_runtime_state
    
    # é©—è­‰ side åƒæ•¸
    side_lower = side.lower()
    if side_lower not in ["long", "short"]:
        raise HTTPException(status_code=400, detail="side å¿…é ˆæ˜¯ 'long' æˆ– 'short'")
    
    # ç¢ºå®šé…ç½® IDï¼ˆ1 for LONG, 2 for SHORTï¼‰
    config_id = 1 if side_lower == "long" else 2
    
    # å¾è³‡æ–™åº«è¼‰å…¥é…ç½®
    try:
        config = db.query(PortfolioTrailingConfig).filter(PortfolioTrailingConfig.id == config_id).first()
        if not config:
            # å¦‚æœä¸å­˜åœ¨ï¼Œå‰µå»ºé è¨­é…ç½®
            try:
                config = PortfolioTrailingConfig(
                    id=config_id,
                    enabled=False,
                    target_pnl=None,
                    lock_ratio=None
                )
                db.add(config)
                db.commit()
                db.refresh(config)
            except Exception as create_error:
                db.rollback()
                logger.error(f"å‰µå»º Portfolio Trailing Config ({side_lower.upper()}) å¤±æ•—: {create_error}", exc_info=True)
                # è¿”å›é è¨­å€¼
                return PortfolioTrailingConfigOut(
                    enabled=False,
                    target_pnl=None,
                    lock_ratio=None,
                    max_pnl_reached=_portfolio_trailing_runtime_state.get(side_lower, {}).get("max_pnl_reached")
                )
        
        return PortfolioTrailingConfigOut(
            enabled=config.enabled,
            target_pnl=config.target_pnl,
            lock_ratio=config.lock_ratio,
            max_pnl_reached=_portfolio_trailing_runtime_state.get(side_lower, {}).get("max_pnl_reached")
        )
    except Exception as db_error:
        logger.error(f"æŸ¥è©¢ Portfolio Trailing Config ({side_lower.upper()}) å¤±æ•—: {db_error}", exc_info=True)
        # è¿”å›é è¨­å€¼è€Œä¸æ˜¯æ‹‹å‡ºç•°å¸¸
        return PortfolioTrailingConfigOut(
            enabled=False,
            target_pnl=None,
            lock_ratio=None,
            max_pnl_reached=_portfolio_trailing_runtime_state.get(side_lower, {}).get("max_pnl_reached")
        )


@app.post("/binance/portfolio/trailing", response_model=PortfolioTrailingConfigOut)
async def update_portfolio_trailing_config(
    payload: PortfolioTrailingConfigUpdate,
    side: str = Query("long", description="Position side: 'long' or 'short'"),
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    æ›´æ–° Portfolio Trailing Stop è¨­å®šï¼ˆåˆ†åˆ¥è™•ç† LONG å’Œ SHORTï¼‰ã€‚
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    è¨­å®šæœƒæŒä¹…åŒ–åˆ°è³‡æ–™åº«ï¼Œå³ä½¿ç³»çµ±é‡å•Ÿä¹Ÿæœƒä¿ç•™ã€‚
    
    Args:
        payload: è¦æ›´æ–°çš„è¨­å®šï¼ˆåªæ›´æ–°æä¾›çš„æ¬„ä½ï¼‰
        side: Position side ('long' or 'short'), é è¨­ç‚º 'long'
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
        db: è³‡æ–™åº« Session
    
    Returns:
        PortfolioTrailingConfigOut: æ›´æ–°å¾Œçš„ Portfolio Trailing è¨­å®š
    
    Raises:
        HTTPException: ç•¶è¨­å®šå€¼ç„¡æ•ˆæ™‚
    """
    global _portfolio_trailing_runtime_state
    
    # é©—è­‰ side åƒæ•¸
    side_lower = side.lower()
    if side_lower not in ["long", "short"]:
        raise HTTPException(status_code=400, detail="side å¿…é ˆæ˜¯ 'long' æˆ– 'short'")
    
    # ç¢ºå®šé…ç½® IDï¼ˆ1 for LONG, 2 for SHORTï¼‰
    config_id = 1 if side_lower == "long" else 2
    
    # å¾è³‡æ–™åº«è¼‰å…¥æˆ–å‰µå»ºé…ç½®
    try:
        config = db.query(PortfolioTrailingConfig).filter(PortfolioTrailingConfig.id == config_id).first()
        if not config:
            try:
                config = PortfolioTrailingConfig(id=config_id, enabled=False, target_pnl=None, lock_ratio=None)
                db.add(config)
                db.flush()
            except Exception as create_error:
                db.rollback()
                logger.error(f"å‰µå»º Portfolio Trailing Config ({side_lower.upper()}) å¤±æ•—: {create_error}", exc_info=True)
                raise HTTPException(
                    status_code=500,
                    detail=f"ç„¡æ³•å‰µå»ºé…ç½®è¨˜éŒ„: {str(create_error)}"
                )
    except HTTPException:
        raise
    except Exception as db_error:
        logger.error(f"æŸ¥è©¢ Portfolio Trailing Config ({side_lower.upper()}) å¤±æ•—: {db_error}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"è³‡æ–™åº«éŒ¯èª¤: {str(db_error)}"
        )
    
    if hasattr(payload, 'model_dump'):
        data = payload.model_dump(exclude_unset=True)
    else:
        data = payload.dict(exclude_unset=True)
    
    # ç¯„åœé˜²å‘†
    if "lock_ratio" in data and data["lock_ratio"] is not None:
        if data["lock_ratio"] < 0:
            raise HTTPException(
                status_code=400,
                detail="lock_ratio ä¸èƒ½å°æ–¼ 0"
            )
        if data["lock_ratio"] > 1:
            logger.warning(f"lock_ratio > 1ï¼ˆå€¼={data['lock_ratio']}ï¼‰ï¼Œå·²å¼·åˆ¶èª¿æ•´ç‚º 1.0")
            data["lock_ratio"] = 1.0
    
    # ç¢ºä¿ side_lower ç‹€æ…‹å­˜åœ¨ï¼ˆé˜²å‘†ï¼‰
    if side_lower not in _portfolio_trailing_runtime_state:
        _portfolio_trailing_runtime_state[side_lower] = {"max_pnl_reached": None, "last_check_time": None}
    
    # æ›´æ–°è¨­å®šä¸¦ä¿å­˜åˆ°è³‡æ–™åº«
    if "enabled" in data:
        config.enabled = data["enabled"]
        # å¦‚æœåœç”¨ï¼Œé‡ç½® max_pnl_reached
        if not data["enabled"]:
            _portfolio_trailing_runtime_state[side_lower]["max_pnl_reached"] = None
    
    if "target_pnl" in data:
        old_target_pnl = config.target_pnl
        new_target_pnl = data["target_pnl"]
        config.target_pnl = new_target_pnl  # Can be None to clear the value
        # å¦‚æœæ›´æ–° target_pnlï¼Œé‡ç½® max_pnl_reachedï¼ˆéœ€è¦é‡æ–°é”åˆ°ç›®æ¨™ï¼‰
        # Compare with explicit None check for proper null handling
        if (old_target_pnl is None) != (new_target_pnl is None) or (old_target_pnl is not None and new_target_pnl is not None and abs(old_target_pnl - new_target_pnl) > 0.0001):
            _portfolio_trailing_runtime_state[side_lower]["max_pnl_reached"] = None
            logger.info(f"Portfolio Trailing ({side_lower.upper()}) target_pnl å·²æ›´æ–°: {old_target_pnl} -> {new_target_pnl}ï¼Œå·²é‡ç½® max_pnl_reached")
    
    if "lock_ratio" in data:
        old_lock_ratio = config.lock_ratio
        new_lock_ratio = data["lock_ratio"]
        config.lock_ratio = new_lock_ratio  # Can be None to clear the value
        logger.debug(f"Portfolio Trailing ({side_lower.upper()}) lock_ratio å·²æ›´æ–°: {old_lock_ratio} -> {new_lock_ratio}")
    
    try:
        db.commit()
        db.refresh(config)
        logger.info(f"æ›´æ–° Portfolio Trailing è¨­å®š ({side_lower.upper()})ï¼ˆå·²æŒä¹…åŒ–ï¼‰: enabled={config.enabled}, "
                    f"target_pnl={config.target_pnl}, lock_ratio={config.lock_ratio}")
    except Exception as commit_error:
        db.rollback()
        logger.error(f"ä¿å­˜ Portfolio Trailing è¨­å®š ({side_lower.upper()}) å¤±æ•—: {commit_error}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"ç„¡æ³•ä¿å­˜è¨­å®š: {str(commit_error)}"
        )
    
    return PortfolioTrailingConfigOut(
        enabled=config.enabled,
        target_pnl=config.target_pnl,
        lock_ratio=config.lock_ratio,
        max_pnl_reached=_portfolio_trailing_runtime_state.get(side_lower, {}).get("max_pnl_reached")
    )


@app.post("/binance/portfolio/trailing/reset-max-pnl")
async def reset_portfolio_max_pnl(
    side: str = Query("long", description="Position side: 'long' or 'short'"),
    user: dict = Depends(require_admin_user)
):
    """
    æ‰‹å‹•é‡ç½® Portfolio Trailing Stop çš„ Max PnL Reachedï¼ˆåˆ†åˆ¥è™•ç† LONG å’Œ SHORTï¼‰
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        side: Position side ('long' or 'short'), é è¨­ç‚º 'long'
    
    Returns:
        dict: æ“ä½œçµæœ
    """
    global _portfolio_trailing_runtime_state
    
    # é©—è­‰ side åƒæ•¸
    side_lower = side.lower()
    if side_lower not in ["long", "short"]:
        raise HTTPException(status_code=400, detail="side å¿…é ˆæ˜¯ 'long' æˆ– 'short'")
    
    # ç¢ºä¿ side_lower ç‹€æ…‹å­˜åœ¨
    if side_lower not in _portfolio_trailing_runtime_state:
        _portfolio_trailing_runtime_state[side_lower] = {"max_pnl_reached": None, "last_check_time": None}
    
    side_state = _portfolio_trailing_runtime_state.get(side_lower, {})
    old_value = side_state.get("max_pnl_reached")
    _portfolio_trailing_runtime_state[side_lower]["max_pnl_reached"] = None
    
    logger.info(f"æ‰‹å‹•é‡ç½® Max PnL Reached ({side_lower.upper()}): {old_value} -> None")
    
    return {
        "success": True,
        "message": f"Max PnL Reached ({side_lower.upper()}) å·²é‡ç½®ï¼ˆåŸå€¼: {old_value}ï¼‰"
    }


@app.delete("/positions/{pos_id}", response_model=dict)
async def delete_position_record(
    pos_id: int,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    åˆªé™¤æŒ‡å®šçš„å€‰ä½è¨˜éŒ„ã€‚
    
    æ³¨æ„ï¼šæ­¤æ“ä½œåƒ…æœƒåˆªé™¤æœ¬åœ°è³‡æ–™åº«ä¸­çš„å€‰ä½è¨˜éŒ„ï¼Œä¸æœƒå° Binance ä¸Šçš„å¯¦éš›å€‰ä½é€²è¡Œä»»ä½•æ“ä½œã€‚
    å¦‚æœå€‰ä½ä»ç‚º OPEN ç‹€æ…‹ï¼Œè«‹å‹™å¿…ç¢ºèªå¯¦éš›å€‰ä½å·²é—œé–‰ï¼Œå¦å‰‡åˆªé™¤ç´€éŒ„å¾Œå°‡ç„¡æ³•è‡ªå‹•è¿½è¹¤è©²å€‰ä½ã€‚
    """
    position = db.query(Position).filter(Position.id == pos_id).first()
    if not position:
        raise HTTPException(status_code=404, detail="Position not found")
    
    symbol = position.symbol
    status = position.status
    db.delete(position)
    db.commit()
    
    logger.info(f"åˆªé™¤å€‰ä½è¨˜éŒ„: id={pos_id}, symbol={symbol}, status={status}")
    
    return {
        "success": True,
        "message": f"Position {pos_id} ({symbol}) å·²åˆªé™¤",
        "position_id": pos_id,
        "status": status,
    }


@app.post("/positions/{pos_id}/close", response_model=dict)
async def close_position(
    pos_id: int,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    é—œé–‰å€‰ä½
    
    å¾è³‡æ–™åº«å–å‡º Positionï¼Œå¦‚æœæ˜¯ OPEN ç‹€æ…‹ï¼Œå‘¼å« close_futures_position é—œå€‰ï¼Œ
    ä¸¦æ›´æ–° status ç‚º CLOSEDï¼Œè¨­å®š closed_at ç‚ºç¾åœ¨æ™‚é–“ã€‚
    
    æ­¤ç«¯é»åƒ…é™å·²ç™»å…¥ä¸”é€šéç®¡ç†å“¡é©—è­‰çš„ä½¿ç”¨è€…ï¼ˆGoogle OAuth + ADMIN_GOOGLE_EMAILï¼‰ä½¿ç”¨ã€‚
    
    Args:
        pos_id: å€‰ä½ ID
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
        db: è³‡æ–™åº« Session
    
    Returns:
        dict: é—œå€‰çµæœ
    """
    # å¾è³‡æ–™åº«å–å‡º Position
    position = db.query(Position).filter(Position.id == pos_id).first()
    
    if not position:
        raise HTTPException(status_code=404, detail="æ‰¾ä¸åˆ°æŒ‡å®šçš„å€‰ä½è¨˜éŒ„")
    
    if position.status != "OPEN":
        raise HTTPException(
            status_code=400,
            detail=f"å€‰ä½ç‹€æ…‹ç‚º {position.status}ï¼Œç„¡æ³•é—œé–‰ã€‚åªæœ‰ OPEN ç‹€æ…‹çš„å€‰ä½å¯ä»¥é—œé–‰ã€‚"
        )
    
    try:
        # å‘¼å«å¹£å®‰ API é—œå€‰
        close_order = close_futures_position(
            symbol=position.symbol,
            position_side=position.side,  # LONG æˆ– SHORT
            qty=position.qty,
            position_id=position.id
        )
        
        # å–å¾—å¹³å€‰åƒ¹æ ¼
        exit_price = get_exit_price_from_order(close_order, position.symbol)
        
        # æ›´æ–° Position è¨˜éŒ„èˆ‡å¹³å€‰è³‡è¨Š
        position.status = "CLOSED"
        position.closed_at = datetime.now(timezone.utc)
        position.exit_price = exit_price
        position.exit_reason = "manual_close"
        db.commit()
        
        return {
            "success": True,
            "message": "å€‰ä½å·²æˆåŠŸé—œé–‰",
            "position_id": position.id,
            "closed_at": position.closed_at.isoformat(),
            "exit_price": exit_price,
            "exit_reason": "manual_close",
            "binance_order": close_order
        }
    
    except Exception as e:
        # å¦‚æœé—œå€‰å¤±æ•—ï¼Œæ›´æ–°ç‹€æ…‹ç‚º ERROR
        position.status = "ERROR"
        db.commit()
        raise HTTPException(status_code=500, detail=f"é—œå€‰å¤±æ•—: {str(e)}")


@app.post("/positions/{pos_id}/trailing", response_model=PositionOut)
async def update_trailing(
    pos_id: int,
    trailing_update: TrailingUpdate,
    user: dict = Depends(require_admin_user),
    db: Session = Depends(get_db)
):
    """
    æ›´æ–°è¿½è¹¤åœæè¨­å®š
    
    å°‡ trailing_callback_percent è½‰æ›ç‚º ratioï¼ˆä¾‹å¦‚ 2.0 -> 0.02ï¼‰ï¼Œ
    å¦‚æœ highest_price ç‚ºç©ºï¼Œåˆå§‹åŒ–ç‚ºç›®å‰çš„æ¨™è¨˜åƒ¹æ ¼ã€‚
    
    æ­¤ç«¯é»åƒ…é™å·²ç™»å…¥ä¸”é€šéç®¡ç†å“¡é©—è­‰çš„ä½¿ç”¨è€…ï¼ˆGoogle OAuth + ADMIN_GOOGLE_EMAILï¼‰ä½¿ç”¨ã€‚
    
    Args:
        pos_id: å€‰ä½ ID
        trailing_update: è¿½è¹¤åœææ›´æ–°è³‡æ–™
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
        db: è³‡æ–™åº« Session
    
    Returns:
        PositionOut: æ›´æ–°å¾Œçš„å€‰ä½è³‡è¨Š
    """
    # å¾è³‡æ–™åº«å–å‡º Position
    position = db.query(Position).filter(Position.id == pos_id).first()
    
    if not position:
        raise HTTPException(status_code=404, detail="æ‰¾ä¸åˆ°æŒ‡å®šçš„å€‰ä½è¨˜éŒ„")
    
    if position.status != "OPEN":
        raise HTTPException(
            status_code=400,
            detail=f"åªèƒ½æ›´æ–° OPEN ç‹€æ…‹å€‰ä½çš„è¿½è¹¤åœæè¨­å®šï¼Œç›®å‰ç‹€æ…‹ç‚º {position.status}"
        )
    
    try:
        pct = trailing_update.trailing_callback_percent
        
        if pct < 0 or pct > 100:
            logger.warning(
                f"æ›´æ–°å€‰ä½ {position.id} ({position.symbol}) çš„ trailing_callback_percent éæ³•: {pct}ï¼Œå¿…é ˆåœ¨ 0~100 ä¹‹é–“"
            )
            raise HTTPException(
                status_code=400,
                detail="trailing_callback_percent å¿…é ˆåœ¨ 0~100 ä¹‹é–“"
            )
        
        if pct == 0:
            logger.info(
                f"å€‰ä½ {position.id} ({position.symbol}) æ›´æ–° trailing_callback_percent=0ï¼Œåƒ…ä½¿ç”¨ base stop-loss"
            )
            position.trail_callback = 0.0
        else:
            trail_callback_ratio = pct / 100.0
            position.trail_callback = trail_callback_ratio
            logger.info(
                f"å€‰ä½ {position.id} ({position.symbol}) æ›´æ–° trailing_callback_percent={pct}%ï¼Œlock_ratio={trail_callback_ratio}"
            )
        
        # å¦‚æœ highest_price ç‚ºç©ºï¼Œåˆå§‹åŒ–ç‚ºç›®å‰çš„æ¨™è¨˜åƒ¹æ ¼
        if position.highest_price is None:
            mark_price = get_mark_price(position.symbol)
            position.highest_price = mark_price
        
        db.commit()
        db.refresh(position)
        
        return PositionOut(**position.to_dict())
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"æ›´æ–°è¿½è¹¤åœæè¨­å®šå¤±æ•—: {str(e)}")


@app.get("/health")
async def health_check():
    """å¥åº·æª¢æŸ¥ç«¯é»"""
    try:
        # æª¢æŸ¥å¹£å®‰é€£ç·š
        client = get_client()
        # ç°¡å–®æ¸¬è©¦ï¼šå–å¾— BTCUSDT æ¨™è¨˜åƒ¹æ ¼
        mark_price = get_mark_price("BTCUSDT")
        
        return {
            "status": "healthy",
            "binance_connected": True,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy",
                "binance_connected": False,
                "error": str(e),
                "timestamp": datetime.now(timezone.utc).isoformat()
            }
        )


# ==================== Admin Prune Endpoints ====================

@app.delete("/admin/signal-logs/prune")
def prune_signal_logs(
    days: int = Query(30, ge=1, le=365, description="ä¿ç•™æœ€è¿‘å¹¾å¤©çš„ signal logs"),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user),
):
    """
    æ¸…ç†èˆŠçš„ TradingView Signal Logs
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        days: ä¿ç•™æœ€è¿‘å¹¾å¤©çš„ logsï¼ˆé è¨­ 30 å¤©ï¼‰
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: æ¸…ç†çµæœï¼ŒåŒ…å«åˆªé™¤çš„ç­†æ•¸å’Œæˆªæ­¢æ™‚é–“
    """
    cutoff = datetime.utcnow() - timedelta(days=days)
    stmt = delete(TradingViewSignalLog).where(TradingViewSignalLog.received_at < cutoff)
    result = db.execute(stmt)
    db.commit()
    
    deleted_count = getattr(result, "rowcount", None)
    logger.info(f"Prune signal logs: åˆªé™¤ {deleted_count} ç­†ï¼ˆä¿ç•™æœ€è¿‘ {days} å¤©ï¼Œæˆªæ­¢æ™‚é–“: {cutoff.isoformat()}ï¼‰")
    
    return {
        "success": True,
        "deleted": deleted_count,
        "cutoff": cutoff.isoformat(),
        "days": days,
    }


@app.delete("/admin/signal-logs/clear")
def clear_all_signal_logs(
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user),
):
    """
    æ¸…é™¤æ‰€æœ‰ TradingView Signal Logs
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    æ­¤æ“ä½œæœƒåˆªé™¤æ‰€æœ‰ signal logsï¼Œç„¡æ³•å¾©åŸï¼Œè«‹è¬¹æ…ä½¿ç”¨ã€‚
    
    Args:
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: æ¸…ç†çµæœï¼ŒåŒ…å«åˆªé™¤çš„ç­†æ•¸
    """
    # å…ˆè¨ˆç®—ç¸½æ•¸
    total_count = db.query(TradingViewSignalLog).count()
    
    # åˆªé™¤æ‰€æœ‰è¨˜éŒ„
    stmt = delete(TradingViewSignalLog)
    result = db.execute(stmt)
    db.commit()
    
    deleted_count = getattr(result, "rowcount", None) or total_count
    logger.info(f"Clear all signal logs: åˆªé™¤ {deleted_count} ç­†")
    
    return {
        "success": True,
        "deleted": deleted_count,
        "message": f"æˆåŠŸæ¸…é™¤ {deleted_count} ç­† signal logs"
    }


@app.delete("/admin/positions/prune-closed")
def prune_closed_positions(
    days: int = Query(30, ge=1, le=365, description="ä¿ç•™æœ€è¿‘å¹¾å¤©å…§é—œé–‰çš„å€‰ä½"),
    include_error: bool = Query(False, description="æ˜¯å¦åŒæ™‚åˆªé™¤ ERROR ç‹€æ…‹çš„å€‰ä½"),
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user),
):
    """
    æ¸…ç†èˆŠçš„å·²é—œé–‰å€‰ä½è¨˜éŒ„
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    
    Args:
        days: ä¿ç•™æœ€è¿‘å¹¾å¤©å…§é—œé–‰çš„å€‰ä½ï¼ˆé è¨­ 30 å¤©ï¼‰
        include_error: æ˜¯å¦åŒæ™‚åˆªé™¤ ERROR ç‹€æ…‹çš„å€‰ä½ï¼ˆé è¨­ Falseï¼‰
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: æ¸…ç†çµæœï¼ŒåŒ…å«åˆªé™¤çš„ç­†æ•¸å’Œæˆªæ­¢æ™‚é–“
    """
    cutoff = datetime.utcnow() - timedelta(days=days)
    
    # å»ºç«‹åˆªé™¤æ¢ä»¶
    if include_error:
        # å¦‚æœ include_error=Trueï¼Œåˆªé™¤ CLOSED ä¸”è¶…éæŒ‡å®šå¤©æ•¸çš„ï¼Œä»¥åŠæ‰€æœ‰ ERROR ç‹€æ…‹çš„
        stmt = delete(Position).where(
            or_(
                # CLOSED ä¸”è¶…éæŒ‡å®šå¤©æ•¸
                and_(
                    Position.status == "CLOSED",
                    Position.closed_at.isnot(None),
                    Position.closed_at < cutoff
                ),
                # æˆ– ERROR ç‹€æ…‹ï¼ˆä¸é™æ—¥æœŸï¼‰
                Position.status == "ERROR"
            )
        )
    else:
        # åªåˆªé™¤ CLOSED ä¸”è¶…éæŒ‡å®šå¤©æ•¸çš„
        stmt = delete(Position).where(
            and_(
                Position.status == "CLOSED",
                Position.closed_at.isnot(None),
                Position.closed_at < cutoff
            )
        )
    result = db.execute(stmt)
    db.commit()
    
    deleted_count = getattr(result, "rowcount", None)
    error_info = "ï¼ˆåŒ…å« ERROR ç‹€æ…‹ï¼‰" if include_error else ""
    logger.info(f"Prune closed positions: åˆªé™¤ {deleted_count} ç­†{error_info}ï¼ˆä¿ç•™æœ€è¿‘ {days} å¤©å…§é—œé–‰çš„å€‰ä½ï¼Œæˆªæ­¢æ™‚é–“: {cutoff.isoformat()}ï¼‰")
    
    return {
        "success": True,
        "deleted": deleted_count,
        "cutoff": cutoff.isoformat(),
        "days": days,
        "include_error": include_error,
    }


@app.delete("/admin/positions/clear-error")
def clear_error_positions(
    db: Session = Depends(get_db),
    user: dict = Depends(require_admin_user),
):
    """
    æ¸…é™¤æ‰€æœ‰ ERROR ç‹€æ…‹çš„å€‰ä½è¨˜éŒ„
    
    åƒ…é™å·²ç™»å…¥çš„ç®¡ç†å“¡ä½¿ç”¨ã€‚
    æ­¤æ“ä½œæœƒåˆªé™¤æ‰€æœ‰ ERROR ç‹€æ…‹çš„å€‰ä½ï¼Œç„¡æ³•å¾©åŸï¼Œè«‹è¬¹æ…ä½¿ç”¨ã€‚
    
    Args:
        db: è³‡æ–™åº« Session
        user: ç®¡ç†å“¡ä½¿ç”¨è€…è³‡è¨Šï¼ˆç”± Depends(require_admin_user) è‡ªå‹•é©—è­‰ï¼‰
    
    Returns:
        dict: æ¸…ç†çµæœï¼ŒåŒ…å«åˆªé™¤çš„ç­†æ•¸
    """
    # å…ˆè¨ˆç®— ERROR ç‹€æ…‹çš„ç¸½æ•¸
    error_count = db.query(Position).filter(Position.status == "ERROR").count()
    
    # åˆªé™¤æ‰€æœ‰ ERROR ç‹€æ…‹çš„è¨˜éŒ„
    stmt = delete(Position).where(Position.status == "ERROR")
    result = db.execute(stmt)
    db.commit()
    
    deleted_count = getattr(result, "rowcount", None) or error_count
    logger.info(f"Clear error positions: åˆªé™¤ {deleted_count} ç­† ERROR ç‹€æ…‹çš„å€‰ä½")
    
    return {
        "success": True,
        "deleted": deleted_count,
        "message": f"æˆåŠŸæ¸…é™¤ {deleted_count} ç­† ERROR ç‹€æ…‹çš„å€‰ä½"
    }


# ==================== ä¸»ç¨‹å¼å…¥å£ ====================

if __name__ == "__main__":
    import uvicorn
    
    # é–‹ç™¼æ¨¡å¼ä¸‹å•Ÿå‹•ä¼ºæœå™¨
    # é è¨­ç›£è½ 0.0.0.0:8000
    # æ­£å¼ç’°å¢ƒå»ºè­°ä½¿ç”¨ç”Ÿç”¢ç´š ASGI ä¼ºæœå™¨ï¼ˆå¦‚ gunicorn + uvicorn workersï¼‰
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=True  # é–‹ç™¼æ¨¡å¼ï¼šç¨‹å¼ç¢¼è®Šæ›´æ™‚è‡ªå‹•é‡è¼‰
    )
